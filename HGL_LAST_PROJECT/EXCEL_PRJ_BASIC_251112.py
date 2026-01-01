"""
-일자:2025.11.12
-내용:엑셀 다루기 기초 다루기
-한계
 : openpyxl은 xlsx만 읽는다.
 : 시트의 컬럼과 로우의 끝을 알기가 애매하다.
-결론
 : 판다스를 이용하자
"""

import os
os.environ['KMP_DUPLICATE_LIB_OK']='True'
os.chdir(r"c:\projects\PY_MLE\HGL_LAST_PROJECT")


import openpyxl as op
import datetime

filename = r"./lastsample/EXCEL_PRJ_TEST01.xlsx"


wb = op.load_workbook(filename)
ws = wb['Sheet1']
vaList = []

forcontinue = True

# 엑셀에 있는 값 불러오기
for row in range(1, ws.max_row + 1):
    imList=[]

    # 첫번째 컬럼의 값이 없으면 빠져나온다.
    if not forcontinue:
        break

    for col in range(1, ws.max_column + 1):

        val=ws.cell(row=row, column=col).value

        # 어떤 엑셀 파일의 끝에 뭔가 한점이 있으면 거기까지 읽는다.
        # 이걸 방지하기 위해 첫번째 컬럼의 값이(id라고 해도 됨) 스톱한다.
        if col==1 and not val:
            forcontinue=False
            break

        # 날짜 타입은 년월일만 나오게 함
        if type(val)==datetime.datetime:
            val = val.strftime('%Y-%m-%d')

        imList.append(val)
    if len(imList) > 0:
        vaList.append(imList)

# 값 입력
maxrow = ws.max_row + 1
ws.cell(row=maxrow, column=1).value = ws.cell(row=maxrow-1, column=1).value + 1
ws.cell(row=maxrow, column=2).value = "hhh"
ws.cell(row=maxrow, column=3).value = ws.cell(row=maxrow-1, column=3).value + 1000
ws.cell(row=maxrow, column=4).value = datetime.datetime.now().strftime("%Y-%m-%d")

for vl in vaList:
    print(vl)

wb.save(filename)
wb.close()
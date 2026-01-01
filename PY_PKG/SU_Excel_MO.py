"""
-일자:24.12.10
"""
import openpyxl as op
import csv


###################################################### 리스트 형식으로 리턴
def Excel_To_List(sourceFile):
    """
    엑셀파일을 리스트로 리턴, 리스트 안의 리스트로 안쪽 리스트는 엑셀의 행의 값
    :param filename: 엑셀파일명
    :return: 엑셀파일의 내용을 담은 리스트, 리스트 안의 리스트는 ","로 구분
    """
    try:
        wb = op.load_workbook(sourceFile, data_only=True)  # data_only옵션을 주지 않으면 수식까지 읽어 온다.
        wlist = wb.sheetnames
        row_list_value = []

        for name in wlist:
            ws = wb[name]
            for row in ws.rows:
                iList = []
                for col in row:
                    strvalue = col.value
                    if strvalue is not None:
                        iList.append(str(strvalue))
                row_list_value.append(",".join(iList))
        return row_list_value
    except ZeroDivisionError as e:
        print(e)
        return []

###################################################### 리스트 형식을 엑셀로 저장
def List_To_Excel(filename, List):
    """
    리스트를 엑셀로 저장
    :param filename: 저장할 엑셀파일명
    :param List: 리스트
    """
    wb = op.Workbook()
    ws = wb.active

    ws.title = "Sheet1"

    # 여기는 컬럼이 하나 일때와 리스트안에 리스트 형태의 다중 컬럼일때를 구분
    # 일반적으로 하나의 컬럼을 가지는 엑셀은  없을 것으로 판단
    for item in List:
        if str(type(item)) != "<class 'list'>":
            rowList = []
            rowList.append(item)
            ws.append(rowList)
        else:   # 리스트 안의 리스트는 그냥 append해버리면 됨
            ws.append(item)
    wb.save(filename)

###################################################### 엑셀 시트 내용 삭제
def Excel_Sheet_Delete(filename, worksheet):
    """
    엑셀파일에서 특정 엑셀 시트 삭제
    :param filename: 엑셀파일명
    :param worksheet: 내용을 삭제할 엑셀시트명
    """
    wb = op.load_workbook(filename)
    ws = wb[worksheet]

    for item in range(1, ws.max_row+1):
        ws.delete_rows(1)  # 로우 수만큼 삭제

    wb.save(filename)


###################################################### csv 파일 읽어서 리스트형으로 돌려주기
def CSV_to_List(filename,delimiter=',',quotechar=' '):
    """
    csv 파일을 읽어서 리스트형으로 리턴
    encoding=utf-8로 하니 맨처음 컬럼명에 \ufeff 알수없는 문자가 나옴
    :param filename: csv 파일명
    :param delimiter:디폴트 구분자
    :param quotechar:인용
    :return: 리스트
    """
    output = []
    with open(filename, newline='',encoding='utf-8-sig') as csvfile: # utf-8로 하니 맨처음 컬럼명에 \ufeff 알수없는 문자가 나옴
        spamreader = csv.reader(csvfile, delimiter=delimiter, quotechar=quotechar)
        for row in spamreader:
            output.append(row)
    return output


########################################################################
# 오라클 조회쿼리를 엑셀로 내보내기
# wooksheetcount : 하나의 워크 시트에 담을 수 있는 개수 정의
def OracleToExcel(cursor, sql, filename, wooksheetcount):
    cursor.execute(sql)

    orgwooksheetcount = wooksheetcount

    # 워크시트 아름명에 순차적으로 붙는다.
    wscount = 1

    wb = op.Workbook()
    ws = wb.active

    # 처음 워크시트명은 "DB1"
    ws.title = 'DB' + str(wscount)

    # 컬럼을 세팅한다.
    rowList = []
    for col in cursor.description:
        rowList.append(col[0])
    ws.append(rowList)

    # 값을 세팅한다.
    rowList = []
    rowCount = 0
    for row in cursor:
        for va in row:
            rowList.append(va)

        print(str(rowCount) + " create complete")

        # 여기부터는 하나의 시트에 담을 수 있는 개수만큼 채우고 다음시트 생성
        if rowCount == wooksheetcount:
            wscount = wscount + 1
            wstitle = "DB" + str(wscount)
            wb.create_sheet(wstitle)
            ws = wb[wstitle]

            rowListimsi = []
            for col in cursor.description:
                rowListimsi.append(col[0])
            ws.append(rowListimsi)
            wooksheetcount = orgwooksheetcount * wscount

        rowCount = rowCount + 1
        ws.append(rowList)
        rowList = []

    wb.save(filename+".xlsx")

    print("excel create complete!!")
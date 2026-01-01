import openpyxl as op
import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import styles
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
# from openpyxl.styles.alignment import Alignment
import random

from pyparsing import col
from sqlalchemy import true

############################################################
# 빈 파일을 하나 생성해서 날짜를 로우마다 추가하는 예제
wb = op.Workbook()
ws = wb.active
ws.title = 'test01'
ws.sheet_properties.tabColor = "1072BA"

# 끝에 시트하나 추가
wb.create_sheet("My Createsheet")
ws2 = wb["My Createsheet"]
ws2.sheet_properties.tabColor = "2072BA"


# 오늘날짜 세팅
now = datetime.datetime.now()

# 40개 행에 순번과 날짜 세팅

ws.append(['순번','날짜'])

for i in range(1, 41):
    nowDate = now.strftime('%Y-%m-%d')  # 날짜 타입정의
    rowList = [i, nowDate]
    ws.append(rowList)  # 행추가
    now = now + datetime.timedelta(hours=24)  # 하루씩 더해감
wb.save('newExcel.xlsx')
############################################################


############################################################
# sheet 네임타이틀정보알기
wb = op.load_workbook('newExcel.xlsx',read_only=True) # true이면 읽기 전용
print(wb.sheetnames) # 한번에 리스트로 뽑기
for sheet in wb:
    print(sheet.title)

wb.close()
############################################################


############################################################
# 엑셀 정렬, 색깔 등 바꾸기
wb = op.load_workbook('newExcel.xlsx')
ws = wb.active
for row in ws.rows:
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(fgColor='00FFFF', fill_type='solid')
        cell.font = Font(color='FF00FF')
        
wb.save('newExcel.xlsx')
############################################################


############################################################
# 숫자를 컬럼명으로 리턴(1 -> A
wb = op.load_workbook('newExcel.xlsx')
ws = wb.active
for item in range(1, ws.max_column + 1):
    print(get_column_letter(item))
wb.close()
############################################################


############################################################
# 포맷을 조회
wb = op.load_workbook('newExcel.xlsx')
ws = wb.active
for item in range(1, ws.max_column + 1):
    print(ws.cell(1, item).data_type)  # n-숫자, s-날짜
wb.close()
############################################################


############################################################
# 읽기 모드로 해서 모든 셀의 값을 출력. 한줄에 컬럼별 읽고 다음줄로 이동
wb = op.load_workbook('newExcel.xlsx', read_only=True)
ws = wb.active
for row in ws.rows:
    for cell in row:
        print(cell.value)
wb.close()
############################################################


############################################################
# 현재 시트의 범위를 출력 A1:D40
print(ws.calculate_dimension())
############################################################


############################################################
# 엑셀에 하나 하나 로루 또는 컬럼을 추가할 수 있다
wb = op.load_workbook('newExcel.xlsx')
ws = wb.active
ws.insert_rows(3)  # 20번 바로전에 하나의 로우 생성
ws.insert_cols(3)  # 3번째 컬럼 생성
ws.delete_rows(5)  # 5번째 로우 삭제
ws.delete_cols(1)  # 1번째 컬럼 삭제
wb.save('newExcel.xlsx')
############################################################


############################################################
# 엑셀을 테이블로 포맷으로 받아서 작업
# 굳이 이럴필요는 없는데 이렇게 하는 방법만
wb = op.Workbook()
ws = wb.active

data = [
    ['Apples', 10000, 5000, 8000, 6000],
    ['Pears',   2000, 3000, 4000, 5000],
    ['Bananas', 6000, 6000, 6500, 6000],
    ['Oranges',  500,  300,  200,  700],
]

ws.append(["Fruit", "2011", "2012", "2013", "2014"])
for row in data:
    ws.append(row)

tab = Table(displayName="Table1", ref=ws.calculate_dimension())
style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=True)
tab.tableStyleInfo = style
ws.add_table(tab)
wb.save("table.xlsx")
############################################################

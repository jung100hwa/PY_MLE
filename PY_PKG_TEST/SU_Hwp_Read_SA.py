# 승인대상 한글 파일에서 뽑아내기

import SU_Hwp_Read_MO as phwp
import os
import re
import openpyxl as op


aValueD = {}
strImsi = ''
excelrow = 1

# 저장할 엑셀파일을 하나 생성
wb = op.Workbook()
ws = wb.active
ws.cell(row=1, column=1, value='승인번호')
ws.cell(row=1, column=2, value='상호(명칭)')
ws.cell(row=1, column=3, value='법인등록번호(사업자등록번호)')
ws.cell(row=1, column=4, value='담당자 성명 및 연락처')
ws.cell(row=1, column=5, value='전자우편')
ws.cell(row=1, column=6, value='소재지(사업장)')
ws.cell(row=1, column=7, value='전화번호')
ws.cell(row=1, column=8, value='팩스번호')
ws.cell(row=1, column=9, value='제품명')
ws.cell(row=1, column=10, value='품목')
ws.cell(row=1, column=11, value='제형')
ws.cell(row=1, column=12, value='용도')
ws.cell(row=1, column=13, value='주성분 및 함량')
ws.cell(row=1, column=14, value='효과효능')
ws.cell(row=1, column=15, value='용법용량')
ws.cell(row=1, column=16, value='사용상 주의사항')
ws.cell(row=1, column=17, value='포장단위')
ws.cell(row=1, column=18, value='저장방법 및 유통기한')
ws.cell(row=1, column=19, value='승인일자')


# 이 함수는 직인이라는 키워드 위에 있는 값들을 세팅 시키는 함수
def findnextkey(strList, strTitle, count):
    cnt = 0
    strList2 = strList[:]
    returnList = []
    strTitle = strTitle.replace(' ', '')

    for item in strList2[count:]:
        item2 = item.replace(' ', '')
        if '효능ㆍ효과' in item2:
            item2 = item2.replace('효능ㆍ효과', '효과ㆍ효능')

        if strTitle not in item2:
            if cnt == 0:
                returnList.append(item)
            else:
                returnList.append('\n' + item)
            cnt += 1
        else:
            break

    return ''.join(returnList).strip()

# 이 함수는 직인이라는 키워드 아래에 있는 함수들을 세팅 시키는 함수
# 사용상 주의, 용법용량 등은 직인 위에 별첨으로 있는 경우가 대부분이고
# 직인 아래에 첨부형식으로 상세하기 되어 있음


def findnextkey2(strList, strTitle, count):
    cnt = 0
    strList2 = strList[:]
    returnList = []

    for item in strList2[count:]:
        item2 = item.replace(' ', '')
        if '효능ㆍ효과' in item2:
            item2 = item2.replace('효능ㆍ효과', '효과ㆍ효능')

        if strTitle not in item2:
            if cnt == 0:
                returnList.append(item)
            else:
                returnList.append('\n' + item)
            cnt += 1
        else:
            break

    return ''.join(returnList).strip()


def hwpFileReadSave(hwpReader):
    hwpReader = hwpReader.bodyStream()
    for keyItem in hwpReader.keys():
        strVList = hwpReader.get(keyItem)
        strVList = strVList.splitlines()

        cnt = 0
        strList2 = strVList[:]
        if "직인" in strList2:
            cnt = strList2.index("직인")
            cnt = cnt + 1

        count = 0
        for valueItem in strVList:  # 라인단위로 해서 특정 키워드를 찾는다.
            count = count + 1
            valueItem = valueItem.strip()
            if '효능ㆍ효과' in valueItem:
                valueItem = valueItem.replace('효능ㆍ효과', '효과ㆍ효능')

            # print(valueItem)

            if count <= cnt:
                # 1.승인번호
                if valueItem == "■ 생활화학제품 및 살생물제의 안전관리에 관한 법률 시행규칙 [별지 제7호서식]":
                    strImsi = strVList[count]
                    aValueD['승인번호'] = strImsi[5:].replace(' ', '')

                # 2.상호(명칭)
                if '상호(명칭)' in valueItem:
                    strImsi = findnextkey(strVList, '법인등록번호(사업자등록번호)', count)
                    aValueD['상호(명칭)'] = strImsi

                # 3.법인등록번호(사업자등록번호)
                if '법인등록번호(사업자등록번호)' in valueItem:
                    strImsi = findnextkey(strVList, '성명(대표자)', count)
                    aValueD['법인등록번호(사업자등록번호)'] = strImsi

                # 4.담당자 성명 및 연락처
                if '담당자 성명 및 연락처' in valueItem:
                    strImsi = findnextkey(strVList, '(전자우편:', count)
                    aValueD['담당자 성명 및 연락처'] = strImsi

                # 5.전자우편
                if '(전자우편' in valueItem:
                    strImsi = findnextkey(
                        strVList, '소재지(사업장)', count-1)  # 여기는 count - 1
                    strImsi = strImsi.replace(
                        '(전자우편:', '').replace(')', '').replace(' ', '')
                    aValueD['전자우편'] = strImsi

                # 6.소재지(사업장)
                if '소재지(사업장)' in valueItem:
                    strImsi = findnextkey(strVList, '(전화번호', count)
                    aValueD['소재지(사업장)'] = strImsi

                # 7.전화번호
                if '(전화번호' in valueItem:
                    strImsi = valueItem
                    findnumloc = strImsi.replace(' ', '').find('(전화번호')
                    findnumloc2 = strImsi.replace(
                        ' ', '').find(')', findnumloc)
                    strImsi = strImsi[:findnumloc2+1].replace('(전화번호:', '')
                    aValueD['전화번호'] = strImsi

                    # 다음과 같이 하면 오류가 많이 남. 전화번호와 팩스번호가 한줄로 나오는 경우가 있음
                    # strImsi = findnextkey(strVList, '(팩스번호:', count - 1)  # 여기는 count - 1
                    # strImsi = strImsi.replace('(전화번호:', '').replace(')', '').replace(' ', '')
                    # aValueD['전화번호'] = strImsi

                # 8.팩스번호
                if '(팩스번호' in valueItem:
                    strImsi = valueItem.replace(' ', '')
                    findnumloc = strImsi.find('팩스번호')
                    findnumloc2 = strImsi.find(')', findnumloc)
                    strImsi = strImsi[findnumloc -
                                      1:findnumloc2].replace('(팩스번호:', '')
                    aValueD['팩스번호'] = strImsi

                    # strImsi = findnextkey(strVList, '제품명', count - 1)  # 여기는 count - 1
                    # strImsi = strImsi.replace('(팩스번호', '').replace(')', '').replace(' ', '')
                    # aValueD['팩스번호'] = strImsi

                # 9.제품명
                if '제품명' in valueItem:
                    strImsi = findnextkey(strVList, '품목', count)
                    aValueD['제품명'] = strImsi

                # 10.품목
                if '품목' in valueItem:
                    strImsi = findnextkey(strVList, '제형', count)
                    aValueD['품목'] = strImsi

                # 11.제형
                if '제형' in valueItem:
                    strImsi = findnextkey(strVList, '용도', count)
                    aValueD['제형'] = strImsi

                # 12.용도
                if '용도' in valueItem:
                    strImsi = findnextkey(strVList, '주성분 및 함량', count)
                    aValueD['용도'] = strImsi

                # 13.주성분 및 함량
                if '주성분 및 함량' in valueItem:
                    strImsi = findnextkey(strVList, '효과ㆍ효능', count)
                    aValueD['주성분 및 함량'] = strImsi

                # 14.효과·효능
                if '효과ㆍ효능' in valueItem:  # 효능 효과 바뀐게 있네...쯧쯧
                    strImsi = findnextkey(strVList, '용법ㆍ용량', count)
                    aValueD['효과ㆍ효능'] = strImsi

                # 15.용법용량
                if '용법ㆍ용량' in valueItem:
                    strImsi = findnextkey(strVList, '사용상 주의사항', count)
                    aValueD['용법ㆍ용량'] = strImsi

                # 16.사용상 주의사항
                if '사용상 주의사항' in valueItem:
                    strImsi = findnextkey(strVList, '포장단위', count)
                    aValueD['사용상 주의사항'] = strImsi

                # 17.포장단위
                if '포장단위' in valueItem:
                    strImsi = findnextkey(strVList, '저장방법 및 유통기한', count)
                    aValueD['포장단위'] = strImsi

                # 18.저장방법 및 유통기한
                if '저장방법 및 유통기한' in valueItem:
                    strImsi = findnextkey(
                        strVList, '「생활화학제품 및 살생물제의 안전관리에 관한 법률」', count)
                    aValueD['저장방법 및 유통기한'] = strImsi

                # 19.직인
                if '직인' in valueItem:
                    strM1 = strVList[count - 4].replace(' ', '')
                    if strM1:
                        aValueD['승인일자'] = strM1
                    else:
                        aValueD['승인일자'] = strVList[count-5].replace(' ', '')
            # else:
            #     # 효과효능이 별첨인 경우
            #     if '효과·효능' in valueItem:
            #         strImsi = findnextkey2(strVList, '용법·용량', count)
            #         aValueD['효과·효능2'] = strImsi
            #
            #     # 사용상 주의사항이 별첨인 경우
            #     if '사용상 주의사항' in valueItem:
            #         strImsi = findnextkey2(strVList, '별규 자료', count)
            #         strImsi = strImsi.replace('■ 생활화학제품 및 살생물제의 안전관리에 관한 법률 시행규칙 [별지 제7호서식] 별첨 자료 - 승인번호 제2420-0000호(3 / 6)','')
            #         strImsi = strImsi.replace('\n\n\n','\n')
            #         aValueD['사용상 주의사항2'] = strImsi


# 여기서부터 프로그램 시작
dirname = 'c:\\work\\PY_MYD_G\\IDT\\HWPDB\\'
dirname2 = 'c:\\work\\PY_MYD_G\\IDT\\HWPDB'
dirname = os.listdir(dirname)

num = 0
for fname in dirname:
    try:
        num = num + 1
        hwpReader = phwp.HwpReader(dirname2, fname)
        hwpFileReadSave(hwpReader)

        print("================================>%d : %s" % (num, fname))

        # 엑셀에 저장한다.
        excelcol = 0
        for item in list(aValueD.keys()):
            str = aValueD.get(item)
            excelcol = excelcol + 1
            ws.cell(row=num+1, column=excelcol, value=str)
            print('%s : %s' % (item, str))

    except PermissionError:
        pass

wb.save('c:\\work\\PY_MYD_G\\IDT\\승인대상.xlsx')
wb.close()

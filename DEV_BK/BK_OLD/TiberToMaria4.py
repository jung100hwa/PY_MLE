import openpyxl as op
import datetime
import os
import platform
import pymysql
import pyodbc
import time

############################################################### 기본 변수 세팅
# 현재 실행 위치 담기
gtm_filepos = os.getcwd()

# 오늘날짜 세팅
gtm_now = datetime.datetime.now().strftime('%Y-%m-%d')

# 플랫폼 담기
gtm_platform = platform.system()

gtm_splitdef = ''

# 필요한 파일들 여기에다 위치시키기
if gtm_platform == 'Windows':
    gtm_filepos = gtm_filepos + '\\'
    gtm_splitdef = '\\'
else:
    gtm_filepos = gtm_filepos + '/'
    gtm_splitdef = '/'

os.putenv('NLS_LANG', '.UTF8')


############################################################### 엑셀열기
wb = op.load_workbook(gtm_filepos+"전환대상테이블.xlsx")
ws = wb.worksheets[0]

############################################################### 데이터베이스 연결
# 티베로 연동
print("==================> 티베로 연결 테스트")
# gtm_Tconn  = pyodbc.connect(DSN='DATAWARE',UID='BKUSER',PWD='KOTRABK')
# gtm_Tcur  = gtm_Tconn.cursor()
gtm_Tconn  = pyodbc.connect(DSN='DATAWARE',UID='bknanal',PWD='kotra123!')
gtm_Tcur  = gtm_Tconn.cursor()

# 마리아 DB연동
print("=================> 마리아 연결 테스트")
# gtm_Mconn = pymysql.connect(host='180.100.215.207',port=13306, user='root', 
#                        password='mariadb2022!', db='db_bk', charset='utf8')

gtm_Mconn = pymysql.connect(host='192.168.50.125',port=3306, user='bkuser', 
                       password='1234', db='bkuser', charset='utf8')

gtm_Mcur = gtm_Mconn.cursor()



############################################################### 티베로 데이터 가져오기 및 컬럼개수 세팅
# 티베로 테이블 값 리스트에 담기
print("=================> 티베로 데이터 가져오기")
startT = time.time()

for rowitem in range(2, ws.max_row + 1):
    try:
        tab = str(ws.cell(row=rowitem,column=2).value).strip()

        # 데이터가져오기
        gtm_Tcur.execute("SELECT * FROM BKUSER."+ tab)
        gtm_Trows = gtm_Tcur.fetchall()
        gtm_Trows2 = []

        for row in gtm_Trows:
            gtm_Trows2.append(tuple(row))
        print("==========>" + tab)

        # 로우개수 구하기
        gtm_Tcur.execute("SELECT count(*) FROM BKUSER."+ tab)
        gtm_Trows = gtm_Tcur.fetchone()
        if gtm_Trows:
            ws.cell(row=rowitem, column=3, value=int(gtm_Trows[0]))
        else:
            ws.cell(row=rowitem, column=3, value=0)


        # 마리에 넣을 인서트 구문 넣기
        print("=================> 마리아에 넣기 위한 인서트 구문 만들기")
        colCount = len(gtm_Trows2[0])
        gtm_Tinsert = "INSERT INTO bkuser." + tab + " VALUES("
        for item in range(0,colCount):
            gtm_Tinsert =gtm_Tinsert+"%s,"
        gtm_Tinsert = gtm_Tinsert[0:-1] +")"


        ############################################################# 마리아에 티베로 데이터 넣기
        print("=================> 마리아 데이터 인서트")
        gtm_Mcur.executemany(gtm_Tinsert, gtm_Trows2)
        gtm_Mconn.commit()
    except:
        ws.cell(row=rowitem, column=4, value="오류")
        continue

############################################################## 양이 많은 데이터 분활해서 넣기
# print("=================> 마리아에 넣기 위한 인서트 구문 만들기")

# gtm_Tinsert = "INSERT INTO bkuser.BK_CM_ZIN_T VALUES("
# for item in range(0,19):
#     gtm_Tinsert =gtm_Tinsert+"%s,"
# gtm_Tinsert = gtm_Tinsert[0:-1] +")"

# gtm_Tcur.execute("SELECT * FROM bkuser.BK_CM_ZIN_T")
# gtm_Trows = gtm_Tcur.fetchall()
# gtm_Trows2 = []

# gtm_Tconn.close()

# for row in gtm_Trows:
#     gtm_Trows2.append(tuple(row))

# gtm_Mcur.executemany(gtm_Tinsert, tuple(gtm_Trows2))
# gtm_Mconn.commit()

# intTotalCount = 0
# intCount      = 0

# while True:
#     row = gtm_Tcur.fetchone()
#     intCount = intCount + 1
#     intTotalCount = intTotalCount + 1

#     if intCount > 10000:
#         gtm_Mconn.commit()
#         intCount = 0

#     if row == None:
#         gtm_Mconn.commit()
#         break;

#     gtm_Trows2 = []
#     gtm_Trows2.append(tuple(row))
#     print(str(intTotalCount) + "====> " + str(row))
#     gtm_Mcur.executemany(gtm_Tinsert, tuple(gtm_Trows2))
    
gtm_Tconn.close()
gtm_Mconn.close()

wb.save(gtm_filepos+"전환대상테이블.xlsx")
print("time : ", time.time() - startT)
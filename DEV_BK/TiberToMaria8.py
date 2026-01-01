import openpyxl as op
import datetime
import os
import platform
import pymysql
import pyodbc
import time

############################################################### 기본 변수 세팅
# 현재 실행 위치 담기
gtm_filepos = os.getcwd()

# 오늘날짜 세팅
gtm_now = datetime.datetime.now().strftime('%Y-%m-%d')

# 플랫폼 담기
gtm_platform = platform.system()

gtm_splitdef = ''

# 필요한 파일들 여기에다 위치시키기
if gtm_platform == 'Windows':
    gtm_filepos = gtm_filepos + '\\'
    gtm_splitdef = '\\'
else:
    gtm_filepos = gtm_filepos + '/'
    gtm_splitdef = '/'

os.putenv('NLS_LANG', '.UTF8')


############################################################### 엑셀열기
wb = op.load_workbook(gtm_filepos+"전환대상테이블.xlsx")
ws = wb.worksheets[0]

############################################################### 데이터베이스 연결

gtm_Tconn = ''
gtm_Tcur  = ''
gtm_Mconn = ''
gtm_Mcur  = ''

# 티베로 연결시에는 odbc를 설정해야 한다.
def DBConnection():
    # 티베로 연동
    print("==================> 티베로 연결")
    gtm_Tconn  = pyodbc.connect(DSN='DATAWARE',UID='bknanal',PWD='kotra123!')
    gtm_Tcur  = gtm_Tconn.cursor()

    # 마리아 DB연동
    print("=================> 마리아 연결")
    gtm_Mconn = pymysql.connect(host='192.168.50.125',port=3306, user='bkuser', 
                        password='1234', db='bkuser', charset='utf8')
    gtm_Mcur = gtm_Mconn.cursor()

# 초기 한번 연결하고
DBConnection()


############################################################### 티베로 로우데이터 조사(검증)
for rowitem in range(2, ws.max_row + 1):
    try:
        tab = str(ws.cell(row=rowitem,column=2).value).strip()
        print("----------->" + tab)

        # 티베로 개수
        gtm_Tcur.execute("SELECT count(*) FROM BKUSER."+ tab)
        gtm_Trows = gtm_Tcur.fetchone()
        
        if gtm_Trows:
            ws.cell(row=rowitem, column=3, value=int(gtm_Trows[0]))
        else:
            ws.cell(row=rowitem, column=3, value=0)

    except:
        ws.cell(row=rowitem, column=5, value="오류")
        continue
wb.save(gtm_filepos+"전환대상테이블.xlsx")


############################################################### 티베로 데이터 가져오기(커넥션 타임에 지장이 있을 때)
# 이 타입이 가장 문안하다. 커넥션 타입에 문제가 없다. 가능하면 rownum으로 하지 말자. 제대로 나오지 않는 정보가 있다.
# 티베로 테이블 값 리스트에 담기
print("=================> 티베로 데이터 가져오기")
startT = time.time()

totalCount  = 1
totalCountS = 1
totalCountM = 500

for rowitem in range(2, ws.max_row + 1):
    try:
        tab = str(ws.cell(row=rowitem,column=2).value).strip()

        # 로우가 없으면 넣을 필요 없음
        irowcount = str(ws.cell(row=rowitem,column=3).value).strip()
        if irowcount == 0:
            ws.cell(row=rowitem, column=5, value="성공")
            continue

        gtm_Tcur.execute("SELECT * FROM BKUSER."+ tab)
        
        # 컬럼개수를 알기 위해
        columns = [column[0] for column in gtm_Tcur.description]
        firstColumn = str(columns[0])
        strColumns  = str(columns)
        strColumns  = strColumns.replace('[','')
        strColumns  = strColumns.replace(']','')
        strColumns  = strColumns.replace("""'""",'')

        # 첫번째 컬럼을 row_number()와 함께 사용, order by 절에 사용
        selectColumns = strColumns + ", ROW_NUMBER () OVER(ORDER BY " + firstColumn + ") ROWNUMAS"

        # row_number() 함수로 범위를 지정해서 가져오기 위함. 이상하게 rownum으로 하면 통계정보가 잘 안됨(티베로)
        firstCountrow  = 1
        secondCountrow = 1000
        addCount       = 1000
        
        # 5000개 단위로 재연결하기 위해 사용
        conditonCount = 5000
        condCount     = 1

        while True:
            sss = conditonCount * condCount
            
            # 아래 부분은 자주 끊겨서 어쩔 수 없다. 5000개 단위로 재연결, 아니면 데이터베이스 커넥션 시간을 늘릴 수도 있다.
            if secondCountrow > sss:
                condCount = condCount + 1
                gtm_Mconn.commit()
                gtm_Mconn.close()
                gtm_Tconn.close()
                
                # 재연결하고
                DBConnection()
 
            sql = "SELECT " + strColumns + " FROM (SELECT " + selectColumns + " FROM BKUSER." + tab +" ORDER BY " + firstColumn + ") TA WHERE ROWNUMAS >= " + str(firstCountrow) + " AND ROWNUMAS < " + str(secondCountrow)
            gtm_Tcur.execute(sql)

            gtm_Trows2 = []
            gtm_Trows = gtm_Tcur.fetchall()
            if not gtm_Trows:
                break
            
            # 여기는 중요하다. 반드시 튜플형태로 재가공해야 한다.
            for row in gtm_Trows:
                gtm_Trows2.append(tuple(row))
                print(tab + "-->" + str(totalCount) + "-->" + str(row))
                totalCount = totalCount + 1

            # 마리에 넣을 인서트 구문 넣기
            colCount = len(gtm_Trows2[0])
            gtm_Tinsert = "INSERT INTO bkuser." + tab + " VALUES("
            
            for item in range(0,colCount):
                gtm_Tinsert =gtm_Tinsert+"%s,"
            gtm_Tinsert = gtm_Tinsert[0:-1] +")"
            ws.cell(row=rowitem, column=5, value="성공")

            # 마리아에 티베로 데이터 넣기
            gtm_Mcur.executemany(gtm_Tinsert, gtm_Trows2)
            gtm_Mconn.commit()

            firstCountrow  = secondCountrow 
            secondCountrow = secondCountrow + addCount

    except:
        ws.cell(row=rowitem, column=5, value="오류")
        continue

gtm_Mconn.commit()

############################################################### 티베로 데이터 가져오기 2번째 방법(커네션 타입에 지장이 없을 때)
# 이 타입은 조금 문제가 있다. 하나의 테이블에 데이터양이 많을 때 커넥션 타임에 걸려 끊겨 버린다.

# 티베로 테이블 값 리스트에 담기
# print("=================> 티베로 데이터 가져오기")
# startT = time.time()

# totalCount = 1
# totalCountS = 1
# totalCountM = 500

# for rowitem in range(2, ws.max_row + 1):
#     try:
#         tab = str(ws.cell(row=rowitem,column=2).value).strip()

#         # 로우가 없으면 넣을 필요 없음
#         irowcount = str(ws.cell(row=rowitem,column=3).value).strip()
#         if irowcount == 0:
#             ws.cell(row=rowitem, column=4, value="성공")
#             continue

#         # 일정수 이상이면 다시 접속하기. 사실 여기는 의미 없다.
#         # 왜냐하면 하나의 테이블을 읽는 도중에 커텍션 타임이 설정되어 있기 때문이다.
#         # 하나의 테이블이 데이터수가 굉장히 많을 때 끊김
#         if totalCount > totalCountM * totalCountS:
#             print("================================>재접속")
#             gtm_Mconn.commit()
#             gtm_Tconn.close()
#             gtm_Mconn.close()

#             DBConnection()

#         gtm_Tcur.execute("SELECT * FROM BKUSER."+ tab)
     

#         while True:
#             gtm_Trows2 = []
#             gtm_Trows = gtm_Tcur.fetchmany(1000) # 이함수가 부분적으로 가져온다.
#             if not gtm_Trows:
#                 break

#             for row in gtm_Trows:
#                 gtm_Trows2.append(tuple(row))
#                 print(tab + "-->" + str(totalCount) + "-->" + str(row))
#                 totalCount = totalCount + 1

#             # 마리에 넣을 인서트 구문 넣기
#             colCount = len(gtm_Trows2[0])
#             gtm_Tinsert = "INSERT INTO bkuser." + tab + " VALUES("
            
#             for item in range(0,colCount):
#                 gtm_Tinsert =gtm_Tinsert+"%s,"
#             gtm_Tinsert = gtm_Tinsert[0:-1] +")"
#             ws.cell(row=rowitem, column=4, value="성공")

#             # 마리아에 티베로 데이터 넣기
#             gtm_Mcur.executemany(gtm_Tinsert, gtm_Trows2)
#             gtm_Mconn.commit()
#     except:
#         ws.cell(row=rowitem, column=4, value="오류")
#         continue

# gtm_Mconn.commit()


wb.save(gtm_filepos+"전환대상테이블.xlsx")
print("time : ", time.time() - startT)

############################################################### 마리아 로우데이터 조사(검증)
for rowitem in range(2, ws.max_row + 1):
    try:
        tab = str(ws.cell(row=rowitem,column=2).value).strip()
        print("----------->" + tab)

        # 마리아 개수
        gtm_Mcur.execute("SELECT count(*) FROM bkuser."+ tab)
        gtm_Trows = gtm_Mcur.fetchone()
        
        if gtm_Trows:
            ws.cell(row=rowitem, column=4, value=int(gtm_Trows[0]))
        else:
            ws.cell(row=rowitem, column=4, value=0)
    except:
        ws.cell(row=rowitem, column=5, value="오류")
        continue
wb.save(gtm_filepos+"전환대상테이블.xlsx")


# 데이터베이스를 닫는다.
gtm_Tconn.close()
gtm_Mconn.commit()
gtm_Mconn.close()
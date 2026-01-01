# 복권, 인터넷 사이트에 동행복권 사이트에 들어가면 엑셀로 받을 수 있다.
import random
import openpyxl as op
import pandas as pd


ft = "c:\\work\\PLangVim\\GIDT\\ft.xlsx"

wb = op.load_workbook(ft)
ws = wb.worksheets[0]

# 날짜와 당점번호를 담을 리스트
fList = []
fColList = ['SD','N01','NO2','NO3','NO4','NO5','NO6','NO7']

# 엑셀 파일 읽어서 리스트에 담기
for irow in range(4, ws.max_row + 1):
    imsiList = []
    for icol in range(3,ws.max_column + 1):
        strValue = str(ws.cell(row=irow, column=icol).value)
        imsiList.append(strValue)
    fList.append(imsiList)

# 인덱스는 복권추천일자로 하고 날짜 타입으로 변경, 나머지는 숫자타입으로 변경
df = pd.DataFrame(fList,columns=fColList)
df['SD'] = pd.to_datetime(df['SD'])
df.set_index('SD', inplace=True)
df = df.astype('int64')

# 6개의 숫자를 선택, 숫자가 겹치지 않아야 한다.
fsList = []
while(True):
    si = random.randint(1,45)
    
    # 겹치지 않을 때까지 수행. 처음에는 아무것도 없겠지
    if si in fsList:
        continue
    else:
        fsList.append(si)

    if len(fsList) == 6:
        break

print(fsList)
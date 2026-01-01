# 임현우 박사 승인유예고시 항목 수행
# 삭제예정

import platform
import openpyxl as op
import cx_Oracle
import datetime
import os



# 현재 실행 위치 담기
G_ExFilePos = os.getcwd()

# 오늘날짜 세팅
now = datetime.datetime.now()
strTime = now.strftime('%Y-%m-%d')

# 플랫폼 담기
G_Platform = platform.system()

G_AptSample = ''    # 파일위치
G_SplitDef = ''

# 작업항목마다 여기다 적는다.
if G_Platform == 'Windows':
    G_AptSample = G_ExFilePos + '\\ALL\\'
    G_SplitDef = '\\'
else:
    G_AptSample = G_ExFilePos + '/ALL'
    G_SplitDef = '/'


# 엑셀파일 열기
wb = op.load_workbook(G_AptSample + "2021-01_나민정3.xlsx")
ws = wb.worksheets[0]


for rowitem in range(2, ws.max_row + 1):
    str1 = str(ws.cell(row=rowitem, column=1).value).strip()  # 연번
    str2 = str(ws.cell(row=rowitem, column=2).value).strip()  # 물질명
    str3 = str(ws.cell(row=rowitem, column=3).value).strip()  # CAS No
    str4 = str(ws.cell(row=rowitem, column=4).value).strip()  # CAS No
    str5 = str(ws.cell(row=rowitem, column=5).value).strip()  # CAS No
    str6 = str(ws.cell(row=rowitem, column=6).value).strip()  # CAS No

    str7 = str1 + str2 + str3 + str4 + str5 + str6
    ws.cell(row=rowitem, column=7, value=str7)
    print(rowitem)

# strList6 = []
# strList7 = []
#
# for rowitem in range(2, ws.max_row + 1):
#     str6 = str(ws.cell(row=rowitem, column=6).value).strip()  # 연번
#     str7 = str(ws.cell(row=rowitem, column=7).value).strip()  # 물질명
#
#     strList6.append(str6)
#     strList7.append(str7)
#     print(rowitem)

# for item in range(0,len(strList6)):
#     if strList6[item] in strList7:
#         ws.cell(row=item+1, column=8, value="존재")
#     else:
#         ws.cell(row=item + 1, column=8, value="미존재")
#
#     print(item)


wb.save(G_AptSample + "2021-01_나민정3.xlsx")
wb.close()
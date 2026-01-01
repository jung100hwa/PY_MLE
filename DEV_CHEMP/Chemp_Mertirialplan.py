# 김려옥박사 물질승인신청계획서 추출하기

import platform
import openpyxl as op
import cx_Oracle
import datetime
import pandas as pd
from functools import reduce
import os, re
import shutil

# 기술원 개발서버
os.putenv('NLS_LANG', '.UTF8')
connection = cx_Oracle.connect('chemp/chemp@192.168.50.60:1521/orcl')
cursor = connection.cursor()

# 기술원 운영서버
# os.putenv('NLS_LANG', '.UTF8')
# connection = cx_Oracle.connect('chemp/chemp!1299@10.10.20.10:1521/chemp')
# cursor = connection.cursor()


# 현재 실행 위치 담기
G_ExFilePos = os.getcwd()

# 오늘날짜 세팅
now = datetime.datetime.now()

# 플랫폼 담기
G_Platform = platform.system()

G_FilePdfPos = ''  # 시험성적서 파일 경로
G_AptSample = ''  # 전국아파트 샘플 파일
G_SplitDef = ''

# 작업항목마다 여기다 적는다.
if G_Platform == 'Windows':
    G_FilePdfPos = G_ExFilePos + '\\IDT\\PDF\\'
    G_AptSample = G_ExFilePos + '\\ODT\\Sample\\'
    G_SplitDef = '\\'
else:
    G_FilePdfPos = G_ExFilePos + '/IDT/PDF/'
    G_AptSample = G_ExFilePos + '/ODT/Sample/'
    G_SplitDef = '/'


# 물질승인신청계획서 통계 산출(김려옥 박사)
# 엑셀파일 열기
wb = op.load_workbook(G_AptSample + "물질승인신청계획서.xlsx")

# 지정한 시트 얻기
ws = wb.worksheets[0]

# 1~89번까지 번호 생성
for item in range(1, 90):
    ws.cell(row=1, column=6 + item, value=item)

# 1~ 89번까지 코드값을 담는다.
codeList = []
cursor.execute(
    """SELECT CATEGORY_CD || TEST_ITEM_CD ITEM FROM TN_NBP_CATEGORY_ITEM WHERE SUBSTR(CATEGORY_CD,1,2) = 'B0' ORDER BY DISPLAY_ORDER""")
for item in cursor:
    codeList.append(item[0])

# 제대로 들어 갔느지 한번 찍어 보기
iNo = 1
for item in codeList:
    print(str(iNo) + "-" + item)
    iNo = iNo + 1

i006NO = 1

for item in range(2, ws.max_row + 1):
    subjectNo = ws.cell(row=item, column=1).value

    iNo = 7

    for item_org in codeList:

        item2 = item_org[0:len(item_org) - 3]
        item3 = item_org[16:len(item_org)]

        # 아래는 DB를 수정해서 다시는 발생하지 않음
        # if subjectNo =='20-EAS-P-0783':
        # 	subjectNo ='20-EAS-P-0783,20-EAS-P-0818'

        print(subjectNo + " >>> " + item2 + " >>> " + item3)

        cursor.execute("""SELECT NVL((SELECT CODE_NM FROM COMTCCMMNDETAILCODE WHERE CODE_ID = A.ACQUISITION_STATUS AND CODE = A.ACQUISITION_METHOD), 
            A.ACQUISITION_STATUS) AS NAME FROM TN_NBP_CATEGORY_ITEM_RESPONSE A WHERE A.MST_ID = (SELECT MST_ID FROM TN_NBP_MST WHERE RCT_NO ='%s') 
            AND A.CATEGORY_CD='%s' AND A.TEST_ITEM_CD='%s'""" % (subjectNo, item2, item3))

        codeValue = cursor.fetchone()

        ws.cell(row=item, column=iNo, value=codeValue[0])

        if codeValue[0] == 'NBP006':
            ws.cell(row=item, column=iNo, value="환경부고시 제2019-103호 [별표 2] 확인")

        if codeValue[0] == 'NBP007':
            ws.cell(row=item, column=iNo, value="환경부고시 제2019-103호 [별표 3] 확인")

        if codeValue[0] == 'NBP008':
            ws.cell(row=item, column=iNo, value="화학제품안전법 제13조제2항제3호에 해당하는 자료")

        iNo = iNo + 1

# 엑셀 파일 저장
wb.save(G_AptSample + "물질승인신청계획서.xlsx")
wb.close()
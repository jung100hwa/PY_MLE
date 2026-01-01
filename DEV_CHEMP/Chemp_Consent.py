# 승인대상 생화화학제품 상세 엑셀로 뽑은거 정렬하기
import openpyxl as op
import SU_Init as SU_Init

# 1일 경우 데이터베이스 연결 불필요.
SU_Init.SU_MO_VarInit(1)

wb = op.load_workbook(SU_Init.G_SU_ExFilePosIn + "승인대상목록.xlsx")
ws = wb.worksheets[0]

for rowitem in range(2, ws.max_row + 1):
    
    curRect = str(ws.cell(row=rowitem, column=1).value).strip()      # 접수번호
    print("======> " + str(rowitem) + " : " + curRect)
    
    if curRect != 'None':
        nrt = rowitem + 1
        i = 0
        for rowindex in range(nrt, nrt + 100):
            nextRect = str(ws.cell(row=rowindex, column=1).value).strip()
            if curRect == nextRect:
                i = i + 1
                for colindex in range(1, 31):
                    ws.cell(row=rowindex, column=colindex, value='')
                    # 아래처럼 하면 안됩니다.!!!
                    # ws.merge_cells(start_row=rowindex-1, start_column=colindex, end_row=rowindex, end_column=colindex)
                    # ws.merge_cells(start_row=4, start_column=colindex, end_row=10, end_column=colindex)
            else:
                break
        for colindex in range(1,32):        
            ws.merge_cells(start_row=rowitem, start_column=colindex, end_row=rowitem+i, end_column=colindex)
    else:
        continue

wb.save(SU_Init.G_SU_ExFilePosOut + "승인대상목록_Result2.xlsx")
wb.close()
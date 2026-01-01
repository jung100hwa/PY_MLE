# 엑셀자료를 컨트롤 한다.

import platform
from re import T
import openpyxl as op
import datetime
import os
from openpyxl.descriptors import base
import pandas as pd

# 현재 실행 위치 담기
G_ExFilePos = os.getcwd()

# 오늘날짜 세팅
now = datetime.datetime.now()
strTime = now.strftime("%Y-%m-%d")

# 플랫폼 담기
G_Platform = platform.system()

G_AptSample = ""  # 파일위치
G_SplitDef = ""

# 작업항목마다 여기다 적는다.
if G_Platform == "Windows":
    G_AptSample = G_ExFilePos + "\\ALL\\"
    G_SplitDef = "\\"
else:
    G_AptSample = G_ExFilePos + "/ALL/"
    G_SplitDef = "/"


# 중복값 추출
# 하나의 컬럼값 중복
# df = pd.read_excel(G_AptSample + "중복값2.xlsx")
# sr = df.duplicated(subset=["col3"], keep="first")
# df["중복체크"] = sr
# writer = pd.ExcelWriter(G_AptSample + "중복값결과.xlsx")
# df.to_excel(writer, sheet_name="중복값")
# writer.save()
# print("ok")


# # 전체 로우를 대상으로 중복 여부 판단
# df = pd.read_excel(G_AptSample + "중복값.xlsx")
# sr = df.duplicated(keep="first")
# df["중복체크"] = sr
# writer = pd.ExcelWriter(G_AptSample + "중복값결과.xlsx")
# df.to_excel(writer, sheet_name="중복값")
# writer.save()
# print("ok")


# 다음행과 특정 셀이나 영역을 비교해서 같은 값은 삭제한다. 비교대상이 아닌값들은 그대로 둔다
# 즉 머지의 개념으로 수행(조승연 책임 엑셀), 판다스보다는 그냥 일반적으로 하는게 나을듯 함
# !!일단 데이터가 순차적으로 되어 있어야 함

# 중복값을 머지할 컬럼 순서 번호
m_list = [1, 2]

wb = op.load_workbook(G_AptSample + "그룹머지.xlsx")
ws = wb.worksheets[0]
basev = ''
currv = ''
chtv = True
for rowitem in range(2, ws.max_row + 1):

    print('=====> ' + str(rowitem))

    if chtv:
        for item in range(0, len(m_list)):      # 그룹중 첫번째 행
            imsi = str(ws.cell(row=rowitem, column=m_list[item]).value).strip()
            basev = basev + imsi
        chtv = False
    else:
        for item in range(0, len(m_list)):      # 두번째행부터
            imsi = str(ws.cell(row=rowitem, column=m_list[item]).value).strip()
            currv = currv + imsi

    if basev == currv:
        for colitem in range(1, ws.max_column + 1):
            if colitem in m_list:
                ws.cell(row=rowitem, column=colitem, value='')

    if basev != currv and len(currv) > 0:
        basev = currv

    currv = ''

wb.save(G_AptSample + "그룹머지결과.xlsx")
wb.close()

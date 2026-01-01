# 물질고유번호 변경, 임현우 박사 과학원 관리번호 변경 건
import platform
import openpyxl as op
import cx_Oracle
import datetime
import os

# 기술원 개발서버
# os.putenv('NLS_LANG', '.UTF8')
# connection = cx_Oracle.connect('chemp/chemp@192.168.50.60:1521/orcl')
# cursor = connection.cursor()

# 기술원 운영서버
os.putenv('NLS_LANG', '.UTF8')
connection = cx_Oracle.connect('chemp/chemp!1299@10.10.20.10:1521/chemp')
cursor = connection.cursor()

# 로컬설버
# os.putenv('NLS_LANG', '.UTF8')
# connection = cx_Oracle.connect('scott/tiger@127.0.0.1:1521/XE')
# cursor = connection.cursor()

# 현재 실행 위치 담기
G_ExFilePos = os.getcwd()

# 오늘날짜 세팅
now = datetime.datetime.now()

# 플랫폼 담기
G_Platform = platform.system()

G_SplitDef  = ''

# 작업항목마다 여기다 적는다.
if G_Platform == 'Windows':
    G_ExFilePos = G_ExFilePos + '\\GODT\\'
    G_SplitDef = '\\'
else:
    G_ExFilePos = G_ExFilePos + '/GODT/'
    G_SplitDef = '/'

# 자진취하할 엑셀 파일 열기
wb = op.load_workbook(G_ExFilePos + "고시번호변경.xlsx")
ws = wb.worksheets[0]

# 변수정의
strMaterial     = ''           # 화학물질명
strCasno        = ''           # CAS_NO


for rowitem in range(2, ws.max_row + 1):
    strMaterial   = str(ws.cell(row=rowitem, column=1).value).strip()      # 물질명
    strCasno      = str(ws.cell(row=rowitem, column=2).value).strip()      # CAS_NO

    strMaterial = strMaterial.replace("'", "' || chr(39) || '")
    strMaterial2 = strMaterial.lower()

    print(strMaterial +"=====" +strMaterial2)


    cursor.execute("""UPDATE TN_APD_MST SET CAS_NO = '%s', MATERIAL_NAME='%s' WHERE LOWER(MATERIAL_NAME) ='%s'""" % (strCasno, strMaterial,strMaterial2))
    connection.commit()

    cursor.execute("""UPDATE TN_APD_MST_HIST SET CAS_NO = '%s', MATERIAL_NAME='%s' WHERE LOWER(MATERIAL_NAME) ='%s'""" % (strCasno, strMaterial, strMaterial2))
    connection.commit()

    cursor.execute("""UPDATE TN_EBMMST SET SERIAL_NO = '%s', CHMCLS_NM='%s' WHERE LOWER(CHMCLS_NM) ='%s'""" % (strCasno, strMaterial, strMaterial2))
    connection.commit()

    cursor.execute("""UPDATE TN_EBMMST_HISTORY SET SERIAL_NO = '%s', CHMCLS_NM='%s' WHERE LOWER(CHMCLS_NM) ='%s'""" % (strCasno, strMaterial, strMaterial2))
    connection.commit()

    cursor.execute("""UPDATE TN_EBM_COMPANY_LIST SET CAS_NO = '%s', MATERIAL_NAME='%s' WHERE LOWER(MATERIAL_NAME) ='%s'""" % (strCasno, strMaterial, strMaterial2))
    connection.commit()

    cursor.execute("""UPDATE TN_NBP_MST SET CAS_NO = '%s', MATERIAL_NAME='%s' WHERE LOWER(MATERIAL_NAME) ='%s'""" % (strCasno, strMaterial, strMaterial2))
    connection.commit()

    cursor.execute("""UPDATE TN_NBP_MST_HISTORY SET CAS_NO = '%s', MATERIAL_NAME='%s' WHERE LOWER(MATERIAL_NAME) ='%s'""" % (strCasno, strMaterial, strMaterial2))
    connection.commit()

    cursor.execute("""UPDATE TN_NBP_MATERIAL SET COM_CAS_NO = '%s', COM_MATERIAL_NAME='%s' WHERE LOWER(COM_MATERIAL_NAME) ='%s'""" % (strCasno, strMaterial, strMaterial2))
    connection.commit()


    cursor.execute("""UPDATE TN_EBM_COOPERATION_T SET CAS_NO = '%s', MATERIAL_NM='%s' WHERE LOWER(MATERIAL_NM) ='%s'""" % (strCasno, strMaterial, strMaterial2))
    connection.commit()

    cursor.execute("""UPDATE TN_NBMMST SET SERIAL_NO = '%s', MTTR_NM='%s' WHERE LOWER(MTTR_NM) ='%s'""" % (strCasno, strMaterial, strMaterial2))
    connection.commit()

    cursor.execute("""UPDATE TN_NBMMST_HISTORY SET SERIAL_NO = '%s', MTTR_NM='%s' WHERE LOWER(MTTR_NM) ='%s'""" % (strCasno, strMaterial, strMaterial2))
    connection.commit()

wb.save(G_ExFilePos + "고시번호변경.xlsx")
wb.close()
# 물질승인신청, 안전성종합자료 코드 넣는 것
# 이것은 마인드 맵에서 추출하여 한글로 만든 다음에 엑셀로 코드를 추출한다.
# 그리고 프로그램에서 항목에 대한 코드를 적은 다음 소스 실행

import platform
import openpyxl as op
import cx_Oracle
import datetime
from sshtunnel import SSHTunnelForwarder
import os

# import re, csv
# import statsmodels.formula.api as smf
# import matplotlib.pyplot as plt
# import numpy as np
# import Excel_Basic_PL as EBP

# 기술원 개발서버
os.putenv('NLS_LANG', '.UTF8')
connection = cx_Oracle.connect('chemp/chemp@192.168.50.60:1521/orcl')
cursor = connection.cursor()

# 기술원 운영서버
# os.putenv('NLS_LANG', '.UTF8')
# connection = cx_Oracle.connect('chemp/chemp!1299@10.10.20.10:1521/chemp')
# cursor = connection.cursor()

# 로컬서버(터널이 한번 오류 나니 안되네)
# HOST = "127.0.0.1"
# REMOTE_PORT = 1521
# LOCAL_PORT = 1521
# USER_NAME = "root"
# ssh_pw = '!!hgl1651ok'
# DSN = "scott/tiger@127.0.0.1:1521/XE"


# server = SSHTunnelForwarder(HOST, ssh_username = USER_NAME,
#                                   remote_bind_address = ("127.0.0.1", REMOTE_PORT),
#                                   ssh_password=ssh_pw,
#                                   local_bind_address = ("", LOCAL_PORT))
# server.start()
# connection = cx_Oracle.connect(DSN)
# cursor = connection.cursor()


# 현재 실행 위치 담기
G_ExFilePos = os.getcwd()

# 오늘날짜 세팅
now = datetime.datetime.now()

# 플랫폼 담기
G_Platform = platform.system()

G_FilePdfPos = ''  # 시험성적서 파일 경로
G_AptSample = ''  # 전국아파트 샘플 파일
G_SplitDef = ''

# 작업항목마다 여기다 적는다.
if G_Platform == 'Windows':
    G_FilePdfPos = G_ExFilePos + '\\IDT\\PDF\\'
    G_AptSample = G_ExFilePos + '\\ODT\\Sample\\'
    G_SplitDef = '\\'
else:
    G_FilePdfPos = G_ExFilePos + '/IDT/PDF/'
    G_AptSample = G_ExFilePos + '/ODT/Sample/'
    G_SplitDef = '/'



wb = op.load_workbook(G_AptSample + "코드_효과효능.xlsx")
strTime = now.strftime('%Y-%m-%d')

# 탭마다 즉 BAD 항목마다 유일한 코드, 탭마다 유일한 코드값
# CHEMP_물질_인체 유해성 항목 체계도-202012080200.emm
# main_code_DEF = ['CP1801','CP1802','CP1803','CP1804','CP1805','CP1806','CP1807','CP1808','CP1809','CP1810'
# ,'CP1811','CP1812','CP1833','CP1823','CP1813','CP1827','CP1826','CP1814','CP1815','CP1816','CP1817','CP1818','CP1819','CP1821','CP1822','CP22132']

# CHEMP_물질_환경(생태) 유해성 항목 체계도_202012140200.emm
# main_code_DEF = ['CP2001','CP2002','CP2003','CP2004','CP2005','CP2006','CP2007','CP2008','CP2009','CP2010','CP2011','CP2012',
# 'CP2013','CP2014','CP2016','CP2017','CP231119']

# CHEMP_물질_환경거동 항목 체계도-202012110200.emm
# main_code_DEF = ['CP2018','CP2035','CP2019','CP2020','CP2021','CP2022','CP2023','CP2024','CP2025','CP2026','CP2027','CP2028','CP2029'
# ,'CP2030','CP2034']

# CHEMP_2.1.5.(3.1.5) 노출정보-20210305.emm
# 2장만 작업
# main_code_DEF = ['CP1701','CP1702']

# CHEMP_2.3.3.(3.3.2) 종민감도 분포-20210305_2장만 완료.emm
# main_code_DEF = ['CP2701','CP2702','CP2703','CP2704']

# CHEMP_물질_2.4. 효과효능 평가-202101120200_2장만 완료.emm
# main_code_DEF = ['CP1301', 'CP1302']
# main_code_DEF = ['RS1301', 'RS1302', 'RG17011', 'RG17021', 'RG17022', 'RG17023', 'RG17024']
# main_code_DEF = ['RG2701', 'RG2702', 'RG2703', 'RG2704', 'RU2201','RU3322', 'RU3323', 'RU3324']
main_code_DEF = ['RG1301', 'RG1302', 'RU1601', 'RU1602']

main_code = ''
wsList = wb.sheetnames

# 모든 시트를 읽어 데이터베이스 넣는다.
for sheet_num in range(0, len(wsList)):
    main_code = main_code_DEF[sheet_num]
    ws = wb.worksheets[sheet_num]
    strMainCode = main_code

    strIndex = 0
    strIndex3 = 1
    strMC = ""
    strIndex4 = ""

    # 각 시트마다 수행
    for row_num in range(1, ws.max_row + 1):
        v1 = ws.cell(row=row_num, column=1).value   # 항목
        v2 = ws.cell(row=row_num, column=2).value   # 항목마다 정해진 코드
        v2 = str(v2).strip()

        # 앞에 '-있은 것을 지우자'
        if v2[0:1] == '-':
            v2 = v2[1:].strip()

        if v2 == '기타 (직접입력)':
            v2 = '기타(직접입력)'

        if v1:
            strIndex3 = 1
            strIndex = strIndex + 1
            strIndex2 = str(strIndex)
            if len(strIndex2) == 1:
                strIndex2 = '0' + strIndex2

            # strMC는 그룹코드가 된다.
            strMC = strMainCode + str(strIndex2)

            # strIndex4는 세부코드
            strIndex4 = str(strIndex3)
            if len(strIndex4) == 1:
                strIndex4 = '00' + strIndex4
            elif len(strIndex4) == 2:
                strIndex4 = '0' + strIndex4
        else:
            strIndex3 = strIndex3 + 1
            strIndex4 = str(strIndex3)

            if len(strIndex4) == 1:
                strIndex4 = '00' + strIndex4
            elif len(strIndex4) == 2:
                strIndex4 = '0' + strIndex4
                
        if '기타(직접입력)' in str(v2).replace(' ',''):
            strIndex4 = '9999'

        ws.cell(row=row_num, column=3, value=strMC)
        ws.cell(row=row_num, column=4, value=strIndex4)
        ws.cell(row=row_num, column=5, value=v2)
        ws.cell(row=row_num, column=6, value=v2)
        ws.cell(row=row_num, column=7, value=strIndex3)
        ws.cell(row=row_num, column=8, value='Y')
        ws.cell(row=row_num, column=9, value='admin')
        ws.cell(row=row_num, column=10, value=strTime)
        ws.cell(row=row_num, column=11, value='admin')
        ws.cell(row=row_num, column=12, value=strTime)
        ws.cell(row=row_num, column=13, value='N')
        ws.cell(row=row_num, column=14, value='')

        print(strMC + " - " + strIndex4)

        v2 = v2.replace("'","' || chr(39) || '")
        
        cursor.execute("""INSERT INTO TC_RP_OECD_PICK_CODE_I(EXMP_GROUP_CODE, EXMP_CODE, ENG_EXMP_NM, KOREAN_EXMP_NM, SORT_ORDR, USE_AT, INPUT_ID,INPUT_DT,UPDT_ID,UPDT_DT,NO_SELECT,LABEL_NM)
            VALUES('%s','%s','%s','%s', '%s','%s','%s',to_date('%s','YYYY-MM-DD'),'%s',to_date('%s','YYYY-MM-DD'),'%s','%s')""" % (strMC, strIndex4, v2, v2, strIndex3, 'Y', 'admin',strTime,'admin',strTime,'N',''))
        connection.commit()

    wb.save(G_AptSample + "코드_효과효능.xlsx")
    wb.close()

print("ok")
# server.stop()
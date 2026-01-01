# bad 중에서도 면제 항목 넣을 때 사용하는 것으로 이제는 거의 사용할 일이 없다.


import cx_Oracle
import openpyxl as op
import time, os
import datetime
import platform

# 기술원 개발서버
# os.putenv('NLS_LANG', '.UTF8')
# connection = cx_Oracle.connect('chemp/chemp@192.168.50.60:1521/orcl')
# cursor = connection.cursor()

# 기술원 운영서버
# os.putenv('NLS_LANG', '.UTF8')
# connection = cx_Oracle.connect('chemp/chemp!1299@10.10.20.10:1521/chemp')
# cursor = connection.cursor()

# 현재 실행 위치 담기
G_ExFilePos = os.getcwd()

# 오늘날짜 세팅
now = datetime.datetime.now()

# 플랫폼 담기
G_Platform = platform.system()

G_FilePdfPos = ''  # 시험성적서 파일 경로
G_AptSample = ''  # 전국아파트 샘플 파일
G_SplitDef = ''

# 작업항목마다 여기다 적는다.
if G_Platform == 'Windows':
    G_FilePdfPos = G_ExFilePos + '\\IDT\\PDF\\'
    G_AptSample = G_ExFilePos + '\\ODT\\Sample\\'
    G_SplitDef = '\\'
else:
    G_FilePdfPos = G_ExFilePos + '/IDT/PDF/'
    G_AptSample = G_ExFilePos + '/ODT/Sample/'
    G_SplitDef = '/'

wb = op.load_workbook(G_AptSample + "코드_1장2장만.xlsx")
strTime = now.strftime('%Y-%m-%d')

# 탭마다 즉 BAD 항목마다 유일한 코드, 탭마다 유일한 코드값
# main_code_DEF = ['RS1301', 'RS1302', 'RG17011', 'RG17021', 'RG17022', 'RG17023', 'RG17024']
main_code_DEF = ['RS1301']

main_code = ''
wsList = wb.sheetnames


# 모든 시트를 읽어 데이터베이스 넣는다.
for sheet_num in range(0, len(wsList)):
    # main_code = main_code_DEF[sheet_num]
    ws = wb.worksheets[sheet_num]
    # strMainCode = main_code

    strIndex = 0
    strIndex3 = 1
    strMC = ""
    strIndex4 = ""

    # 각 시트마다 수행
    iCount = 0
    for row_num in range(1, ws.max_row + 1):

        v1 = ws.cell(row=row_num, column=1).value   # 앞에 구분자. 의미 없는 것
        v2 = ws.cell(row=row_num, column=2).value   # 항목마다 코드에 해당되는 명칭
        v2 = str(v2).strip()

        # 앞에 '-있은 것을 지우자'
        if v2[0:1] == '-':
            v2 = v2[1:].strip()

        # 구분자가 있으면. 있으면 그룹코드가 변경된다.
        if v1:
            strIndex3 = 1
            strIndex = strIndex + 1
            strIndex2 = str(strIndex)
            if len(strIndex2) == 1:
                strIndex2 = '0' + strIndex2

            # strMC는 그룹코드가 된다.
            strMainCode = main_code_DEF[iCount]
            strMC = strMainCode + '00'

            iCount = iCount + 1

            # strIndex4는 세부코드
            strIndex4 = str(strIndex3)
            if len(strIndex4) == 1:
                strIndex4 = '00' + strIndex4
            elif len(strIndex4) == 2:
                strIndex4 = '0' + strIndex4

        else:
            strIndex3 = strIndex3 + 1
            strIndex4 = str(strIndex3)

            if len(strIndex4) == 1:
                strIndex4 = '00' + strIndex4
            elif len(strIndex4) == 2:
                strIndex4 = '0' + strIndex4
                
        if '기타(직접입력)' in v2:
            strIndex4 = '9999'

        ws.cell(row=row_num, column=3, value=strMC)
        ws.cell(row=row_num, column=4, value=strIndex4)
        ws.cell(row=row_num, column=5, value=v2)
        ws.cell(row=row_num, column=6, value=v2)
        ws.cell(row=row_num, column=7, value=strIndex3)
        ws.cell(row=row_num, column=8, value='Y')
        ws.cell(row=row_num, column=9, value='admin')
        ws.cell(row=row_num, column=10, value=strTime)
        ws.cell(row=row_num, column=11, value='admin')
        ws.cell(row=row_num, column=12, value=strTime)
        ws.cell(row=row_num, column=13, value='N')
        ws.cell(row=row_num, column=14, value='')


        print(strMC + " - " + strIndex4)

        # v2 = v2.replace("'","' || chr(39) || '")

        # cursor.execute("""INSERT INTO TC_RP_OECD_PICK_CODE_I_BAK(EXMP_GROUP_CODE, EXMP_CODE, ENG_EXMP_NM, KOREAN_EXMP_NM, SORT_ORDR, USE_AT, INPUT_ID,INPUT_DT,UPDT_ID,UPDT_DT,NO_SELECT,LABEL_NM)
        # 	VALUES('%s','%s','%s','%s', '%s','%s','%s',to_date('%s','YYYY-MM-DD'),'%s',to_date('%s','YYYY-MM-DD'),'%s','%s')""" % (strMC, strIndex4, v2, v2, strIndex3, 'Y', 'admin',strTime,'admin',strTime,'N',''))
        # connection.commit()

    wb.save(G_AptSample + "코드_1장2장만.xlsx")
    wb.close()

print("ok")

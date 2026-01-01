# 임현우 연구사 자진취하
import platform
import openpyxl as op
import cx_Oracle
import datetime
import os

# 기술원 개발서버
# os.putenv('NLS_LANG', '.UTF8')
# connection = cx_Oracle.connect('chemp/chemp@192.168.50.60:1521/orcl')
# cursor = connection.cursor()

# 기술원 운영서버
os.putenv('NLS_LANG', '.UTF8')
connection = cx_Oracle.connect('chemp/chemp!1299@10.10.20.10:1521/chemp')
cursor = connection.cursor()

# 로컬설버
# os.putenv('NLS_LANG', '.UTF8')
# connection = cx_Oracle.connect('scott/tiger@127.0.0.1:1521/XE')
# cursor = connection.cursor()

# 현재 실행 위치 담기
G_ExFilePos = os.getcwd()

# 오늘날짜 세팅
now = datetime.datetime.now()

# 플랫폼 담기
G_Platform = platform.system()

G_SplitDef  = ''

# 작업항목마다 여기다 적는다.
if G_Platform == 'Windows':
    G_ExFilePos = G_ExFilePos + '\\GODT\\'
    G_SplitDef = '\\'
else:
    G_ExFilePos = G_ExFilePos + '/GODT/'
    G_SplitDef = '/'

# 자진취하할 엑셀 파일 열기
wb = op.load_workbook(G_ExFilePos + "자진취하_0128.xlsx")
ws = wb.worksheets[0]

# 변수정의
strMstid        = ''           # 공동협의체 MST_ID
strMainMstid    = ''           # 공동협의체 MST_ID
strNo           = ''           # 승인유예고시 연번
strMaterial     = ''           # 화학물질명
strCasno        = ''           # CAS_NO
strCasno2       = ''           # 원 CAS_NO를 저장한다. 과학원 CAS_NO가 통합되고 분리되고 한다.
strUsecode      = ''           # 제품유형 코드
strGigan        = ''           # 승인유예기간
strRectno       = ''           # 접수번호
strWrid         = ''           # 사용자 아이디
strBizno        = ''           # 사업자번호
strAplcnttype   = ''           # 수입,제조,선임(4). 결국 선임을 구하기 위함
strCnnmtn       = ''           # 선임일때 국외제조자
strCnnmtnImsi   = ''           # 선임일때 국외제조자의 공백을 제거
strManageno     = ''           # 선임사실신고증 번호


for rowitem in range(2, ws.max_row + 1):
    strCasno   = str(ws.cell(row=rowitem, column=4).value).strip()      # CAS No
    strUsecode = str(ws.cell(row=rowitem, column=5).value).strip()      # 제품유형코드
    strRectno  = str(ws.cell(row=rowitem, column=7).value).strip()      # 접수번호

    print("ROWITEM = %s" % (rowitem-1))

    # 약간 특수한 경우(2개의 물질을 합친것...참)
    strCasno2 = ''
    if strCasno:
        strCasno2 = strCasno
        if strCasno == '27083-27-8' or strCasno == '32289-58-0':
            strCasno = '27083-27-8, 32289-58-0'
        if strCasno == '105827-78-9' or strCasno == '138261-41-3':
            strCasno = '138261-41-3(105827-78-9)'
        # 고유번호-8이 고유번호-6으로 변경
        # if strCasno == '7681-52-9' or strCasno == '고유번호 없음-6':
        #     strCasno = '7681-52-9'
        if strCasno == '7778-54-3' or strCasno == '고유번호 없음-7':
            strCasno = '7778-54-3'
    else:
        continue

    strUsecode = strUsecode.replace(' ', '')
    if strUsecode=='4-선박·수중시설용오염방지제':
        strUsecode = '4-가.선박·수중시설용오염방지제'
    
    strRectno  = 'EAS-N-' + strRectno

    
    # 유형코드를 조회
    cursor.execute("""SELECT CODE FROM COMTCCMMNDETAILCODE WHERE CODE_ID='BCC003' AND REPLACE(CODE_NM,' ','')='%s'""" % (strUsecode))
    ptcode = cursor.fetchone()
    if ptcode:
        strUsecode = ptcode[0]
    else:
        ws.cell(row=rowitem, column=13, value='제품유형코드 존재하지 않음')
        continue


    # 기존 살생물물질 신고에서 아이디를 먼저 구하고 회원테이블에서 사업자번호를 구한다.
    cursor.execute("""SELECT WR_ID, APLCNT_TYPE, CN_NM_TN FROM TN_EBMMST WHERE RCPTNO='%s'""" % (strRectno))
    strResult     = cursor.fetchone()
    
    if strResult:
        strWrid         = str(strResult[0])
        strAplcnttype   = str(strResult[1])  # 선임여부
        strCnnmtn       = str(strResult[2])  # 국외제조자

        # 기업회원 테이블에서 먼저 조회
        cursor.execute("""SELECT BIZRNO FROM COMTNENTRPRSMBER WHERE ENTRPRS_MBER_ID='%s'""" % (strWrid))
        strBizno = cursor.fetchone()
        
        if strBizno:
            strBizno = str(strBizno[0])
            strBizno = strBizno.replace('-', '')
        else:  
            # 기업회원에 없으면 일반회원에서 구한다.
            cursor.execute("""SELECT BIZRNO FROM COMTNENTRPRSMBER WHERE ENTRPRS_MBER_ID=
            (SELECT ENTRPRS_MBER_ID FROM COMTNGNRLMBER WHERE MBER_ID='%s')""" % (strWrid))
            strBizno = cursor.fetchone()
            
            if strBizno:
                strBizno = str(strBizno[0])
                strBizno = strBizno.replace('-', '')
            else:
                ws.cell(row=rowitem, column=13, value='접수번호에 해당되는 아이디 미존재')
                continue

    else:
        ws.cell(row=rowitem, column=13, value='접수번호에 해당되는 아이디 미존재')
        continue


    # TN_EBM_COMPANY_LIST 이 테이블에 DEL_YN를 'Y'로 변경, 물질승인신청계획서 제출을 위함
    cursor.execute("""UPDATE TN_EBM_COMPANY_LIST SET DEL_YN='Y' WHERE RCT_NO='%s' AND (CAS_NO='%s' OR CAS_NO='%s') AND ITEM_CD='%s'"""
        % (strRectno, strCasno, strCasno2, strUsecode))
    connection.commit()


    # 선임자테이블에서 CAS_NO, ITEM_CD, 국외제조자를 조건으로 해서 선임번호를 구한다.
    # 동일 국외제조자 인데 추가로 신청하면 선임사실신고번호가 2개 이상이 생김에 따라 아래와 같이 모두 , 로 묶는다.
    # 동일물질 동일유형에 대해 선임할 때 수입자 개수와 해임할때 수입자 개수를 비교해서 같은면 아래에서 유협협의체츨 삭제. 다르면 유지해야 한다.
    if strAplcnttype == '4':
        strCnnmtnImsi = strCnnmtn.replace(' ','')
        strCnnmtnImsi = strCnnmtnImsi.replace("'", "' || chr(39) || '")
        
        cursor.execute("""SELECT LISTAGG(MANAGE_NO, ',') WITHIN GROUP(ORDER BY MANAGE_NO) OVER(PARTITION BY APP_CN) AS MANAGE_NO 
        FROM TN_APD_MST WHERE CAS_NO='%s' AND ITEM_CD='%s' AND GUBUN='001' AND
        BIZRNO='%s' AND STS = '005' AND REPLACE(APP_CN,' ','')='%s'""" % (strCasno, strUsecode, strBizno, strCnnmtnImsi))

        strResult = cursor.fetchone()
        if strResult:
            strManageno = strResult[0]
    else:
        strManageno     = 'N'
        strCnnmtn       = '-'
        strCnnmtnImsi   = '-'


    # 공동협의체 작업
    # 유형협의체의 MST_ID를 구한다. 없으면 기존에 아에 등록하지 않았거나 중간에 취하, 삭제한 것임
    # 여기가 좀 이상함. CAS_No. 중간에 합쳐치고 바뀌었기 때문에 하나의 값이 나오지 않을 수 있음. 이것은 요청이 있을 때 수정해줘야 함
    # 여기를 TN_EBM_COMPANY_LIST 이것처럼 (CAS_NO='%s' OR CAS_NO='%s') 이렇게 해버리면 두개의 값이 나오는게 있을 것임 분명히
    cursor.execute("""SELECT MST_ID FROM TN_EBM_COOPERATION_T WHERE CAS_NO='%s' AND USE_CODE='%s'""" % (strCasno, strUsecode))
    strMstid = cursor.fetchone()

    if strMstid:
        strMstid = strMstid[0]
    else:
        ws.cell(row=rowitem, column=13, value='제품유형 협의체가 존재하지 않음')
        continue


    # 물질협의체 메인에서 해당 유형협의체 MST_ID를 구한다
    cursor.execute("""SELECT MST_ID FROM TN_EBM_COOPERATION_T WHERE CAS_NO='%s' AND USE_CODE='000'""" % (strCasno))
    strMainMstid = cursor.fetchone()

    if strMainMstid:
        strMainMstid = str(strMainMstid[0])
    else:
        strMainMstid = 'kkk'

    # 유형협의체에서 해당건에 롤을 구한다. 대표자, 구성원
    cursor.execute("""SELECT GROUP_ROLE FROM TN_EBM_COOPERATION_MAP_T WHERE MST_ID='%s' AND BIZ_NO='%s' AND 
    AGENT_YN='%s' AND REPLACE(AGENT_NAME,' ','')='%s'""" % (strMstid, strBizno, strManageno, strCnnmtnImsi))
    strGrouprole = cursor.fetchone()

    if strGrouprole:
        strGrouprole = str(strGrouprole[0])

    # 대표자 일경우 유형협의체의 상태값을 바꾼다. "구성중으로 바꾼다"
    if strGrouprole == '001':
        cursor.execute("""UPDATE TN_EBM_COOPERATION_T SET STS='001' WHERE MST_ID='%s'""" % (strMstid))
        connection.commit()

    # 유형협의체에서 삭제한다. 여기서 조심해야 한다. 선임일 경우 수입자 10명 2명만 자진취하 했다고 해서 삭제하면 안된다. 좀더 조사해야 함.
    cursor.execute("""DELETE FROM TN_EBM_COOPERATION_MAP_T WHERE MST_ID='%s' AND BIZ_NO='%s' AND 
    AGENT_YN='%s' AND REPLACE(AGENT_NAME,' ','')='%s'""" % (strMstid, strBizno, strManageno, strCnnmtnImsi))
    connection.commit()

    # 유형협의체 참여자수를 갱신한다. 009-아직 공동협의체 가입하지 하지 않음. 008-Y : 탈퇴자 제외
    # 사실 여기도 별로 의미 없다. 왜나면 로우수로 불러오기 때문(2020년에 소수 수정)
    cursor.execute("""UPDATE TN_EBM_COOPERATION_T SET PARTICIPANT_COUNT = (SELECT COUNT(*) FROM TN_EBM_COOPERATION_MAP_T WHERE
        (GROUP_ROLE <> '009' AND (GROUP_ROLE, APPROVE_YN) NOT IN (('008', 'Y'))) AND MST_ID = '%s') WHERE MST_ID ='%s'""" % (
    strMstid, strMstid))
    connection.commit()

    # ************************************** 중요사항
    # 원래는 아이디까지 해야 하는데 같은 업체이면서 아이디가 틀린 경우가 있음. 나중에 협의체랑 협의봐야 함
    # 따라서 유형에서는 삭제되었어도 물질협의체는 삭제된 아이디로 해야 되는 경우가 생김
    # 아래는 하나의 물질에 여러 유형에 삭제되는 기업이 자진취하 대상이 아닌 경우가 제품유형협의체에 하나라도 존재하는지 검토(만약에 있으면 물질협의체에서 삭제 하면 안됨)
    cursor.execute("""SELECT COUNT(*) AS GA FROM TN_EBM_COOPERATION_MAP_T WHERE MST_ID IN (SELECT MST_ID FROM TN_EBM_COOPERATION_T
        WHERE CAS_NO IN (SELECT CAS_NO FROM TN_EBM_COOPERATION_T WHERE MST_ID='%s') AND CATEGORY ='C')
        AND BIZ_NO='%s'""" % (strMainMstid, strBizno))
    strMstidexist = cursor.fetchone()

    # 만약에 자진취하 대상이 존재하면
    if strMstidexist:
        strMstidexist = int(strMstidexist[0])
    else:
        continue

    # 더이상 해당업체가 존재하지 않으면 물질협의체 대표자 인지 검사하고 대표자이면 물질협의체 상태를 구성중으로 변경
    # 또한 물질협의체에서 삭제하고 물질협의체 구성원수를 하나 뺀다.
    if strMainMstid == 'kkk':
        continue

    if strMstidexist == 0:
        # 물질협의체에서 해당건이 대표자 이면 물질협의체의 상태값을 "구성중"으로 변경, 물질협의체에는 무조건 하나의 기업만 존재 함
        cursor.execute("""SELECT GROUP_ROLE FROM TN_EBM_COOPERATION_MAP_T WHERE MST_ID='%s' AND BIZ_NO='%s'""" % (
            strMainMstid, strBizno))
        strMainGrouprole = cursor.fetchone()

        if strMainGrouprole:
            strMainGrouprole = str(strMainGrouprole[0])

            # 위에는 유형협의체의 상태이고 삭제된 이 기업이 물질협의체의 대표일경우에 수행
            if strMainGrouprole == '001':
                cursor.execute("""UPDATE TN_EBM_COOPERATION_T SET STS='001' WHERE MST_ID='%s'""" % (strMainMstid))
                connection.commit()

        # 물질협의체에서 삭제한다.
        cursor.execute("""DELETE FROM TN_EBM_COOPERATION_MAP_T WHERE MST_ID='%s' AND BIZ_NO='%s'""" % (
        strMainMstid, strBizno))
        connection.commit()

        # 물질협의체의 구성원수를 하나 감소한다. !! 사실 이부분도 크게 의미는 없다. 개수를 로우수로 하기 때문
        cursor.execute("""UPDATE TN_EBM_COOPERATION_T SET PARTICIPANT_COUNT = (SELECT COUNT(*) FROM TN_EBM_COOPERATION_MAP_T
        WHERE MST_ID='%s') WHERE MST_ID ='%s'""" % (strMainMstid, strMainMstid))
        connection.commit()

wb.save(G_ExFilePos + "자진취하_0128.xlsx")
wb.close()
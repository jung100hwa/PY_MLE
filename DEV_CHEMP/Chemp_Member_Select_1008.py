# 기존살생물물질 신고번호에 해당하는 회원정보 추출
# 국립환경과학원 요청이 있을 때 수행

import platform
import openpyxl as op
import cx_Oracle
import datetime
import os


# 기술원 개발서버
# os.putenv('NLS_LANG', '.UTF8')
# connection = cx_Oracle.connect('chemp/chemp@192.168.50.60:1521/orcl')
# cursor = connection.cursor()

# 기술원 운영서버
os.putenv('NLS_LANG', '.UTF8')
connection = cx_Oracle.connect('chemp/chemp!1299@10.10.20.10:1521/chemp')
cursor = connection.cursor()


# 로컬설버
# os.putenv('NLS_LANG', '.UTF8')
# connection = cx_Oracle.connect('scott/tiger@127.0.0.1:1521/XE')
# cursor = connection.cursor()


# 현재 실행 위치 담기
G_ExFilePos = os.getcwd()

# 오늘날짜 세팅
now = datetime.datetime.now()
strTime = now.strftime('%Y-%m-%d')

# 플랫폼 담기
G_Platform = platform.system()


G_SplitDef = ''

# 작업항목마다 여기다 적는다.
if G_Platform == 'Windows':
    G_ExFilePos = G_ExFilePos + '\\GODT\\'
    G_SplitDef = '\\'
else:
    G_ExFilePos = G_ExFilePos + '/GODT/'
    G_SplitDef = '/'


# 엑셀파일 열기
wb = op.load_workbook(G_ExFilePos + "자진취하 현황 2021-12-31.xlsx")
ws = wb.worksheets[0]

# 변수정의
strRectno        = ''           # 접수번호
strJurirno       = ''           # 법인등록번호
strBizno         = ''           # 사업자등록번호
strCompany       = ''           # 회사명
strCxfc          = ''           # 대표자명
strApplicannm    = ''           # 담당자명
strApplicantel   = ''           # 담당자전화번호
strApplicanmail  = ''           # 담당자이메일
strApplicanuno   = ''           # 우편번호
strApplicanujuso = ''           # 주소


# 26부터 시작
for rowitem in range(2, ws.max_row + 1):
    strRectno = str(ws.cell(row=rowitem, column=6).value).strip()                   # 연번
    strRectno = 'EAS-N-' + strRectno
    print("==>" + str(rowitem) + "==>" + strRectno)

    cursor.execute("""SELECT WR_ID FROM TN_EBMMST WHERE RCPTNO='%s'""" % (strRectno))
    strResult     = cursor.fetchone()

    # 해당아이디 저장
    if strResult:
        strWrid         = str(strResult[0])
        ws.cell(row=rowitem, column=28, value=strWrid)

        # 기업회원 테이블에서 먼저 조회
        cursor.execute("""SELECT JURIRNO, BIZRNO, CMPNY_NM, CXFC, APPLCNT_NM, 
        CB_FRST_TELNO || '-' || CB_MDDL_TELNO || '-' || CB_LAST_TELNO, 
        APPLCNT_EMAIL_ADRES, ZIP, ADRES || ' ' || DETAIL_ADRES FROM COMTNENTRPRSMBER WHERE ENTRPRS_MBER_ID='%s'""" % (strWrid))
        strResult = cursor.fetchone()

        if strResult:
            strJurirno = str(strResult[0])
            strJurirno = strJurirno.replace('-', '')
            if strJurirno=='None':
                strJurirno = ''
            ws.cell(row=rowitem, column=29, value=strJurirno)

            strBizno = str(strResult[1])
            strBizno = strBizno.replace('-', '')
            ws.cell(row=rowitem, column=30, value=strBizno)

            strCompany = str(strResult[2])
            strCompany = strCompany.replace(' ', '')
            ws.cell(row=rowitem, column=31, value=strCompany)

            strCxfc = str(strResult[3])
            strCxfc = strCxfc.replace(' ', '')
            ws.cell(row=rowitem, column=32, value=strCxfc)

            strApplicannm = str(strResult[4])
            strApplicannm = strApplicannm.replace(' ', '')
            ws.cell(row=rowitem, column=33, value=strApplicannm)

            strApplicantel = str(strResult[5])
            strApplicantel = strApplicantel.replace(' ', '')
            ws.cell(row=rowitem, column=34, value=strApplicantel)

            strApplicanmail = str(strResult[6])
            strApplicanmail = strApplicanmail.replace(' ', '')
            ws.cell(row=rowitem, column=35, value=strApplicanmail)

            strApplicanuno = str(strResult[7])
            strApplicanuno = strApplicanuno.replace(' ', '')
            ws.cell(row=rowitem, column=36, value=strApplicanuno)

            strApplicanujuso = str(strResult[8])
            ws.cell(row=rowitem, column=37, value=strApplicanujuso)


        else:
            # 기업회원에 없으면 일반회원에서 구한다.


            cursor.execute("""SELECT TB.JURIRNO, TB.BIZRNO, TB.CMPNY_NM, TB.CXFC, 
            TA.MBER_NM, TA.CB_FRST_CERL_TELNO || '-' || TA.CB_MDDL_CERL_TELNO || '-' || TA.CB_LAST_CERL_TELNO,
            TA.CB_EMAIL_ID || '@' || TA.CB_EMAIL_DOMN_NM,TA.ZIP, TA.ADRES || ' ' || TA.DETAIL_ADRES 
            FROM COMTNGNRLMBER TA LEFT OUTER JOIN COMTNENTRPRSMBER TB ON TA.ENTRPRS_MBER_ID = TB.ENTRPRS_MBER_ID WHERE TA.MBER_ID='%s'""" % (strWrid))
            strResult2 = cursor.fetchone()

            if strResult2:
                strJurirno = str(strResult2[0])
                strJurirno = strJurirno.replace('-', '')
                if strJurirno=='None':
                    strJurirno=''
                ws.cell(row=rowitem, column=29, value=strJurirno)

                strBizno = str(strResult2[1])
                strBizno = strBizno.replace('-', '')
                ws.cell(row=rowitem, column=30, value=strBizno)

                strCompany = str(strResult2[2])
                strCompany = strCompany.replace(' ', '')
                ws.cell(row=rowitem, column=31, value=strCompany)

                strCxfc = str(strResult2[3])
                strCxfc = strCxfc.replace(' ', '')
                ws.cell(row=rowitem, column=32, value=strCxfc)

                strApplicannm = str(strResult2[4])
                strApplicannm = strApplicannm.replace(' ', '')
                ws.cell(row=rowitem, column=33, value=strApplicannm)

                strApplicantel = str(strResult2[5])
                strApplicantel = strApplicantel.replace(' ', '')
                ws.cell(row=rowitem, column=34, value=strApplicantel)

                strApplicanmail = str(strResult2[6])
                strApplicanmail = strApplicanmail.replace(' ', '')
                ws.cell(row=rowitem, column=35, value=strApplicanmail)

                strApplicanuno = str(strResult2[7])
                strApplicanuno = strApplicanuno.replace(' ', '')
                ws.cell(row=rowitem, column=36, value=strApplicanuno)

                strApplicanujuso = str(strResult2[8])
                ws.cell(row=rowitem, column=37, value=strApplicanujuso)



wb.save(G_ExFilePos + "자진취하 현황 2021-12-31.xlsx")
wb.close()

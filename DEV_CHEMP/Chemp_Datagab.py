# 데이터갭분석 자료 리스트 만들기
# 갭분석 DB 폴더안에 *.XLSM 파일의 일반정보만
# CAS_NO가 없는 것은 CASNO_NO.xlsx 파일을 보고 마지막 검증한다.

import openpyxl as op
import os
import datetime

import cx_Oracle

#######################################여기는 xlsm -> xlsx, 필요시만 주석 풀고 실행
import win32com.client as win32
import os
orgdirname = 'c:\\work\\PY_MYD_G\\IDT\\DB\\'
orgdirname2 = 'c:\\work\\PY_MYD_G\\IDT\\DB2\\'
dirname = os.listdir(orgdirname)


# ############################################################## 1. 파일명을 변경
# 파일명에 특수문자 등으로 잘 익혀지지 않을 때는 파일명을 변경해서 수행
# # 어차피 파일명은 현재 주어진 파일명으로 하지 않음
# excel = win32.gencache.EnsureDispatch('Excel.Application')
# for fname in dirname:
#     try:
#         full_fname = os.path.join(orgdirname, fname)
#         print(full_fname)
#         wb = excel.Workbooks.Open(full_fname)
#         excel.DisplayAlerts = False
#         wb.DoNotPromptForConvert = True
#         wb.CheckCompatibility = False
#         full_fname2 = 'c:\\work\\PY_MYD_G\\IDT\\DB2\\'+ fname.replace('xlsm','xlsx')
#         wb.SaveAs(full_fname2, FileFormat=51, ConflictResolution=2)
#
#     except PermissionError:
#         pass
#
# excel.Application.Quit()



#######################################  DB 넣기 위한 Result파일을 만듬
# now = datetime.datetime.now()
# nowDate = now.strftime('%Y-%m-%d')

# wb2 = op.Workbook()
# ws2 = wb2.active
# ws2.title = 'sheet1'

# def dirsearch(orgdirname, strExt):
#     try:
#         dirname = os.listdir(orgdirname) # 리스트 형식으로 리턴
#         count = 0
#         for fname in dirname:
#             count = count + 1
#             full_fname = os.path.join(orgdirname, fname)
#             if strExt == "":
#                 print(full_fname)
#             elif os.path.splitext((full_fname))[-1] == "." + strExt:
#                 wb = op.load_workbook(full_fname, read_only=True)
#                 ws = wb["일반정보"]

#                 num = ws.cell(row=8, column=1).value # 일련번호
#                 if num:
#                     num = int(num)
#                 chemname = ws.cell(row=8, column=2).value # 화학물질명
#                 if chemname:
#                     chemname = str(chemname).strip()
#                     chemname = chemname.replace('\n','')

#                 casno = ws.cell(row=8, column=3).value # CAS No
#                 casnoyn = ''
#                 if not casno:
#                     casno = ''
#                 else:
#                     casno = str(casno).strip()
#                     casno = casno.replace('\n','')
#                     casno = casno.replace(' ','')
#                     if casno.find('-')==-1 or casno == '고유번호없음':
#                         casnoyn = 'CAS No 확인 필요'

#                 kno = ws.cell(row=8, column=4).value # 기존화학물질 고유번호
#                 if not kno:
#                     kno = ''
#                 else:
#                     kno = str(kno).strip()
#                     kno = kno.replace('\n','')
#                     kno = kno.replace('-', '')

#                 ws2.cell(row=count, column=1, value=count)
#                 ws2.cell(row=count, column=2, value=num)
#                 ws2.cell(row=count, column=3, value=casno)
#                 ws2.cell(row=count, column=4, value=chemname)
#                 ws2.cell(row=count, column=5, value=kno)
#                 ws2.cell(row=count, column=6, value=fname)
#                 ws2.cell(row=count, column=7, value=nowDate) # 수정일자
#                 ws2.cell(row=count, column=8, value=nowDate) # 등록일자
#                 ws2.cell(row=count, column=9, value='관리자') # 등록자
#                 ws2.cell(row=count, column=10, value=casnoyn) # 이상무 검증


#                 print("count = %s, filename = %s" % (count, fname))
#                 wb.close()
#             else:
#                 pass
#     except PermissionError:
#         pass

# dirsearch(orgdirname2,"xlsx")
# wb2.save("GAB_DB_RESULT.xlsx")


#######################################  casno가 없는 것 등 검증
#
# wb = op.load_workbook('GAB_DB_RESULT.xlsx')
# ws = wb.active

# # CAS No가 없는 것을 정의한 엑셀 열기
# wb2 = op.load_workbook('CASNO_NO.xlsx')
# ws2 = wb2.active

# for item in range(1, ws.max_row+1):
#     vb = ws.cell(row=item, column=10).value
#     print('---------------> 검증중')
#     if vb:
#         camnm = ws.cell(row=item, column=4).value
#         camnm = camnm.replace(' ','')

#         # 만약에 CASNO_NO가 없는 것이 CASNO_NO.xlsx에 존재하면 값을 세팅
#         for subitem in range(1, ws2.max_row+1):
#             subcamnm = ws2.cell(row=subitem, column=2).value
#             subcamnm = subcamnm.replace(' ','')
#             if camnm==subcamnm:
#                 subcasno = ws2.cell(row=subitem, column=1).value
#                 ws.cell(row=item, column=3, value=subcasno)
#                 ws.cell(row=item, column=10, value='')
#                 pass

# wb.save('GAB_DB_RESULT.xlsx')

###############################################################


#######################################  파일명을 casno로 바꾸기
# strdir = orgdirname2

# wb = op.load_workbook('GAB_DB_RESULT.xlsx')
# ws = wb.active
# for item in range(1, ws.max_row+1):
#     casno = ws.cell(row=item, column=3).value
#     xlsxname = ws.cell(row=item, column=6).value
#     src = os.path.join(strdir, xlsxname)
#     dst = str(casno) + '.xlsx'
#     ws.cell(row=item, column=6, value=dst)
#     dst = os.path.join(strdir, dst)
#     os.rename(src, dst)
#     print(dst)

# wb.save('GAB_DB_RESULT.xlsx')
# wb.close()


#######################################  산업계 지원단에서 주는 자료로 물질명 바꾸기
wb = op.load_workbook('GAB_DB_RESULT.xlsx')
ws = wb.active

# CAS No가 없는 것을 정의한 엑셀 열기
wb2 = op.load_workbook('GAB.xlsx')
ws2 = wb2.active

for item in range(1, ws.max_row+1):
    vb = ws.cell(row=item, column=3).value
    print('---------------> 검증중')

    # 만약에 CASNO_NO가 없는 것이 CASNO_NO.xlsx에 존재하면 값을 세팅
    for subitem in range(1, ws2.max_row+1):
        subcamnm = ws2.cell(row=subitem, column=1).value
        if vb==subcamnm:
            subcasno = ws2.cell(row=subitem, column=2).value
            ws.cell(row=item, column=4, value=subcasno)
            ws.cell(row=item, column=10, value='물질명 변경')
            pass
wb.save('GAB_DB_RESULT.xlsx')





# #############################################################################산업계 지원단에서 주는 자료로 물질명 바꾸기
# 이것은 이미 들어간 자료를 업데이트 시켰던 것으로 앞으로 수행할일이 없음
# 오라클 연결
# 다이퀘스트
# os.putenv('NLS_LANG', '.UTF8')
# connection = cx_Oracle.connect('chemp/chemp@133.186.171.48:42521/orcl')
# cursor = connection.cursor()

# 기술원 개발서버
# os.putenv('NLS_LANG', '.UTF8')
# connection = cx_Oracle.connect('chemp/chemp@192.168.50.60:1521/orcl')
# cursor = connection.cursor()

# 기술원 운영서버
# os.putenv('NLS_LANG', '.UTF8')
# connection = cx_Oracle.connect('chemp/chemp!1299@10.10.20.10:1521/chemp')
# cursor = connection.cursor()


# # 엑셀파일 열기
# wb = op.load_workbook("c:\work\PY_MYD_G\IDT\GAB.xlsx")
# ws = wb.worksheets[0]

# strCasno= ''
# strNM = ''

# for rowitem in range(1, ws.max_row + 1):
#     strCasno = str(ws.cell(row=rowitem, column=1).value).strip()
#     strNM = str(ws.cell(row=rowitem, column=2).value).strip()

#     cursor.execute("""SELECT CASNO FROM TN_EBM_GAB WHERE CASNO='%s'""" % (strCasno))
#     strga = cursor.fetchone()
#     if strga:
#     	print(strga[0])
#     	strNM = strNM.replace("'","' || chr(39) || '")
#     	cursor.execute("""UPDATE TN_EBM_GAB SET CHEMNM = '%s' WHERE CASNO='%s'""" % (strNM,strCasno))
#     	connection.commit()

#     print(strCasno +'--->'+strNM)


# wb.close()
# 승인유예고시 및 자진취하 DB 전면 검토 용

import platform
import openpyxl as op
import cx_Oracle
import datetime
import os


# 기술원 개발서버
os.putenv('NLS_LANG', '.UTF8')
connection = cx_Oracle.connect('chemp/chemp@192.168.50.60:1521/orcl')
cursor = connection.cursor()

# 기술원 운영서버
# os.putenv('NLS_LANG', '.UTF8')
# connection = cx_Oracle.connect('chemp/chemp!1299@10.10.20.10:1521/chemp')
# cursor = connection.cursor()


# 로컬설버
# os.putenv('NLS_LANG', '.UTF8')
# connection = cx_Oracle.connect('scott/tiger@127.0.0.1:1521/XE')
# cursor = connection.cursor()


# 현재 실행 위치 담기
G_ExFilePos = os.getcwd()


# 오늘날짜 세팅
now = datetime.datetime.now()
strTime = now.strftime('%Y-%m-%d')

# 플랫폼 담기
G_Platform = platform.system()


G_SplitDef = ''

# 작업항목마다 여기다 적는다.
if G_Platform == 'Windows':
    G_ExFilePos = G_ExFilePos + '\\GODT\\'
    G_SplitDef = '\\'
else:
    G_ExFilePos = G_ExFilePos + '/GODT/'
    G_SplitDef = '/'


# 엑셀파일 열기
wb = op.load_workbook(G_ExFilePos + "승인유예고시.xlsx")
ws = wb.worksheets[0]



for rowitem in range(3, ws.max_row + 1):
    print('=============> %s' % (rowitem))
    # 변수정의
    strChmn  = ''         # 화학물질명
    strCasno = ''         # CAS No
    strUsecode = ''       # 유형코드
    strRctno = ''         # 접수번호

    strChmn     = str(ws.cell(row=rowitem, column=2).value).strip()         # 화학물질명
    strCasno    = str(ws.cell(row=rowitem, column=3).value).strip()         # 고유번호
    strUsecode  = str(ws.cell(row=rowitem, column=4).value).strip()         # 유형코드
    strRctno    = str(ws.cell(row=rowitem, column=6).value).strip()         # 기존살생물물질 접수번호

    # 데이터정제
    strState = 0
    
    # #########################################################물질정보가 고시에 있는지 검토
    if strChmn != "None" and strCasno !='None':
        strChmn = strChmn.replace("'", "' || chr(39) || '")
        cursor.execute("""SELECT CAS_NO FROM TN_EBM_CHMCLS WHERE CMS_MTR_ENG_NM='%s' AND CAS_NO='%s'""" % (strChmn,strCasno))
        strReuslt = cursor.fetchone()
        if strReuslt is None:
            ws.cell(row=rowitem, column=21, value='과학원 고시정보에 해당 물질이 없음')
            strState = 1
    else:
        ws.cell(row=rowitem, column=20, value='엑셀에 물질정보가 없음')
        strState = 1

    # #########################################################유형이 DB에 있는지 검토
    if strUsecode == '4-선박·수중시설용오염방지제':
        strUsecode = '4-가.선박·수중시설용오염방지제'
    if strUsecode != 'None':
        strUsecode = strUsecode.replace(' ', '')
        cursor.execute("""SELECT CODE FROM COMTCCMMNDETAILCODE WHERE CODE_ID='BCC003' AND REPLACE(CODE_NM,' ','')='%s'""" % (strUsecode))
        strReuslt = cursor.fetchone()
        if strReuslt is None:
            ws.cell(row=rowitem, column=22, value='과학원 고시정보에 해당 물질이 없음')
            strState = 1
        else:
            strUsecode = strReuslt[0]
    else:
        ws.cell(row=rowitem, column=20, value='엑셀에 유형값이 없음')
        strState = 1

    # #########################################################물질승인신청계획서 테이블에 있는지 검토
    strRctno = 'EAS-N-' + strRctno
    cursor.execute("""SELECT RCT_NO FROM TN_EBM_COMPANY_LIST WHERE RCT_NO='%s' AND CAS_NO='%s' AND ITEM_CD='%s' AND DEL_YN = 'N'""" % (strRctno, strCasno, strUsecode))
    strReuslt = cursor.fetchone()
    if strReuslt is None:
        ws.cell(row=rowitem, column=23, value='물질승인신청계획서 항목에 없습니다')
        strState = 1


    # strSubid = ''       # Subid
    # strMailid = ''      # 이메일 아이디
    # strMaildomain = ''  # 이메일 도메인
    # strSmgrid = ''      # 선임증번호
    # strSregid = ''      # 기존 아이디(TN_EBMMST)에서 구한다.
    # strSmstid = ''      # 선임관리번호
    # strRcptno = ''      # 기존살생물물질 신고 접수번호
    # strNbprcptno = ''   # 물질승인신청계획서 접수번호
    # strBizno = ''       # 사업자등록번호
    # strCasno = ''       # CAS No
    # strUsecode = ''     # 유형코드
    # strAgentyn = ''     # 선임신고증번호
    # strAgentName = ''   # 공동협의체에 등록되어 있는 국외제조자명(물질승인신청에서 이것으로 조인)
    # strMstid = ''       # 공동협의체 유현협의체 관리 아이디
    # strMstidmain = ''   # 공동협의체 물질협의체 관리아이디
    #
    # strSubid     = str(ws.cell(row=rowitem, column=1).value).strip()                   # 서브아이디
    # strSmgrid    = str(ws.cell(row=rowitem, column=2).value).strip()                   # 선임증관리번호
    # strRcptno    = str(ws.cell(row=rowitem, column=3).value).strip()                   # 기존살생물물질 신고번호
    # strNbprcptno = str(ws.cell(row=rowitem, column=4).value).strip()                   # 물질승인신청계획서 접수번호
    # strBizno     = str(ws.cell(row=rowitem, column=5).value).strip()                   # 사업자등록번호
    #
    # if strBizno != "None":
    #     strBizno     = strBizno.replace('-', '')
    # else:
    #     ws.cell(row=rowitem, column=6, value='사업자등록번호가 없음')
    #     continue
    #
    # print('=============> %s' % (strSubid))
    #
    # # 검증해서 잘못된 경우는 아에 등록하지 못하게 함
    # # 1차 서브아이디가 맞는지 확인 그리고 승인이 되었는지 검토
    # cursor.execute("""SELECT MBER_ID FROM COMTNGNRLMBER WHERE MBER_ID='%s' AND MBER_STTUS ='P'""" % (strSubid))
    # sss = cursor.fetchone()
    # if sss is None:
    #     ws.cell(row=rowitem, column=6, value='신규아이디가 존재하지 않거나 승인상태가 아님')
    #     continue
    #
    # # 2차 신고의 접수번호의 사업자번호와 엑셀의 사업자번호를 비교 검토
    # cursor.execute("""SELECT BIZ_NO FROM TN_EBMMST WHERE RCPTNO='%s'""" % (strRcptno))
    # sss = cursor.fetchone()
    # if sss:
    #     strv = str(sss[0])
    #     strv = strv.replace('-','')
    #     if strBizno != strv:
    #         ws.cell(row=rowitem, column=7, value='신고와 사업자번호 불일치')
    #         continue
    # else:
    #     ws.cell(row=rowitem, column=7, value='신고 접수번호 존재하지 않음')
    #     continue
    #
    #
    # # 3차 서브아이디의사업자번호와 엑셀의 사업자번호를 비교
    # cursor.execute("""SELECT BIZRNO FROM COMTNENTRPRSMBER WHERE ENTRPRS_MBER_ID = (SELECT ENTRPRS_MBER_ID FROM COMTNGNRLMBER WHERE MBER_ID = '%s')""" % (strSubid))
    # sss = cursor.fetchone()
    # if sss:
    #     strv = str(sss[0])
    #     strv = strv.replace('-','')
    #     if strBizno != strv:
    #         ws.cell(row=rowitem, column=7, value='일반회원과 사업자번호 불일치')
    #         continue
    # else:
    #     ws.cell(row=rowitem, column=7, value='기업회원이 존재하지 않음')
    #     continue
    #
    #
    # # 4차 선임일 경우 선임자의 사업자번호와 엑셀의 사업자 번호 비교
    # if strSmgrid !='None':
    #     cursor.execute("""SELECT BIZRNO FROM TN_APD_MST WHERE GUBUN = '001' AND STS = '005' AND MANAGE_NO='%s'""" % (strSmgrid))
    #     sss = cursor.fetchone()
    #     if sss:
    #         strv = str(sss[0])
    #         strv = strv.replace('-','')
    #         if strBizno != strv:
    #             ws.cell(row=rowitem, column=7, value='선임자 사업자번호 불일치')
    #             continue
    #     else:
    #         ws.cell(row=rowitem, column=7, value='선임자가 존재하지 않음')
    #         continue
    #
    #
    # # 기존아이디 구하기(물질승인신청테이블 기존 아이디를 신규아이디로 바꾸기 위해)
    # cursor.execute("""SELECT WR_ID FROM TN_EBMMST WHERE RCPTNO='%s'""" % (strRcptno))
    # sss = cursor.fetchone()
    # if sss:
    #     strSregid = str(sss[0])
    # else:
    #     ws.cell(row=rowitem, column=7, value='기존 아이디 정보가 없음')
    #     continue
    #
    # # 일반회원정보 조회(이메일 정보를 업데이트 해주기 위해)
    # cursor.execute("""SELECT CB_EMAIL_ID, CB_EMAIL_DOMN_NM FROM COMTNGNRLMBER WHERE MBER_ID='%s'""" % (strSubid))
    # sss = cursor.fetchone()
    # if sss:
    #     strMailid       = str(sss[0])
    #     strMaildomain   = str(sss[1])
    # else:
    #     ws.cell(row=rowitem, column=7, value='일반아이디 정보가 없음')
    #     continue
    #
    # # 선임증번호
    # if strSmgrid != "None":
    #     cursor.execute("""SELECT MST_ID FROM TN_APD_MST WHERE GUBUN = '001' AND STS = '005' AND MANAGE_NO='%s'""" % (strSmgrid))
    #     sss = cursor.fetchone()
    #     if sss:
    #         strSmstid = str(sss[0])
    #         ws.cell(row=rowitem, column=8, value=strSregid)
    #         strAgentyn = strSmgrid
    #     else:
    #         ws.cell(row=rowitem, column=8, value='선임자아이디가 없음')
    #         continue
    # else:
    #     strAgentyn = 'N'
    #
    # # 선임자테이블 업데이트
    # cursor.execute("""UPDATE TN_APD_MST SET REG_ID='%s', SUBMIT_ID = '%s', CHP_EMAIL_ID='%s', CHP_EMAIL_DOMN_NM = '%s'
    # WHERE MANAGE_NO = '%s'""" % (strSubid, strSubid, strMailid, strMaildomain, strSmgrid))
    # connection.commit()
    #
    #
    # # 선임증 수입자 업데이트
    # cursor.execute("""UPDATE TN_APD_IMPORT SET REG_ID = '%s', DEL_ID = '%s' WHERE MST_ID = '%s'""" % (strSubid, strSubid, strSmstid))
    # connection.commit()
    #
    #
    # # 기존살생물물질 신고테이블 업데이트
    # cursor.execute("""UPDATE TN_EBMMST SET WR_ID='%s', APLCNT_EMAIL_ID='%s', APLCNT_EMAIL_DOMN = '%s'
    # WHERE RCPTNO = '%s'""" % (strSubid, strMailid, strMaildomain, strRcptno))
    # connection.commit()
    #
    # # 기존살생물물질 승인신청계획서 테이블 업데이트
    # cursor.execute("""UPDATE TN_NBP_MST SET SUBMIT_ID='%s', MOD_SUBMIT_ID ='%s', MODIFY_ID = '%s', CREATE_ID='%s', APLCNT_ID='%s'
    # WHERE RCT_NO = '%s'""" % (strSubid, strSubid, strSubid, strSubid,strSubid, strNbprcptno))
    # connection.commit()
    #
    #
    # # 공동협의체 관리시스템
    # # cursor.execute("""SELECT CAS_NO, ITEM_CD FROM TN_EBM_COMPANY_LIST WHERE RCT_NO='%s' AND DEL_YN='N'""" % (strRcptno))
    # cursor.execute("""SELECT CAS_NO, ITEM_CD FROM TN_NBP_MST WHERE RCT_NO = '%s'""" % (strNbprcptno))
    # sss = cursor.fetchone()
    #
    # if sss:
    #     strCasno = str(sss[0])
    #     strUsecode = str(sss[1])
    # else:
    #     ws.cell(row=rowitem, column=10, value='공동협의체에 존재하지 않음')
    #     continue
    #
    # # 제품유형협의체 아이디 업데이트
    # cursor.execute("""SELECT MST_ID FROM TN_EBM_COOPERATION_T WHERE CAS_NO='%s' AND USE_CODE ='%s'""" % (strCasno, strUsecode))
    # sss = cursor.fetchone()
    #
    # if sss:
    #     strMstid = str(sss[0])
    #     cursor.execute("""UPDATE TN_EBM_COOPERATION_MAP_T SET CREATE_ID ='%s', UPDATE_ID='%s'
    #     WHERE MST_ID='%s' AND AGENT_YN='%s' AND BIZ_NO='%s'""" %(strSubid, strSubid, strMstid, strAgentyn, strBizno))
    #     connection.commit()
    # else:
    #     ws.cell(row=rowitem, column=11, value=' 제품유형협의체가 존재하지 않음')
    #     continue
    #
    # # 물질협의체 아이디 업데이트
    # cursor.execute("""SELECT MST_ID FROM TN_EBM_COOPERATION_T WHERE CAS_NO='%s' AND USE_CODE ='%s'""" % (strCasno, '000'))
    # sss = cursor.fetchone()
    #
    # if sss:
    #     strMstidmain = str(sss[0])
    #     cursor.execute("""UPDATE TN_EBM_COOPERATION_MAP_T SET CREATE_ID ='%s', UPDATE_ID='%s'
    #     WHERE MST_ID='%s' AND BIZ_NO='%s'""" %(strSubid, strSubid, strMstidmain, strBizno))
    #     connection.commit()
    # else:
    #     ws.cell(row=rowitem, column=11, value='물질협의체가 존재하지 않음')
    #
    # # 유형협의체 국외제조자명(물질승인신청에서 아이디, casno, 유형, 국외제조자를 검색조건으로 업데이트 하기 위해)
    # cursor.execute("""SELECT AGENT_NAME FROM TN_EBM_COOPERATION_MAP_T
    # WHERE MST_ID='%s' AND AGENT_YN='%s' AND BIZ_NO='%s'""" %(strMstid, strAgentyn, strBizno))
    # sss = cursor.fetchone()
    #
    # if sss:
    #     strAgentName = str(sss[0])
    #
    # ws.cell(row=rowitem, column=13, value=strSregid)
    # ws.cell(row=rowitem, column=14, value=strCasno)
    # ws.cell(row=rowitem, column=15, value=strUsecode)
    # ws.cell(row=rowitem, column=16, value=strSmgrid)
    # ws.cell(row=rowitem, column=17, value=strAgentName)
    #
    #
    # # 물질승인신청테이블 업데이트
    # cursor.execute("""UPDATE TN_NBMMST SET WR_ID ='%s', RV_CHARGER_MAIL1='%s', RV_CHARGER_MAIL2='%s'
    # WHERE WR_ID ='%s' AND SERIAL_NO='%s' AND NBMPRD_TY='%s' AND AGENT_NAME='%s' AND REQ_TYPE=1""" %
    #                (strSubid, strMailid, strMaildomain, strSregid, strCasno, strUsecode, strAgentName))
    # connection.commit()

wb.save(G_ExFilePos + "승인유예고시.xlsx")
wb.close()

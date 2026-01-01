# 승인대상 심기태 연구원 엑셀을 가지고 주성분 업데이트

import platform
import openpyxl as op
import cx_Oracle
import datetime
import os
import re


# 기술원 개발서버
# os.putenv('NLS_LANG', '.UTF8')
# connection = cx_Oracle.connect('chemp/chemp@192.168.50.60:1521/orcl')
# cursor = connection.cursor()

# 기술원 운영서버
# os.putenv('NLS_LANG', '.UTF8')
# connection = cx_Oracle.connect('chemp/chemp!1299@10.10.20.10:1521/chemp')
# cursor = connection.cursor()


# 현재 실행 위치 담기
G_ExFilePos = os.getcwd()

# 오늘날짜 세팅
now = datetime.datetime.now()
strTime = now.strftime('%Y-%m-%d')

# 플랫폼 담기
G_Platform = platform.system()


G_SplitDef = ''

# 작업항목마다 여기다 적는다.
if G_Platform == 'Windows':
    G_ExFilePos = G_ExFilePos + '\\GODT\\'
    G_SplitDef = '\\'
else:
    G_ExFilePos = G_ExFilePos + '/GODT/'
    G_SplitDef = '/'


# 엑셀파일 열기
wb = op.load_workbook(G_ExFilePos + "안생_통합본_주성분_함유량.xlsx")
ws = wb.worksheets[0]

# 변수정의
strCnsno        = ''           # CNS_NO
strMainpr       = ''           # 주성분



for rowitem in range(2, ws.max_row + 1):
    print("==>" + str(rowitem))

    # CNS_NO 정체
    strCnsno  = str(ws.cell(row=rowitem, column=1).value).strip()
    indexno   = strCnsno.find('(')
    if indexno >=0:
        strCnsno = strCnsno[indexno+1 : len(strCnsno)-1]
    ws.cell(row=rowitem, column=4, value=strCnsno)

    # 제품명 정제
    strMainpr = str(ws.cell(row=rowitem, column=2).value).strip()
    p = re.compile('\[M[0-9]{6}\]')
    aList = p.findall(strMainpr)

    if len(aList) > 0:
        for item in aList:
            strMainpr = strMainpr.replace(item,'')

    ws.cell(row=rowitem, column=5, value=strMainpr)

wb.save(G_ExFilePos + "안생_통합본_주성분_함유량.xlsx")
wb.close()

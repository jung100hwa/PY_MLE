# 권혜림 연구원 수입요건 확인번호 통계 뽑는 것

import datetime
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import random
import platform
import openpyxl as op
import cx_Oracle
import os

# 기술원 운영서버
os.putenv('NLS_LANG', '.UTF8')
connection = cx_Oracle.connect('chemp/chemp!1299@10.10.20.10:1521/chemp')
cursor  = connection.cursor()
cursor2 = connection.cursor()

# 현재 실행 위치 담기
G_ExFilePos = os.getcwd()

# 오늘날짜 세팅
now = datetime.datetime.now()
strTime = now.strftime('%Y-%m-%d')

# 플랫폼 담기
G_Platform = platform.system()

G_SplitDef = ''

# 작업항목마다 여기다 적는다.
if G_Platform == 'Windows':
    G_ExFilePos = G_ExFilePos + '\\All\\'
    G_SplitDef = '\\'
else:
    G_ExFilePos = G_ExFilePos + '/All/'
    G_SplitDef = '/'

wb = op.load_workbook(G_ExFilePos + "권혜림수입요건2.xlsx")
ws = wb.worksheets[0]

cursor.execute("""SELECT
    TA.NO,
    TO_CHAR(SAVE_DATE,'YYYY-MM-DD') AS "제출일",
    TO_CHAR(SEND_DATE,'YYYY-MM-DD') AS "접수일",
    TA.RCT_NO    AS "접수번호",
    TO_CHAR(SEND_DATE,'YYYY-MM-DD') AS "발급일", --(접수일과 동일)
    (SELECT CODE_NM FROM COMTCCMMNDETAILCODE WHERE CODE_ID='IRN010' AND CODE=TA.PRO_STS) AS "상태",
    TA.INCOME_REQUEST_NO AS "수입요건확인번호",
    TA.CMPNYNM AS "상호",
    TA.BIZ_NO  AS "사업자번호",
    TA.CXFC    AS "성명(대표자)",
    TA.MBERNM  AS "담당자",
    (TA.CBFRSTCERLTELNO || '-' || TA.CBMDDLCERLTELNO || '-' || TA.CBLASTCERLTELNO) AS "담당자 연락처",
    (TA.CBEMAILID || '@' || TA.CBEMAILDOMNNM) AS "담당자 이메일",
    (TA.CP_ADRES || ' ' || TA.CP_DETAIL_ADRES) AS "소재지(사업장) 주소",
    (TA.CP_TEL1 || '-' || TA.CP_TEL2 || '-' || TA.CP_TEL3) AS "소재지 전화번호",
    (TA.CP_FAX1 || '-' || TA.CP_FAX2 || '-' || TA.CP_FAX3) AS "소재지 팩스번호",
    TA.PDT_NAME AS "제품명",
    CASE WHEN TYPE =2 THEN '신고대상 안전확인대상생활화학제품'
         ELSE '위해우려제품' END AS "종류",
    CASE WHEN TYPE =2 THEN (SELECT DISTINCT(GET_CODE_NM(ITEM_C, DETAIL_ITEM_C)) FROM TN_NLC_ITEM WHERE MST_ID = TA.MST_ID AND EST_NO=TA.EST_NO AND MAJOR_YN='Y')
         ELSE (SELECT DISTINCT(DANGER_DIV) FROM TN_HARM_MST WHERE SELF_CHECK_NO=TA.MST_ID) END AS "품목",
    CASE WHEN TYPE =2 THEN (SELECT DISTINCT(CNS_NO) FROM TN_NLC_MST WHERE MST_ID = TA.MST_ID AND EST_NO=TA.EST_NO)
         ELSE TA.MST_ID END AS "신고번호 또는 자가검사번호",
    TA.EFFTPRD_TRM_DY AS "성적서 유효기간",
    (TO_CHAR(SEND_DATE,'YYYY-MM-DD') || '~' || TO_CHAR(ADD_MONTHS(TO_DATE(TO_CHAR(SEND_DATE,'YYYYMMDD'),'YYYYMMDD'),12)-1,'YYYY-MM-DD')) AS "수입요건 유효기간",
    TA.PDT_EA || '(' || TA.PDT_EA_TYPE || ')' AS "수입량(단위)",
    TA.HSK_NO AS "HSK No.",
    CASE WHEN SUBSTR(TA.MAKE_COUNTRY,0,2)='RG' THEN (SELECT CODE_NM FROM COMTCCMMNDETAILCODE WHERE CODE_ID='CPC003' AND CODE=TA.MAKE_COUNTRY)
         WHEN (SUBSTR(TA.MAKE_COUNTRY,0,2)!='RG' AND TYPE=1) THEN (SELECT DISTINCT(COUNTRY) FROM TN_HARM_MST WHERE SELF_CHECK_NO=TA.MST_ID)
         ELSE TA.MAKE_COUNTRY END AS "제조국",
    TA.MAKE_NAME AS "제조회사명",
    TA.MAKE_ADDRESS AS "제조자 소재지"
FROM
    TN_IRN_MST TA
WHERE
    (TA.TYPE = 2 OR TA.TYPE = 1)
    """)

rowcount = 2
strNo = ''

strList  = [7,8,9,10,11,12,13,14,15,16,17,18,19,20,21]
strListValue = []

# str1='' # 상호
# str2='' # 사업자번호
# str3='' # 성명(대표자)
# str4='' # 담당자
# str5='' # 담당자 연락처
# str6='' # 담당자 이메일
# str7='' # 소재지사업장주소
# str8='' # 소재지 전화번호
# str9='' # 소재지 팩스번호
# str10='' # 소재지 팩스번호
# str11='' # 제품명
# str12='' # 종류
# str13='' # 품목
# str14='' # 신고번호 및 자가검사번호



for row in cursor:
    print(row)
    strListValue = []
    for col in range(0,len(cursor.description)):
        if col == 0:
            strNo=str(row[col])

        va = str(row[col])
        if '' in va:
            va = va.replace('', ' ')
        if '' in va:
            va = va.replace('', ' ')
        if '' in va:
            va = va.replace('', ' ')
        if va == 'None':
            va=''
        ws.cell(row=rowcount, column=col+1, value=va)

        if col in strList:
            if col == 20 or col==21:
                va = ''
            strListValue.append(va)

    rowcount = rowcount + 1

    cursor2.execute(""" SELECT 
        TA.MOD_NO,
        TO_CHAR(SAVE_DATE,'YYYY-MM-DD') AS "제출일",
        TO_CHAR(SEND_DATE,'YYYY-MM-DD') AS "접수일",
        TA.RCT_NO    AS "접수번호",
        TO_CHAR(SEND_DATE,'YYYY-MM-DD') AS "발급일", 
        '재발급' AS "상태",
        TA.INCOME_REQUEST_NO AS "수입요건확인번호",
        '' AS "상호",
        '' AS "사업자번호",
	    '' AS "성명(대표자)",
	    '' AS "담당자",
	    '' AS "담당자 연락처",
	    '' AS "담당자 이메일",
	    '' AS "소재지(사업장) 주소",
	    '' AS "소재지 전화번호",
	    '' AS "소재지 팩스번호",
	    '' AS "제품명",
	    '' AS "종류",
	    '' AS "품목",
	    '' AS "신고번호",
        '' AS "성적서 유효기간",
        '' AS "수입요건 유효기간",
        TA.PDT_EA || '(' || TA.PDT_EA_TYPE || ')' AS "수입량(단위)",
        TA.HSK_NO AS "HSK No.",
        CASE WHEN SUBSTR(TA.MAKE_COUNTRY,0,2)='RG' THEN (SELECT CODE_NM FROM COMTCCMMNDETAILCODE WHERE CODE_ID='CPC003' AND CODE=TA.MAKE_COUNTRY)
             ELSE TA.MAKE_COUNTRY END AS "제조국",
        TA.MAKE_NAME AS "제조회사명",
        TA.MAKE_ADDRESS AS "제조자 소재지"
        FROM TN_IRN_MST_MOD TA WHERE TA.PRO_STS = '04' AND TA.SUBMIT_DATE IS NOT NULL AND TA.SEND_DATE IS NOT NULL AND TA.MOD_NO='%s'""" %(strNo))

    for row in cursor2:
        for col in range(0,len(cursor2.description)):
            va = str(row[col])

            if va == 'None':
                va = ''

            if col in strList:
                ws.cell(row=rowcount, column=col+1, value=strListValue[col-7])
            else:
                ws.cell(row=rowcount, column=col+1, value=va)
    
        rowcount = rowcount + 1


wb.save(G_ExFilePos + "권혜림수입요건2.xlsx")
wb.close()
# 신고대상 전성분 데이터 담기(오화용 연구원)
# 계속해서 이것으로 했으나 오화용연구원이 나민전 연구원것으로 제출해달라고 함

import datetime
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import random
import platform
import openpyxl as op
import cx_Oracle
import os

# 기술원 운영서버
os.putenv('NLS_LANG', '.UTF8')
connection = cx_Oracle.connect('chemp/chemp!1299@10.10.20.10:1521/chemp')
cursor = connection.cursor()

# 현재 실행 위치 담기
G_ExFilePos = os.getcwd()

# 오늘날짜 세팅
now = datetime.datetime.now()
strTime = now.strftime('%Y-%m-%d')

# 플랫폼 담기
G_Platform = platform.system()

G_SplitDef = ''

# 작업항목마다 여기다 적는다.
if G_Platform == 'Windows':
    G_ExFilePos = G_ExFilePos + '\\All\\'
    G_SplitDef = '\\'
else:
    G_ExFilePos = G_ExFilePos + '/All/'
    G_SplitDef = '/'

# strYear = ['2019', '2020', '2021']
strYear = ['2022']

for wsYear in strYear:
    for wsMonth in range(1,2):
        wsMonth = str(wsMonth)
        if len(wsMonth) == 1:
            wsMonth = '0' + wsMonth

        strvalue = wsYear + '-' + wsMonth +'%'
        print(strvalue)
        cursor.execute("""
            SELECT
                AA.MST_ID ,
                AA.EST_NO ,
                AA.RCT_NO AS "접수번호" ,
                AA.CNS_NO AS "신고번호" ,
                AA.ISSUDY AS "발급일자" ,
                AA.CN_NM AS "상호" ,
                CASE
                    WHEN AA.ST = 'represent'
                    AND AA.cl = 'register' THEN '대표'
                    WHEN AA.st = 'represent'
                    AND AA.cl = 'change' THEN '변경(대표)'
                    WHEN AA.st = 'derive'
                    AND AA.cl = 'register' THEN '파생'
                    WHEN AA.st = 'derive'
                    AND AA.cl = 'change' THEN '변경(파생)'
                    ELSE ''
                END "제품구분" ,
                AA.PDT_NM AS "제품명" ,
                CASE
                    AA.ST WHEN 'derive' THEN AA.PDTNM
                    WHEN 'represent' THEN
                    CASE
                        AA.CL WHEN 'change' THEN AA.PDTNM
                    END
                END AS "파생/변경제품명" ,
                BB.ITEMS AS "품목 세부품목:용도-세부용도"
                --, CASE WHEN AA.DETAIL_DF IS NULL THEN AA.DF ELSE AA.DF || '-' || AA.DETAIL_DF END AS "제형"
            ,
                CC.FORM_NM "제형" ,
                AA.USMTR_ID AS "물질ID" ,
                AA.UPPER_USMTR_ID AS "하위 물질ID" ,
                decode(AA.ST, 'represent', '대표', 'derive', '파생') AS "대표/파생 구분" ,
                decode(AA.CL, 'register', '신규', 'change', '변경') AS "신규/변경 분류" ,
                decode(AA.STS, '1', '', '2', '주요물질', '3', '', '4', '', '5', '') AS "물질구분" ,
                AA.CAS_NO ,
                AA.MTRNM AS "물질명" ,
                AA.CBN_RAT AS "함량" ,
                AA.PRPOS_C_NM "용도(기능)" ,
                AA.MTR_ST_NANO AS "물질구분 나노"
                --, AA.MTR_ST_EFFT AS "물질구분 유효"
            ,
                AA.MTR_ST_EFFT AS "물질구분 주요"
                --, AA.MTR_ST_BCM AS "살생물질여부"
            ,
                AA.MTR_ST_BCM AS "물질구분 살생물" ,
                AA.MTR_ST_COMP AS "혼합물질여부" /*사업자번호 및 제조수입여부 추가 요청*/
                ,
                BSNMRG_NO "사업자번호" ,
                DECODE(MF_ICM, 'jejo', '제조', 'suib', '수입') "제조/수입"
            FROM
                (
                SELECT
                    X.*,
                    ROWNUM RN
                FROM
                    (
                    SELECT
                        a.MST_ID ,
                        a.EST_NO ,
                        a.RCT_NO ,
                        a.CNS_NO ,
                        a.ISSUDY AS ISSUDY ,
                        a.CN_NM ,
                        (
                        SELECT
                            PDTNM
                        FROM
                            TN_NLC_MST
                        WHERE
                            MST_ID = a.MST_ID
                            AND EST_NO = '1') AS PDT_NM ,
                        a.PDTNM ,
                        a.DF ,
                        a.DETAIL_DF ,
                        b.USMTR_ID ,
                        b.UPPER_USMTR_ID ,
                        a.ST ,
                        a.CL ,
                        b.ST AS STS ,
                        b.CAS_NO ,
                        b.MTRNM
                        --, CASE WHEN B.UPPER_USMTR_ID IS NOT NULL THEN TO_CHAR(TO_NUMBER(B.CBN_RAT)) ELSE '' END CBN_RAT_
            ,
                        b.CBN_RAT
                        --, (CASE WHEN B.UPPER_USMTR_ID IS NOT NULL AND REGEXP_INSTR(B.CBN_RAT, '^\d*(\.?\d*)$') = 1 THEN TO_CHAR(TO_NUMBER(B.CBN_RAT) / 1000, 'FM9990.99999999999999999999') ELSE B.CBN_RAT END) CVT_CBN_RAT 
            ,
                        b.MTR_ST_NANO ,
                        b.MTR_ST_EFFT ,
                        b.MTR_ST_BCM ,
                        b.MTR_ST_COMP ,
                        c.code_nm || ' ' || b.PRPOS PRPOS_C_NM ,
                        a.MF_ICM ,
                        a.BSNMRG_NO
                    FROM
                        TN_NLC_MST a,
                        TN_NLC_USMTR b,
                        COMTCCMMNDETAILCODE c
                    WHERE
                        a.MST_ID = b.MST_ID
                        AND a.EST_NO = b.EST_NO
                        AND b.PRPOS_C = c.CODE(+)
                        AND c.code_id(+) = 'NLC004'
                        AND a.STS = 'ST008'
                        AND a.DELETE_AT = 'N'
                        AND A.ISSUDY LIKE '%s'
                        --AND A.ISSUDY BETWEEN '2019-01-01' AND '2020-08-31'
            ) X ) AA,
                (
                SELECT
                    a.MST_ID,
                    a.EST_NO,
                    LISTAGG
                    (CASE
                        WHEN a.ITEM_C IS NOT NULL
                            AND a.DETAIL_ITEM_C IS NOT NULL
                            AND a.PRPOS_C IS NOT NULL
                            AND a.DETAIL_PRPOS_C IS NOT NULL THEN c.code_nm || ' ' || d.code_nm || ' : ' || e.code_nm || '-' || f.code_nm
                            WHEN a.ITEM_C IS NOT NULL
                            AND a.DETAIL_ITEM_C IS NOT NULL
                            AND a.PRPOS_C IS NOT NULL
                            AND a.DETAIL_PRPOS_C IS NULL THEN c.code_nm || ' ' || d.code_nm || ' : ' || e.code_nm
                            WHEN a.ITEM_C IS NOT NULL
                            AND a.DETAIL_ITEM_C IS NOT NULL
                            AND a.PRPOS_C IS NULL
                            AND a.DETAIL_PRPOS_C IS NULL THEN c.code_nm || ' ' || d.code_nm
                            WHEN a.ITEM_C IS NOT NULL
                            AND a.DETAIL_ITEM_C IS NULL
                            AND a.PRPOS_C IS NULL
                            AND a.DETAIL_PRPOS_C IS NULL THEN c.code_nm
                        END ,
                        ', ' ) WITHIN GROUP (
                    ORDER BY a.item_id DESC) items
                FROM
                    TN_NLC_ITEM a,
                    tn_nlc_mst b,
                    COMTCCMMNDETAILCODE c ,
                    COMTCCMMNDETAILCODE d ,
                    COMTCCMMNDETAILCODE e ,
                    COMTCCMMNDETAILCODE f
                WHERE
                    a.est_no = b.EST_NO
                    AND a.MST_ID = b.MST_ID
                    AND a.ITEM_C = c.CODE(+)
                    AND a.DETAIL_ITEM_C = d.CODE(+)
                    AND a.PRPOS_C = e.CODE(+)
                    AND a.DETAIL_PRPOS_C = f.CODE(+)
                    AND b.DELETE_AT = 'N'
                    AND b.sts = 'ST008'
                    --AND B.ISSUDY BETWEEN '2019-01-01' AND '2020-08-31'
                    AND b.ISSUDY LIKE '%s'
                GROUP BY
                    a.MST_ID,
                    a.EST_NO ) BB,
                (
                SELECT
                    A.MST_ID,
                    A.EST_NO,
                    LISTAGG(B.CODE_NM || DECODE(C.CODE_NM, NULL, '', '-' || C.CODE_NM) || DECODE(A.DF_F_CHK, NULL, '', '(' || REPLACE(REPLACE(A.DF_F_CHK, '(', ''), ')', '') || ')'), ', ') WITHIN GROUP(
                ORDER BY
                    A.FORM_ID) FORM_NM
                FROM
                    TN_NLC_FORM A,
                    COMTCCMMNDETAILCODE B,
                    COMTCCMMNDETAILCODE C
                WHERE
                    A.DF_C = B.CODE(+)
                    AND A.DETAIL_DF_C = C.CODE(+)
                GROUP BY
                    A.MST_ID,
                    A.EST_NO ) CC
            WHERE
                AA.MST_ID = BB.MST_ID
                AND AA.EST_NO = BB.EST_NO
                AND AA.MST_ID = CC.MST_ID
                AND AA.EST_NO = CC.EST_NO
            ORDER BY
                ISSUDY,
                AA.MST_ID,
                AA.EST_NO,
                NVL(UPPER_USMTR_ID, USMTR_ID),
                USMTR_ID""" % (strvalue, strvalue))

        rowcount=2
        wb = op.load_workbook(G_ExFilePos + "전성분.xlsx")
        ws = wb.worksheets[0]

        for row in cursor:
            print(row)
            for col in range(0,len(cursor.description)):
                va = str(row[col])
                if '' in va:
                    va = va.replace('', ' ')

                if '' in va:
                    va = va.replace('', ' ')

                if '' in va:
                    va = va.replace('', ' ')

                if va == 'None':
                    va=''
                ws.cell(row=rowcount, column=col+1, value=va)
                # ws.cell(row=1, column=col, value='사업자등록번호가 없음')
            rowcount = rowcount + 1

        savefile = wsYear + '-' + wsMonth +".xlsx"
        ws.title = wsYear + '-' + wsMonth
        wb.save(G_ExFilePos + savefile)
        wb.close()
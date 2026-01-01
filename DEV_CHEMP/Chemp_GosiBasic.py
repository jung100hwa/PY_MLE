# 임현우 박사 승인유예고시 항목 수행
# 물질승인신청은 유형협의체에 기반하에 국외제조자별 신청할 계획이었는데
# 국립환경과학원의 선임자는 국외제조자 중심 나머지는 수입자 중심으로 변경해서 이 파일은 계속사용해야 함
# 먼저 이 파일을 수행하고 다음에 선임만 Excel_Oracle_Chemp_Gosi_OR.py를 한번 더 수행해서
# 공동협의체의 선임자일 경우만 국외제조자로 분리해야 함

import platform
import openpyxl as op
import cx_Oracle
import datetime
import os


# 기술원 개발서버
os.putenv('NLS_LANG', '.UTF8')
connection = cx_Oracle.connect('chemp/chemp@192.168.50.60:1521/orcl')
cursor = connection.cursor()

# 기술원 운영서버
# os.putenv('NLS_LANG', '.UTF8')
# connection = cx_Oracle.connect('chemp/chemp!1299@10.10.20.10:1521/chemp')
# cursor = connection.cursor()


# 로컬설버
# os.putenv('NLS_LANG', '.UTF8')
# connection = cx_Oracle.connect('scott/tiger@127.0.0.1:1521/XE')
# cursor = connection.cursor()


# 현재 실행 위치 담기
G_ExFilePos = os.getcwd()

# 오늘날짜 세팅
now = datetime.datetime.now()
strTime = now.strftime('%Y-%m-%d')

# 플랫폼 담기
G_Platform = platform.system()

G_AptSample = ''  # 전국아파트 샘플 파일
G_SplitDef = ''

# 작업항목마다 여기다 적는다.
if G_Platform == 'Windows':
    G_AptSample = G_ExFilePos + '\\GODT\\'
    G_SplitDef = '\\'
else:
    G_AptSample = G_ExFilePos + '/GODT/'
    G_SplitDef = '/'


# 엑셀파일 열기
wb = op.load_workbook(G_AptSample + "GOSI_test.xlsx")
ws = wb.worksheets[0]

strMstid = ''
strMainMstid = ''

for rowitem in range(1, ws.max_row + 1):
    strNo = str(ws.cell(row=rowitem, column=1).value).strip()           # 연번
    strMaterial = str(ws.cell(row=rowitem, column=2).value).strip()     # 물질명
    strCasno = str(ws.cell(row=rowitem, column=3).value).strip()        # CAS No
    strCasno2 = ''

    if strCasno:
        strCasno = strCasno.strip()
        strCasno2 = strCasno

        if strCasno == '27083-27-8' or strCasno == '32289-58-0':
            strCasno = '27083-27-8, 32289-58-0'
        if strCasno == '105827-78-9' or strCasno == '138261-41-3':
            strCasno = '138261-41-3(105827-78-9)'
        # if strCasno == '고유번호 없음-6' or strCasno == '7681-52-9':
        #     strCasno = '7681-52-9'
        if strCasno == '고유번호 없음-7' or strCasno == '7778-54-3':
            strCasno = '7778-54-3'
    else:
        continue

    strUsecode = str(ws.cell(row=rowitem, column=4).value).strip()      # 제품유형코드
    strGigan = str(ws.cell(row=rowitem, column=5).value).strip()        # 승인유예기간
    strRectno = str(ws.cell(row=rowitem, column=6).value).strip()       # 접수번호

    # 제품유형코드를 정제
    strUsecode = strUsecode.replace(' ', '')
    if strUsecode == '4-선박·수중시설용오염방지제':
        strUsecode = '4-가.선박·수중시설용오염방지제'

    # 날짜 끝에 . 없애기
    if strGigan:
        if len(strGigan) == 11:
            strGigan = strGigan[0:len(strGigan) - 1]
            strGigan = strGigan.replace('.', '-')

    # 접수번호를 형식에 맞게 수정
    strRectno = 'EAS-N-' + strRectno

    # 화학물질명에는 수만은 특수문자...아래는 ' 를 처리한다.
    strMaterial = strMaterial.replace("'", "' || chr(39) || '")

    # 유형코드를 조회
    cursor.execute("""SELECT CODE FROM COMTCCMMNDETAILCODE WHERE CODE_ID='BCC003' AND REPLACE(CODE_NM,' ','')='%s'""" % (strUsecode))
    ptcode = cursor.fetchone()

    if ptcode:
        strUsecode = str(ptcode[0])
    else:
        ws.cell(row=rowitem, column=13, value='유형이 존재하지 않음')
        continue

    # 기존 살생물물질 신고에서 아이디를 먼저 구하고 회원테이블에서 사업자번호를 구한다.
    cursor.execute("""SELECT WR_ID FROM TN_EBMMST WHERE RCPTNO='%s'""" % (strRectno))
    strWrid = cursor.fetchone()
    strBizno = ''

    if strWrid:
        strWrid = str(strWrid[0])

        # 기업회원 테이블에서 먼저 조회
        cursor.execute("""SELECT BIZRNO FROM COMTNENTRPRSMBER WHERE ENTRPRS_MBER_ID='%s'""" % (strWrid))
        strBizno = cursor.fetchone()
        
        if strBizno:
            strBizno = str(strBizno[0])
            strBizno = strBizno.replace('-', '')
        else:  
            # 기업회원에 없으면 일반회원에서 구한다.
            cursor.execute("""SELECT BIZRNO FROM COMTNENTRPRSMBER WHERE ENTRPRS_MBER_ID=
            (SELECT ENTRPRS_MBER_ID FROM COMTNGNRLMBER WHERE MBER_ID='%s')""" % (strWrid))
            strBizno = cursor.fetchone()
            
            if strBizno:
                strBizno = str(strBizno[0])
                strBizno = strBizno.replace('-', '')
            else:
                ws.cell(row=rowitem, column=13, value='접수번호에 해당되는 아이디 미존재')
                continue

    else:
        ws.cell(row=rowitem, column=13, value='접수번호에 해당되는 아이디 미존재')
        continue

    # TN_EBM_COMPANY_LIST에 추가 한다. 먼저 있는지 검토한다.
    cursor.execute("""SELECT COUNT(*) FROM TN_EBM_COMPANY_LIST WHERE RCT_NO='%s' AND (CAS_NO='%s' OR CAS_NO='%s') 
        AND ITEM_CD='%s' AND DEL_YN = 'N'""" % (strRectno, strCasno, strCasno2, strUsecode))
    strExist = cursor.fetchone()

    if strExist:
        strExist = int(strExist[0])

        # TN_EBM_COMPANY에 없으면 추가
        if strExist == 0:
            cursor.execute("""INSERT INTO TN_EBM_COMPANY_LIST(RCT_NO, CAS_NO, ITEM_CD, MATERIAL_NAME, NO, DEL_YN, GRACE_DATE) 
            VALUES('%s','%s','%s','%s', %s,'%s','%s')""" % (strRectno, strCasno, strUsecode, strMaterial, strNo, 'N', strGigan))
            connection.commit()
        else:
            ws.cell(row=rowitem, column=13, value='TN_EBM_COMPANY 존재함')

    # 제품유형이 있는지 확인한다.
    cursor.execute("""SELECT MST_ID FROM TN_EBM_COOPERATION_T WHERE CAS_NO='%s' AND USE_CODE='%s'""" % (strCasno, strUsecode))
    strMstid = cursor.fetchone()

    if strMstid:  # 제품유형이 존재하면
        strMstid = str(strMstid[0])

        # 업체가 제품유형협의체에 있는지 조사
        cursor.execute("""SELECT MST_ID FROM TN_EBM_COOPERATION_MAP_T WHERE MST_ID='%s' AND BIZ_NO='%s'""" % (strMstid, strBizno))
        sss = cursor.fetchone()
        if sss:
            ws.cell(row=rowitem, column=13, value='공동협의체에 해당물질 제품유형에 존재함')
        else:
            # 업체가 제품유형협의체에 존재하지 않으면 제품유형에 업체를 추가한다.
            cursor.execute("""INSERT INTO TN_EBM_COOPERATION_MAP_T(MST_ID, BIZ_NO, GROUP_ROLE, AGREE_YN, USE_YN, INDIVISUAL_SUBMIT_YN,
            REQUEST_PLAN_YN, HOPE_REPRESENT_YN, CREATE_DATE, CREATE_ID, UPDATE_DATE, UPDATE_ID, APPROVE_YN, AGENT_YN, AGENT_NAME)
            VALUES('%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', to_date('%s','YYYY-MM-DD'), '%s', to_date('%s','YYYY-MM-DD'), '%s', '%s', '%s', '%s')
            """ % (strMstid, strBizno, '009', 'Y', 'Y', 'N', 'N', '', strTime, strWrid, strTime, strWrid, 'Y', 'N', '-'))
            connection.commit()

    else:  # 제품 유형이 없을 때 신규로 만들어 넣는다
        # 마스터테이블에 유형을 넣고
        strSeq = """'TEC_' || LPAD(SEQ_EBM_COOPERATION_01.NEXTVAL,7,'0')"""
        cursor.execute("""INSERT INTO TN_EBM_COOPERATION_T(MST_ID, CAS_NO, MATERIAL_NM, USE_CODE, PARTICIPANT_COUNT, STS, CATEGORY, 
        CREATE_DATE, CREATE_ID, UPDATE_DATE, UPDATE_ID) VALUES(%s,'%s','%s','%s','%s','%s','%s',to_date('%s','YYYY-MM-DD'),'%s', 
        to_date('%s','YYYY-MM-DD'), '%s')""" % (strSeq, strCasno, strMaterial, strUsecode, '0', '001', 'C', strTime, 'SYSTEM', strTime,'SYSTEM'))
        connection.commit()

        # 제품유형코드를 다시 구한다
        cursor.execute("""SELECT MST_ID FROM TN_EBM_COOPERATION_T WHERE CAS_NO='%s' AND USE_CODE='%s'""" % (strCasno, strUsecode))
        strMstid = cursor.fetchone()
        strMstid = str(strMstid[0])

        # 해당 제품유형코드에 업체정보를 추가한다.
        cursor.execute("""INSERT INTO TN_EBM_COOPERATION_MAP_T(MST_ID, BIZ_NO, GROUP_ROLE, AGREE_YN, USE_YN, INDIVISUAL_SUBMIT_YN, 
        REQUEST_PLAN_YN, HOPE_REPRESENT_YN, CREATE_DATE, CREATE_ID, UPDATE_DATE, UPDATE_ID, APPROVE_YN, AGENT_YN, AGENT_NAME) 
        VALUES('%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', to_date('%s','YYYY-MM-DD'), '%s', to_date('%s','YYYY-MM-DD'), '%s', '%s', '%s', '%s')
        """ % (strMstid, strBizno, '009', 'Y', 'Y', 'N', 'N', '', strTime, strWrid, strTime, strWrid, 'Y', 'N', '-'))
        connection.commit()

    # 유형협의체의 참여자수를 증가시킨다. !! 지금은 소스가 변경되서 큰 의미는 없다
    cursor.execute("""UPDATE TN_EBM_COOPERATION_T SET PARTICIPANT_COUNT = PARTICIPANT_COUNT + 1 WHERE MST_ID='%s'""" % (strMstid))
    connection.commit()

    # !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!여기가 아주 중요함
    # 물질협의체에 인서트 하기 위해 자기가 포함된 유형중에 대표자로 있는 것이 있는지 없으면 
    # 2개 이상 멤버를 가진 유형협의체가 있는지 왜냐하면 멤버가 1명인것은 협의체 가입이 안되기 때문
    # 먼저 CAS NO 중에 유형중에 멤버가 2개 이상인것 중 대표자로 유형협의체가 있는지 검토
    cursor.execute("""SELECT MST_ID FROM TN_EBM_COOPERATION_MAP_T WHERE MST_ID IN (SELECT MST_ID FROM TN_EBM_COOPERATION_MAP_T WHERE
            MST_ID IN (SELECT MST_ID FROM TN_EBM_COOPERATION_T WHERE CAS_NO = '%s' AND USE_CODE <> '000') AND BIZ_NO = '%s' AND GROUP_ROLE='001') GROUP BY 
            MST_ID HAVING COUNT(*) > 1 """ % (strCasno, strBizno))
    strHglExist = cursor.fetchone()

    if strHglExist:
        strMstid = strHglExist[0]
    else:
        # 유형 대표자가 없다면 그냥 개수가 2개 이상 포함된것 있는지 검토(사실 이것은 말이 안됨!!!!!!!!!!!!!나주에 수정해야 함. 무조건 대표자여 함)
        cursor.execute("""SELECT MST_ID FROM TN_EBM_COOPERATION_MAP_T WHERE MST_ID IN (SELECT MST_ID FROM TN_EBM_COOPERATION_MAP_T WHERE
            MST_ID IN (SELECT MST_ID FROM TN_EBM_COOPERATION_T WHERE CAS_NO = '%s' AND USE_CODE <> '000') AND BIZ_NO = '%s') GROUP BY 
            MST_ID HAVING COUNT(*) > 1 """ % (strCasno, strBizno))
        strHglExist = cursor.fetchone()
    if strHglExist:
        strMstid = strHglExist[0]
    else:
        # 해당업체가 포함된 2개 이상의 멤버를 가진 유형이 없으면 물질협의체 코드 수행 필요 없음
        # 하나라도 상관없으면 이부분을 계속 진행하면 됨!!!!!!!!!!!!!!!!!!!!
        continue
    # 물질협의체 존재하는지 본다
    cursor.execute("""SELECT MST_ID FROM TN_EBM_COOPERATION_T WHERE CAS_NO='%s' AND USE_CODE='000'""" % (strCasno))
    strMainMstid = cursor.fetchone()

    if strMainMstid:  # 물질협의체가 존재하면
        strMainMstid = str(strMainMstid[0])
        cursor.execute("""SELECT COUNT(*) FROM TN_EBM_COOPERATION_MAP_T WHERE MST_ID='%s' AND BIZ_NO='%s'""" % (strMainMstid, strBizno))
        strMstidexist = cursor.fetchone()
        strMstidexist = int(strMstidexist[0])

        # !!!!!여기가 아주 중요함. 물질협의체에 멤버로 존재하지 않으면
        # 1. 해당업체가 CAS No의 유형협의체 중 2개 이상 멤버를 가진 유형이 존재할때만 물질협의체에 포함
        # 2. 물질협의체 원본 소속은 2개 이상 구성원이 존재하는 유형협의체 중 유형협의체 대표자인 경우 우선 적용 MST_ID가 들어가야 함.
        # 3. 만약에 유형협의체 대표자가 아닌 경우 일반 구성원으로 추가 하되, 이 업체는 물질협의체의 대표자가 되면 안됨(물질승인신청시 문제가 발생)

        if strMstidexist == 0:
            # 물질협의체에 넣을 때에는 항상 '003'과 대표자 희망도 'N'
            cursor.execute("""INSERT INTO TN_EBM_COOPERATION_MAP_T(MST_ID, BIZ_NO, GROUP_ROLE, AGREE_YN, USE_YN, INDIVISUAL_SUBMIT_YN,
                REQUEST_PLAN_YN, HOPE_REPRESENT_YN, CREATE_DATE, CREATE_ID, UPDATE_DATE, UPDATE_ID, APPROVE_YN,TARGET_MST_ID, AGENT_YN, AGENT_NAME) 
                VALUES('%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', to_date('%s','YYYY-MM-DD'), '%s', to_date('%s','YYYY-MM-DD'), '%s', '%s', '%s', '%s', '%s')"""
                           % (strMainMstid, strBizno, '003', 'Y', 'Y', 'N', 'N', 'N', strTime, strWrid, strTime, strWrid, 'Y', strMstid, 'N', '-'))
            connection.commit()

            cursor.execute("""UPDATE TN_EBM_COOPERATION_T SET PARTICIPANT_COUNT = PARTICIPANT_COUNT + 1 WHERE MST_ID='%s'""" % (strMainMstid))
            connection.commit()
        else:
            ws.cell(row=rowitem, column=13, value='물질협의체에 이미 존재함')

    else:  # 물질협의체가 없으면 물질협의체를 만들고
        strSeq = """'TEC_' || LPAD(SEQ_EBM_COOPERATION_01.NEXTVAL,7,'0')"""

        cursor.execute("""INSERT INTO TN_EBM_COOPERATION_T(MST_ID, CAS_NO, MATERIAL_NM, USE_CODE, PARTICIPANT_COUNT, STS, CATEGORY, 
            CREATE_DATE, CREATE_ID, UPDATE_DATE, UPDATE_ID) VALUES(%s,'%s','%s','%s','%s','%s','%s',to_date('%s','YYYY-MM-DD'),'%s',to_date('%s','YYYY-MM-DD'),'%s')"""
                       % (strSeq, strCasno, strMaterial, '000', '0', '001', 'M', strTime, 'SYSTEM', strTime, 'SYSTEM'))
        connection.commit()

        # 물질협의체의 mst_id 다시 구한다. 새로 넣었으니까
        cursor.execute("""SELECT MST_ID FROM TN_EBM_COOPERATION_T WHERE CAS_NO='%s' AND USE_CODE='000'""" % (strCasno))
        strMainMstid = cursor.fetchone()
        strMainMstid = str(strMainMstid[0])

        cursor.execute("""INSERT INTO TN_EBM_COOPERATION_MAP_T(MST_ID, BIZ_NO, GROUP_ROLE, AGREE_YN, USE_YN, INDIVISUAL_SUBMIT_YN, 
            REQUEST_PLAN_YN, HOPE_REPRESENT_YN, CREATE_DATE, CREATE_ID, UPDATE_DATE, UPDATE_ID, APPROVE_YN,TARGET_MST_ID, AGENT_YN, AGENT_NAME) 
            VALUES('%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', to_date('%s','YYYY-MM-DD'), '%s', to_date('%s','YYYY-MM-DD'), '%s', '%s', '%s', '%s', '%s')"""
                       % (strMainMstid, strBizno, '003', 'Y', 'Y', 'N', 'N', 'N', strTime, strWrid, strTime, strWrid, 'Y', strMstid, 'N', '-'))
        connection.commit()

        # 물질협의체 구성원수를 업데이트 한다. 여기도 마찬가지임.
        cursor.execute("""UPDATE TN_EBM_COOPERATION_T SET PARTICIPANT_COUNT = PARTICIPANT_COUNT + 1 WHERE MST_ID='%s'""" % (strMainMstid))
        connection.commit()

    ws.cell(row=rowitem, column=13, value='OK')
    print('%s=== %s === %s===%s' % (rowitem, strMstid, strWrid, strBizno))

wb.save(G_AptSample + "GOSI_test.xlsx")
wb.close()

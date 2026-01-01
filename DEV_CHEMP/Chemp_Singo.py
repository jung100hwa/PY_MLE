# 하동윤 박사 요청사항

import platform
import openpyxl as op
import cx_Oracle
import datetime
import pandas as pd
from functools import reduce
import os, re
import shutil

# 현재 실행 위치 담기
G_ExFilePos = os.getcwd()

# 오늘날짜 세팅
now = datetime.datetime.now()

# 플랫폼 담기
G_Platform = platform.system()

G_FilePdfPos = ''  # 시험성적서 파일 경로
G_AptSample = ''  # 전국아파트 샘플 파일
G_SplitDef = ''

# 작업항목마다 여기다 적는다.
if G_Platform == 'Windows':
    G_FilePdfPos = G_ExFilePos + '\\IDT\\PDF\\'
    G_AptSample = G_ExFilePos + '\\ODT\\Sample\\'
    G_SplitDef = '\\'
else:
    G_FilePdfPos = G_ExFilePos + '/IDT/PDF/'
    G_AptSample = G_ExFilePos + '/ODT/Sample/'
    G_SplitDef = '/'

# 기술원 개발서버
# os.putenv('NLS_LANG', '.UTF8')
# connection = cx_Oracle.connect('chemp/chemp@192.168.50.60:1521/orcl')
# cursor = connection.cursor()

# 기술원 운영서버
os.putenv('NLS_LANG', '.UTF8')
connection = cx_Oracle.connect('chemp/chemp!1299@10.10.20.10:1521/chemp')
cursor = connection.cursor()

wb = op.load_workbook(G_AptSample + "신고제품_품목별_화학물질 확인결과 추출_0319.xlsx")
ws = wb.worksheets[0]

cursor.execute(""" SELECT tnm.RCT_NO AS 신고번호, DECODE(tnm.ST,'represent','대표','derive','파생') AS 대표파생구분
    ,DECODE(tnm.CL,'register','등록','change','변경') AS 등록변경
    ,T2.CODE_NM AS 품목군 
    ,T3.CODE_NM AS 품목
    ,(SELECT REPLACE(B.PDTNM, CHR(11),'') FROM TN_NLC_MST B WHERE B.MST_ID =  tnm.MST_ID AND B.EST_NO=1) AS 제품명 
    ,(CASE WHEN tnm.ST = 'represent' AND tnm.CL = 'register' THEN '' ELSE tnm.PDTNM END )  AS 파생변경제품명
    ,T4.CODE_NM AS 용도
    ,GET_NLC_FORM_MERGE_NAME(tnm.MST_ID,tnm.EST_NO) AS 제형
    ,DECODE(tnm.MF_ICM,'jejo','제조','suib','수입') AS 제조수입
    ,MF_NATION ||
( SELECT LISTAGG(GET_CODE_NM('',MF_NATION) , ',') WITHIN GROUP (ORDER BY MF_NATION)AS CD
    FROM TN_NLC_MNFCTUR M
    WHERE M.MST_ID = tnm.MST_ID  AND M.EST_NO = tnm.EST_NO
    ) AS 제조국가
    ,CN_NM AS 기업명
    ,INSPCT_CMS_MTRNM AS 확인항목
    ,INSPCT_RST AS 확인결과
    FROM TN_NLC_MST tnm
    ,TN_NLC_ITEM tni 
    ,TN_NLC_INSPCTSCRE ins
    , TN_INSPCTSCRE_DTLINSPCTRST insDtl
    ,COMTCCMMNDETAILCODE T2
    ,COMTCCMMNDETAILCODE T3
    ,COMTCCMMNDETAILCODE T4
    WHERE 
    tnm.MST_ID =tni.MST_ID AND TNM.EST_NO = TNI.EST_NO 
    AND tni.MST_ID =ins.MST_ID AND tni.EST_NO = ins.EST_NO AND tni.ITEM_ID = ins.ITEM_ID 
    AND ins.MST_ID =insDtl.MST_ID AND ins.EST_NO = insDtl.EST_NO AND ins.INSPCTSCRE_ID = insDtl.INSPCTSCRE_ID 
    AND tni.ITEM_C         = T2.CODE(+)
    AND tni.DETAIL_ITEM_C  = T3.CODE(+)
    AND tni.PRPOS_C        = T4.CODE(+)
    AND tnm.sts='ST008'
    -- AND tnm.ISSUDY LIKE '2020-06%'
    ORDER BY RCT_NO""")

row_num = 1
for name in cursor:
    print(str(name[0]) + '================>' + str(row_num))

    if name[0]:
        ws.cell(row=row_num + 1, column=1, value=name[0])

    if name[1]:
        ws.cell(row=row_num + 1, column=2, value=name[1])

    if name[2]:
        ws.cell(row=row_num + 1, column=3, value=name[2])

    if name[3]:
        ws.cell(row=row_num + 1, column=4, value=name[3])

    if name[4]:
        ws.cell(row=row_num + 1, column=5, value=name[4])

    if name[5]:
        ws.cell(row=row_num + 1, column=6, value=name[5])

    if name[6]:
        ws.cell(row=row_num + 1, column=7, value=name[6])

    if name[7]:
        ws.cell(row=row_num + 1, column=8, value=name[7])

    if name[8]:
        ws.cell(row=row_num + 1, column=9, value=name[8])

    if name[9]:
        ws.cell(row=row_num + 1, column=10, value=name[9])

    if name[10]:
        ws.cell(row=row_num + 1, column=11, value=name[10])

    if name[11]:
        ws.cell(row=row_num + 1, column=12, value=name[11])

    if name[12]:
        ws.cell(row=row_num + 1, column=13, value=name[12])

    if name[13]:
        ws.cell(row=row_num + 1, column=14, value=name[13])

    row_num = row_num + 1

wb.save(G_AptSample + "신고제품_품목별_화학물질 확인결과 추출_0319.xlsx")
wb.close()

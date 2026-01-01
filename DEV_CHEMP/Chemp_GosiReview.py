# 일시적인 것으로 승인유예고시 항목과 TN_EBM_COMPANY_LIST와 비교
# 공동협의체에 누락된 것과 자진취하인데 존재하는 것 존재
# 2계의 단계로 하는데 1단계는 승인유예고시가 제대로 이루어졌는지
# 2단계는 자진취하가 제대로 이루어졌는지
# 1단계 결과물을 가지고 승인유예고시 Excel_Oracle_Chemp_Gosi.py를 수행하고
# 2단계 결과물을 가지고 자진취하 Excel_Oracle_Chemp_Cancel.py를 수행함
# 이 때 결과물 엑셀을 가지고 실제 공동협의체나 승인신청계획서에 비교 검토해야 한다. 

import platform
import openpyxl as op
import cx_Oracle
import datetime
import sys, os

# 이렇게 할려면 설정에서 오토세이브 디폴트를 python으로 하면 안된다.
sys.path.append(os.path.dirname(os.path.abspath(os.path.dirname(__file__))))
from SU_MO.SU_Excel_MO import *




# 기술원 개발서버
# os.putenv('NLS_LANG', '.UTF8')
# connection = cx_Oracle.connect('chemp/chemp@192.168.50.60:1521/orcl')
# cursor = connection.cursor()

# 기술원 운영서버
# os.putenv('NLS_LANG', '.UTF8')
# connection = cx_Oracle.connect('chemp/chemp!1299@10.10.20.10:1521/chemp')
# cursor = connection.cursor()

# 로컬설버
os.putenv('NLS_LANG', '.UTF8')
connection = cx_Oracle.connect('scott/tiger@127.0.0.1:1521/XE')
cursor = connection.cursor()

# 현재 실행 위치 담기
G_ExFilePos = os.getcwd()

# 오늘날짜 세팅
now = datetime.datetime.now()
strTime = now.strftime('%Y-%m-%d')

# 플랫폼 담기
G_Platform = platform.system()
G_SplitDef = ''

# 작업항목마다 여기다 적는다.
if G_Platform == 'Windows':
    G_ExFilePos = G_ExFilePos + '\\ODT\\'
    G_SplitDef = '\\'
else:
    G_ExFilePos = G_ExFilePos + '/ODT/'
    G_SplitDef = '/'


# 엑셀파일 열기
wb = op.load_workbook(G_ExFilePos + "승인유예고시_전체.xlsx")
ws = wb.worksheets[0]

strMstid = ''
strMainMstid = ''

# 비교를 위한 변수 정의
eListRectno = []
eListCasno  = []
eListItem   = []
tListRectno = []
tListCasno  = []
tListItem   = []
rTotList    = []

# 여기서는 승인유예고시 파일을 기준이 될 수 있도록 유형코드, 접수번호를 정비한다.
for rowitem in range(2, ws.max_row+1):
    print('====> '+str(rowitem))

    strNo = str(ws.cell(row=rowitem, column=1).value).strip()           # 연번
    strMaterial = str(ws.cell(row=rowitem, column=2).value).strip()     # 물질명
    strCasno = str(ws.cell(row=rowitem, column=3).value).strip()        # CAS No

    # 중간에 cas_no값 합쳐짐. 기존 살생물물질이 합쳐지고 붙고 함
    strCasno2 = ''
    if strCasno:
        strCasno = strCasno.strip()
        strCasno2 = strCasno
        if strCasno == '27083-27-8' or strCasno == '32289-58-0':
            strCasno = '27083-27-8, 32289-58-0'
        if strCasno == '105827-78-9' or strCasno == '138261-41-3':
            strCasno = '138261-41-3(105827-78-9)'
        # if strCasno == '고유번호 없음-6' or strCasno == '7681-52-9':
        #     strCasno = '7681-52-9'
        if strCasno == '고유번호 없음-7' or strCasno == '7778-54-3':
            strCasno = '7778-54-3'
    else:
        continue

    #  접수번호로 조회하면 선임일 경우 선임자 정보가 나옴, 선임자가 아닐 경우는 신고기업 정보
    strUsecode = str(ws.cell(row=rowitem, column=4).value).strip()      # 제품유형코드
    strGigan = str(ws.cell(row=rowitem, column=5).value).strip()        # 승인유예기간
    strRectno = str(ws.cell(row=rowitem, column=6).value).strip()       # 접수번호
    strComp1 = str(ws.cell(row=rowitem, column=7).value).strip()        # 신고기업

    # 제품유형코드를 정제
    strUsecode = strUsecode.replace(' ', '')

    # 날짜 끝에 . 없애기, 승인유예고시 엑셀에 꼭 날짜 마지막에 '.'를 찍어 줌. 이상하네.
    if strGigan:
        if len(strGigan) == 11:
            strGigan = strGigan[0:len(strGigan) - 1]
            strGigan = strGigan.replace('.', '-')

    # 접수번호를 형식에 맞게 수정
    strRectno = 'EAS-N-' + strRectno
   

    # 화학물질명에는 수만은 특수문자...아래는 ' 를 처리한다.
    strMaterial = strMaterial.replace("'", "' || chr(39) || '")

    if strUsecode=='4-선박·수중시설용오염방지제':
        strUsecode = '4-가.선박·수중시설용오염방지제'

    # 유형코드를 조회
    cursor.execute("""SELECT CODE FROM COMTCCMMNDETAILCODE WHERE CODE_ID='BCC003' AND REPLACE(CODE_NM,' ','')='%s'""" % (strUsecode))
    ptcode = cursor.fetchone()
    if ptcode:
        strUsecode = str(ptcode[0])

        ######################################################## 1단계 검토. 승인유예고시가 제대로 이루어졌는지.
        # TN_EBM_COMPANY_LIST에 추가 한다. 먼저 있는지 검토한다.
        cursor.execute("""SELECT COUNT(*) FROM TN_EBM_COMPANY_LIST WHERE RCT_NO='%s' AND (CAS_NO='%s' OR CAS_NO='%s') 
        AND ITEM_CD='%s' AND DEL_YN = 'N'""" % (strRectno, strCasno, strCasno2, strUsecode))
        strExist = cursor.fetchone()

        if strExist:
            strExist = int(strExist[0])

            # TN_EBM_COMPANY에 없으면 추가
            if strExist == 0:
                ws.cell(row=rowitem, column=9, value='TN_EBM_COMPANY 존재하지 않음')
        ########################################################

        # 물질승인신청 테이블 검사를 빠르게 하기 위해
        eListRectno.append(strRectno)
        eListCasno.append(strCasno)
        eListItem.append(strUsecode)
    else:
        ws.cell(row=rowitem, column=8, value='유형코드가 존재하지 않음')
        continue

wb.save(G_ExFilePos + "승인유예고시_전체.xlsx")
wb.close()


# 승인유예고시 테이블을 변수에 담는다.
cursor.execute("""SELECT RCT_NO, CAS_NO, ITEM_CD FROM TN_EBM_COMPANY_LIST WHERE DEL_YN = 'N'""")
for row in cursor:

    strCasno    = row[1]
    
    if strCasno:
        strCasno = strCasno.strip()
        strCasno2 = strCasno
        if strCasno == '27083-27-8' or strCasno == '32289-58-0':
            strCasno = '27083-27-8, 32289-58-0'
        if strCasno == '105827-78-9' or strCasno == '138261-41-3':
            strCasno = '138261-41-3(105827-78-9)'
        if strCasno == '고유번호 없음-6' or strCasno == '7681-52-9':
            strCasno = '7681-52-9'
        if strCasno == '고유번호 없음-7' or strCasno == '7778-54-3':
            strCasno = '7778-54-3'
    else:
        continue
    
    tListRectno.append(row[0])
    tListCasno.append(strCasno)
    tListItem.append(row[2])


######################################################## 2단계 검토. 자진취하가 제대로 이루어졌는지 검토
# TN_EBM_COMPANY_LIST 중심으로 현재 승인유예고시 엑셀과 비교
# TN_EBM_COMPANY_LIST 여기에만 존재하는 삭제되지 않은 물질
rTotList = []
for itIndex in range(0, len(tListRectno)):
    tsRect   = tListRectno[itIndex]
    tsCasno  = tListCasno[itIndex]
    tsItem   = tListItem[itIndex]

    exist = 0
    for exIndex in range(0, len(eListRectno)):
       esRect  = eListRectno[exIndex]
       esCasno = eListCasno[exIndex]
       esItem  = eListItem[exIndex]

    # 같은 것이 있으면 통과
       if tsRect == esRect and tsCasno == esCasno and tsItem == esItem:
           exist = 1
           break
    
    if exist == 0:
        rList = []
        rList.append(tsRect)
        rList.append(tsCasno)
        rList.append(tsItem)
        rTotList.append(rList)

    print("======> " + str(itIndex))

List_To_Excel(G_ExFilePos+"물질승인신청서 검토.xlsx", "검토내용", rTotList)


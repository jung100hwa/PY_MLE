# 전성분 추출에 대한 혼합물질 재계산
# 한상곤 부장이 전성분 추출한 값에 대한 혼합물질일때 전성분 재계산
# 한부장이 먼저 월별 전성분을 추출한 파일에 대해 함량 계산을 다시함
# 즉 한부장이 먼저 전성분을 추출해야 함

import platform
import openpyxl as op
import datetime
import glob
import os

# 현재 실행 위치 담기
G_ExFilePos = os.getcwd()

# 오늘날짜 세팅
now = datetime.datetime.now()
strTime = now.strftime('%Y-%m-%d')

# 플랫폼 담기
G_Platform = platform.system()

G_SplitDef = ''


# 작업항목마다 여기다 적는다.
if G_Platform == 'Windows':
    G_ExFilePos = G_ExFilePos + '\\ODT\\Sample\\'
    G_SplitDef = '\\'
else:
    G_ExFilePos = G_ExFilePos + '/ODT/Sample/'
    G_SplitDef = '/'


# 엑셀파일 열기
os.chdir(G_ExFilePos)
for filename in glob.iglob('**/*.xlsx', recursive=True):

    print("===========>" + filename + "===>start")

    wb = op.load_workbook(G_ExFilePos + filename)
    ws = wb.worksheets[0]

    # 함량재계산 컬럼을 추가
    namecheck = ws.cell(row=1, column=20).value
    if "함량재계산" not in namecheck:
        ws.insert_cols(20)
        ws.cell(row=1, column=20, value="함량재계산")

    strPri = 0          # 주요물질 혼합물질의 원래 합계
    strMajindex = 0     # 주요물질 혼합물질
    strMajtotal = 0     # 주요물질에 혼함물질에 대한 합계

    ws_rowcount = ws.max_row
    for rowitem in range(2,  ws_rowcount + 1):
        strMaterial = ws.cell(row=rowitem, column=16).value    # 물질구분
        strMf = ws.cell(row=rowitem, column=19).value          # 함량
        strMf2 = ws.cell(row=rowitem, column=20).value         # 함량재계산
        strGubun = ws.cell(row=rowitem, column=25).value       # 혼합물질 여부

        if strMaterial == '주요물질' and strGubun == 'Y':
            # 이전 혼합물질에 대한 합계을 재계산하여 넣고 시작한다.
            if strMajindex != 0:
                ws.cell(row=strMajindex, column=20, value=strMajtotal)

            strMajindex = rowitem
            strMajtotal = 0

            # 주요물질 중 혼합물질 값(합계)을 미리 담아둔다.
            try:
                if float(strMf) != 0.0:
                    strPri = strMf
                    ws.cell(row=rowitem, column=20, value=strMf)
                else:
                    strPri = 0
            except:
                strPri = 0
                ws.cell(row=rowitem, column=20, value="0")

        elif strMaterial == '주요물질' and strGubun == 'N':
            if strMajindex != 0:
                ws.cell(row=strMajindex, column=20, value=strMajtotal)

            ws.cell(row=rowitem, column=20, value=strMf)
            strMajindex = 0
            strMajtotal = 0

        else:
            try:
                if float(strMf):
                    strMf2 = float(strPri) * float(strMf) / 100
                    ws.cell(row=rowitem, column=20, value=strMf2)
                    strMajtotal = float(strMajtotal) + float(strMf2)

                    # 파일의 끝이라면
                    if rowitem == ws_rowcount:
                        ws.cell(row=strMajindex, column=20, value=strMajtotal)
            except:
                ws.cell(row=rowitem, column=20, value="0")
                continue

        print('======>' + str(rowitem))

    wb.save(G_ExFilePos + filename)
    wb.close()

    print("===========>" + filename + " ===> end")
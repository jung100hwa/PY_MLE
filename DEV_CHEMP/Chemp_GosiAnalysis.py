# 승인고시된 항목 전체 검사

import cx_Oracle
import openpyxl as op
import time, os

strTime = time.strftime('%Y-%m-%d', time.localtime(time.time()))

# 오라클 연결
# 다이퀘스트
# os.putenv('NLS_LANG', '.UTF8')
# connection = cx_Oracle.connect('chemp/chemp@133.186.171.48:42521/orcl')
# cursor = connection.cursor()

# 기술원 개발서버
# os.putenv('NLS_LANG', '.UTF8')
# connection = cx_Oracle.connect('chemp/chemp@192.168.50.60:1521/orcl')
# cursor = connection.cursor()

# 기술원 운영서버
os.putenv('NLS_LANG', '.UTF8')
connection = cx_Oracle.connect('chemp/chemp!1299@10.10.20.10:1521/chemp')
cursor = connection.cursor()

chdirpath = os.getcwd()
os.chdir(chdirpath)

# ####################################################################################### 승인유예고시 및 자진취하, 공동협의체 검증
# 엑셀파일 열기
wb = op.load_workbook('c:\\work\\PY_MYD_G\\ODT\\승인취하.xlsx')

# 지정한 시트 얻기
ws = wb.worksheets[0]

# 헤더정의
ws.cell(row=1, column=14, value='사전신고_접수번호')
ws.cell(row=1, column=15, value='사전신고_아이디')
ws.cell(row=1, column=16, value='사전신고_아이디구분')
ws.cell(row=1, column=17, value='사전신고_법인등록번호')
ws.cell(row=1, column=18, value='사전신고_사업자등록번호')
ws.cell(row=1, column=19, value='사전신고_신고타입')
ws.cell(row=1, column=20, value='사전신고_제조수입')
ws.cell(row=1, column=21, value='사전신고_국외제조국')
ws.cell(row=1, column=22, value='사전신고_국외제조자')
ws.cell(row=1, column=23, value='사전신고_우편번호')
ws.cell(row=1, column=24, value='사전신고_주소')
ws.cell(row=1, column=25, value='사전신고_상세주소')
ws.cell(row=1, column=26, value='회원정보_법인등록번호')
ws.cell(row=1, column=27, value='회원정보_사업자등록번호')
ws.cell(row=1, column=28, value='사전신고와 회원정보_법인등록번호_일치여부')
ws.cell(row=1, column=29, value='사전신고와 회원정보_사업자등록번호_일치여부')
ws.cell(row=1, column=30, value='물질승인신청계획서_유예대상존재')
ws.cell(row=1, column=31, value='물질승인신청계획서_진행상태')
ws.cell(row=1, column=32, value='공동협의체_존재여부')
ws.cell(row=1, column=33, value='공동협의체_역할')
ws.cell(row=1, column=34, value='공동협의체_가입아이디')
ws.cell(row=1, column=35, value='사전신고와 공동협의체_아이디일치여부')
ws.cell(row=1, column=36, value='공동협의체_유형협의체구성원수')
ws.cell(row=1, column=37, value='공동협의체_물질협의체존재')
ws.cell(row=1, column=38, value='물질협의체가입')
ws.cell(row=1, column=39, value='물질협의체오류판단')
ws.cell(row=1, column=40, value='물질협의체출처')
ws.cell(row=1, column=41, value='물질협의체출처오류판단')
ws.cell(row=1, column=42, value='물질협의체에서의 역할')
ws.cell(row=1, column=43, value='물질협의체출처 추천')
ws.cell(row=1, column=44, value='물질협의체출처 추천 근거')




# ***************************************************************** 1. 고시엑셀정보 읽어 옴
for item in range(2, ws.max_row+1):
    chnm = ws.cell(row=item, column=2).value

    # 화학물질명 세팅
    if chnm:
        chnm = chnm.strip()
    else:
        continue

    # CAS No 세팅
    casno = ws.cell(row=item, column=3).value
    casno2 = ''
    if strCasno:
        strCasno=strCasno.strip()
        strCasno2=strCasno
        if strCasno == '27083-27-8' or strCasno == '32289-58-0':
            strCasno='27083-27-8, 32289-58-0'
        if strCasno == '105827-78-9' or strCasno == '138261-41-3':
            strCasno='138261-41-3(105827-78-9)'
        # 고유번호-8이 고유번호-6으로 변경
        # if strCasno == '7681-52-9' or strCasno == '고유번호 없음-6':
        #     strCasno = '7681-52-9'
        if strCasno == '7778-54-3' or strCasno == '고유번호 없음-7':
            strCasno='7778-54-3'
    else:
        continue

    # 제품유형 세팅
    pt = ws.cell(row=item, column=4).value
    if pt:
        pt = pt.replace(' ','')
    else:
        continue

    # 승인유예기간 세팅. 과학원자료에 의하면 날짜끝에 "." 이 있음. 이것을 제거
    gigan = ws.cell(row=item, column=5).value
    if gigan:
        gigan = gigan.strip()
        if len(gigan) == 11:
            gigan = gigan[0:len(gigan)-1]


    # 접수번호를 가져온다.
    rctno = ws.cell(row=item, column=6).value
    if rctno:
        rctno = rctno.replace(' ','')
        rctno = 'EAS-N-' + rctno
    else:
        continue

    # 승인유예고시인지 취하인지 구분
    AddCancel = ws.cell(row=item, column=13).value
    if AddCancel:
        AddCancel = AddCancel.replace(' ', '')
    else:
        continue

    strwrid = ''    # 신고아이디
    strbizno = ''   # 신고사업자번호
    strrowno = ''   # 신고법인번호
    strSt = ''      # 수입제조 구분
    strNation = ''  # 국외제조국
    strCnnmtn = ''  # 국외제조자
    strBizno = ''   # 회원정보 사업자등록번호
    strRowno = ''   # 회원정보 법인등록번호

# ***************************************************************** 2. 기존살생물물질 신고에 대한 검증
    if rctno:
        # WR_ID를 구한다
        cursor.execute("""SELECT WR_ID, BIZ_NO, ST1, BPLC_NATION, CN_NM_TN, BPLC_ZIP, BPLC_ADDR_TN, BPLC_D_ADDR_TN, APLCNT_TYPE FROM TN_EBMMST WHERE RCPTNO='%s'""" %(rctno))
        wrid = cursor.fetchone()

        if wrid:
            strwrid = wrid[0]
            ws.cell(row=item, column=15, value=strwrid)

            # 물질신고시(사전신고) 사업자등록번호와 법인등록번호를 구한다.
            if wrid[1]:
                strbizrawnum = wrid[1]
                strbizrawnum = strbizrawnum.replace('-','').strip()

                if len(strbizrawnum) == 10: # 사업자등록번호
                    ws.cell(row=item, column=18, value=strbizrawnum)
                    strbizno = strbizrawnum
                else: # 법인등록번호
                    ws.cell(row=item, column=17, value=strbizrawnum)
                    strrowno = strbizrawnum
                    
            else:
                ws.cell(row=item, column=17, value='')
                ws.cell(row=item, column=18, value='')

            if wrid[2]: # 수입제조
                strSt = wrid[2]
                if strSt == 'income':
                    strSt = '수입'
                else:
                    strSt = '제조'
            else:
                strSt = ''

            ws.cell(row=item, column=20, value=strSt)

            if wrid[3]:  # 국외제조국
                strNation = wrid[3]
                ws.cell(row=item, column=21, value=strNation)
            else:
                ws.cell(row=item, column=21, value='')

            if wrid[4]:  # 국외제조자
                strCnnmtn = wrid[4]
                ws.cell(row=item, column=22, value=strCnnmtn)
            else:
                ws.cell(row=item, column=22, value='')

            if wrid[5]:  # 우편번호
                strCnnmtn1 = wrid[5]
                ws.cell(row=item, column=23, value=strCnnmtn1)
            else:
                ws.cell(row=item, column=23, value='')

            if wrid[6]:  # 주소1
                strCnnmtn2 = wrid[6]
                if strSt == '제조':
                    if wrid[7]:
                        strCnnmtn3= wrid[7]
                        strCnnmtn2 = strCnnmtn2 + ' '+ strCnnmtn3
                    else:
                        strCnnmtn2 = wrid[6] 
                ws.cell(row=item, column=24, value=strCnnmtn2)
            else:
                ws.cell(row=item, column=24, value='')

            if wrid[7]:  # 주소2
                strCnnmtn3 = wrid[7]
                if strSt == '수입':
                    ws.cell(row=item, column=25, value=strCnnmtn3)
            else:
                ws.cell(row=item, column=25, value='')

            if wrid[8]:  # 신고타입
                strCnnType = wrid[8]
                if strCnnType == 1:
                    strCnnType = '제조자'
                if strCnnType == 2:
                    strCnnType = '제조자(위탁제조자)'
                if strCnnType == 3:
                    strCnnType = '수입자'
                if strCnnType == 4:
                    strCnnType = '선임자'
                ws.cell(row=item, column=19, value=strCnnType)
            else:
                ws.cell(row=item, column=19, value='')

            # 접수번호를 세팅
            ws.cell(row=item, column=14, value=rctno)

# ***************************************************************** 3. 회원정보에서 구한다.
            cursor.execute("""SELECT BIZRNO, JURIRNO  FROM COMTNENTRPRSMBER WHERE ENTRPRS_MBER_ID='%s'""" %(wrid[0]))
            bizno = cursor.fetchone()

            if bizno:
                if bizno[0]: # 사업자번호 세팅
                    strBizno = str(bizno[0])
                    strBizno = strBizno.replace('-','')

                if bizno[1]:
                    strRowno = str(bizno[1])
                    strRowno = strRowno.replace('-', '')

                ws.cell(row=item, column=16, value='기업회원')

            else: # 기업회원에 없다면 일반회원에서 구한다.
                cursor.execute("""SELECT BIZRNO, JURIRNO  FROM COMTNENTRPRSMBER WHERE ENTRPRS_MBER_ID=(SELECT ENTRPRS_MBER_ID FROM COMTNGNRLMBER WHERE MBER_ID='%s')""" %(wrid[0]))
                bizno = cursor.fetchone()

                if bizno:
                    if bizno[0]:  # 사업자번호 세팅
                        strBizno = str(bizno[0])
                        strBizno = strBizno.replace('-', '')
                    if bizno[1]:
                        strRowno = str(bizno[1])
                        strRowno = strRowno.replace('-', '')
                    ws.cell(row=item, column=16, value='일반회원')

            ws.cell(row=item, column=27, value=strBizno)
            ws.cell(row=item, column=26, value=strRowno)

            # 사전신고와 일반회원의 법인등록번호 일치 여부
            if strrowno == strRowno:
                ws.cell(row=item, column=28, value='일치')
            else:
                ws.cell(row=item, column=28, value='불일치')

            # 사전신고와 일반회원의 사업자등록번호 일치 여부
            if strbizno == strBizno:
                ws.cell(row=item, column=29, value='일치')
            else:
                ws.cell(row=item, column=29, value='불일치')

# ***************************************************************** 4. 공동협의체를 세팅한다.
        # 제품유형코드가 구한다.
        cursor.execute("""SELECT CODE FROM COMTCCMMNDETAILCODE WHERE CODE_ID='BCC003' AND REPLACE(CODE_NM,' ','')='%s'""" % (pt))
        ptcode = cursor.fetchone()

        if not ptcode:
            # ws.cell(row=item, column=24, value=ptcode[0])
            continue
        # else:
            # continue

        # 물질승인신청계획서 제출을 위한 TN_EBM_COMPANY_LIST에 있는지 확인, 승인유예고시와 자진취하를 구분
        # 이부분은 만약에 잘 맞지 않는 다면 일괄업데이트 해버리면 됨
        strPlan = ''
        if AddCancel == '고시':
            cursor.execute("""SELECT RCT_NO FROM TN_EBM_COMPANY_LIST WHERE DEL_YN='N' AND RCT_NO='%s' AND (CAS_NO='%s' OR CAS_NO='%s') AND ITEM_CD='%s'""" %(rctno, casno, casno2, ptcode[0]))
            strPlan = cursor.fetchone()
            if strPlan:
               ws.cell(row=item, column=30, value='고시존재')
            else:
               ws.cell(row=item, column=30, value='고시미존재')
        else:
            cursor.execute("""SELECT RCT_NO FROM TN_EBM_COMPANY_LIST WHERE DEL_YN='Y' AND RCT_NO='%s' AND (CAS_NO='%s' OR CAS_NO='%s') AND ITEM_CD='%s'""" % (rctno, casno, casno2, ptcode[0]))
            strPlan = cursor.fetchone()
            if strPlan:
                ws.cell(row=item, column=30, value='취하존재')
            else:
                ws.cell(row=item, column=30, value='취하미존재')


        # 물질신청계획서 작성 상태
        if strPlan:
            cursor.execute("""SELECT (SELECT CODE_NM FROM COMTCCMMNDETAILCODE WHERE CODE_ID='MIA003' AND CODE=TA.STS) as sts
                FROM TN_NBP_MST TA WHERE TA.EXIST_RCT_NO LIKE '%s'""" %('%'+rctno+'%'))
            strPlanSts = cursor.fetchone()
            if strPlanSts:
                strPlanSts = strPlanSts[0]
                ws.cell(row=item, column=31, value=strPlanSts)

        # MST_ID를 구한다
        strmstid = ''
        cursor.execute("""SELECT MST_ID FROM TN_EBM_COOPERATION_T WHERE CAS_NO='%s' AND USE_CODE='%s'""" % (casno, ptcode[0]))
        mstid = cursor.fetchone()

        if mstid:
            # ws.cell(row=item, column=23, value=mstid[0])
            strmstid = mstid[0]
        else:
            continue

        # 공동협의체 상태
        if strmstid and strBizno:
            # 아래는 공동협의체 아이디와 그룹롤을 구함.
            cursor.execute("""SELECT CREATE_ID, (SELECT CODE_NM FROM COMTCCMMNDETAILCODE WHERE CODE_ID='TEC001' AND CODE=GROUP_ROLE)
                FROM TN_EBM_COOPERATION_MAP_T WHERE MST_ID='%s' AND (BIZ_NO='%s' OR BIZ_NO='%s')""" %(strmstid, strBizno,strbizno))
            ptyn = cursor.fetchone()

            if ptyn:
               ws.cell(row=item, column=32, value='존재')
               strCreateid = str(ptyn[0])                               # 공동협의체 아이디
               strGroupcode = str(ptyn[1])                              # 역할
               ws.cell(row = item, column = 34, value = strCreateid)
               ws.cell(row = item, column = 33, value = strGroupcode)

                # 신고아이디와 공동협의체 아이디의 일치 여부
               if strwrid == strCreateid:
                  ws.cell(row=item, column=35, value='일치')
               else:
                  ws.cell(row=item, column=35, value='불일치')         # 불일치 부분은 대부분 지원단에서 아이디 변경, 권리의무 승계건으로 추정
            else:
               ws.cell(row=item, column=32, value='미존재')
        else:
            ws.cell(row=item, column=32, value='미존재')


        # 유형의체의 개수를 구한다. 구성원수가 1이면서 물질협의체에 추가되었는지 확인하기 위해 필요함
        if strmstid:
            cursor.execute("""SELECT COUNT(*) FROM TN_EBM_COOPERATION_MAP_T WHERE MST_ID='%s'""" %(strmstid))
            strstidcount = cursor.fetchone()
            ws.cell(row=item, column=36, value = strstidcount[0])


        # 물질협의체가 존재하는지 여부, 물질협의체 없는 것은 별도 코드 수행
        if strmstid:
            cursor.execute("""SELECT MST_ID FROM TN_EBM_COOPERATION_T WHERE CAS_NO='%s' AND USE_CODE='000'""" %(casno))
            ptyn = cursor.fetchone()

            if ptyn:
                ws.cell(row=item, column=37, value='존재')
            else:
                ws.cell(row=item, column=37, value='미존재')
                continue


        # 물질협의체에 등록되어 있는지 여부, 신고 사업자번호 또는 회원 사업자번호 둘중에 하나만 있으면 등록으로 체크
        if strmstid and strBizno :
            cursor.execute("""SELECT MST_ID FROM TN_EBM_COOPERATION_MAP_T WHERE MST_ID = (
            SELECT MST_ID FROM TN_EBM_COOPERATION_T WHERE CAS_NO='%s' AND USE_CODE='000') AND (BIZ_NO='%s' OR BIZ_NO='%s')""" %(casno,strBizno,strbizno))
            ptyn = cursor.fetchone()
            if ptyn:
                ws.cell(row=item, column=38, value='등록')
            else:
                ws.cell(row=item, column=38, value='미등록')

            # 물질협의체 등록되어 있는 것이 맞는지 판단여부, casno, 사업자번호를 가진 유형협의체 중에 구성원수가 1이상인것
            cursor.execute("""SELECT MST_ID FROM TN_EBM_COOPERATION_MAP_T WHERE MST_ID IN (SELECT MST_ID FROM TN_EBM_COOPERATION_MAP_T WHERE
            MST_ID IN (SELECT MST_ID FROM TN_EBM_COOPERATION_T WHERE CAS_NO = '%s' AND USE_CODE <> '000') AND (BIZ_NO='%s' OR BIZ_NO='%s')) GROUP BY 
            MST_ID HAVING COUNT(*) > 1 """ %(casno,strBizno,strbizno))

            strMstCountList = []
            for strMstCountItem in cursor:
                strMstCountList.append(strMstCountItem[0])
            strExist = len(strMstCountList)

            if strExist > 0:
                ws.cell(row=item, column=39, value='등록해야 맞음')
            else:
                ws.cell(row=item, column=39, value='동독되면 안됨')

            # 물질협의체에 등록된 유형협의체(대표자, 구성원)이 어디유형에서 왔는지. 여기서 유형협의체가 구성될수 없는 1개 유형을 적으면 안된다.
            if ptyn:
                cursor.execute("""SELECT TARGET_MST_ID, (SELECT CODE_NM FROM COMTCCMMNDETAILCODE WHERE CODE_ID='TEC001' AND CODE=GROUP_ROLE) FROM TN_EBM_COOPERATION_MAP_T WHERE MST_ID = (
                           SELECT MST_ID FROM TN_EBM_COOPERATION_T WHERE CAS_NO='%s' AND USE_CODE='000') AND (BIZ_NO='%s' OR BIZ_NO ='%s')""" % (casno, strBizno,strbizno))
                TargetMst = cursor.fetchone()
                if TargetMst:
                    ws.cell(row=item, column=40, value=TargetMst[0])
                    ws.cell(row=item, column=42, value=TargetMst[1]) #물질협의체에서의 역할을 의미
                else:
                    ws.cell(row=item, column=40, value='존재하지 않음')

                # target_mst_id가 구성원수가 1이상인지 검토, 1이하이면 협의체가 성립이 안되니까 이 유형은 타켓유형이 될 수 없다.
                if TargetMst:
                    cursor.execute("""SELECT COUNT(*) FROM TN_EBM_COOPERATION_MAP_T WHERE MST_ID = '%s'""" % (TargetMst[0]))
                    TargetMstCount = cursor.fetchone()
                    TargetMstCount = int(TargetMstCount[0])

                    if TargetMstCount > 1:
                        ws.cell(row=item, column=41, value='정상')
                    else:
                        ws.cell(row=item, column=41, value='비정상')

            # 물질협의체 출처를 일단 대표자를 아래의 순서로 추천함
            # 1.유형협의체의 대표자

            strMstGrouploeExist = ''
            if strExist > 0:

                # 1.유형협의체의 대표자
                for item22 in strMstCountList:
                    cursor.execute("""SELECT MST_ID FROM TN_EBM_COOPERATION_MAP_T WHERE MST_ID = '%s' 
                    AND GROUP_ROLE = '001' AND (BIZ_NO='%s' OR BIZ_NO ='%s')""" % (item22, strBizno, strbizno))
                    strMstGrouploeExist = cursor.fetchone()

                    if strMstGrouploeExist:
                        ws.cell(row=item, column=43, value=strMstGrouploeExist[0])
                        ws.cell(row=item, column=44, value='유형협의체 대표')
                        break

                # 2.유형협의체 대표자가 없으면 1.살균제부터 승인유예기간이 짧은 순서대로 함. 사실 이부분은 할 필요가 없음.
                # 물질협의체 대표자는 무조건 유형 협의체 대표자가 되어야 함
                if not strMstGrouploeExist:
                    for item22 in strMstCountList:
                        cursor.execute("""SELECT MST_ID FROM TN_EBM_COOPERATION_MAP_T WHERE MST_ID = '%s'
                         AND (BIZ_NO='%s' OR BIZ_NO ='%s')""" % (item22, strBizno, strbizno))
                        strMstGrouploeExist2 = cursor.fetchone()

                        if strMstGrouploeExist2:
                            ws.cell(row=item, column=43, value=strMstGrouploeExist2[0])
                            ws.cell(row=item, column=44, value='회원수가 2개 이상인것')
                            break
    
    print('순번 = %s' % (item))

# 엑셀 파일 저장
wb.save('c:\\work\\PY_MYD_G\\ODT\\승인취하.xlsx')
wb.close()


# ####################################################################################### 물질협의체가 없는 것 조회해서 무조건 하나는 만든다.
# notExistMn = {}
# cursor.execute("""SELECT DISTINCT CAS_NO, MATERIAL_NM FROM TN_EBM_COOPERATION_T WHERE CAS_NO NOT IN(SELECT CAS_NO FROM TN_EBM_COOPERATION_T WHERE CATEGORY = 'M')""")
# for name in cursor:
#     notExistMn[str(name[0])] = str(name[1])
#
# for item in notExistMn.keys():
#     strCasno = item
#     strMaterial = notExistMn[strCasno]
#
#     # 여기서 한번 찍어 본다.
#     print(strCasno, strMaterial)
#
#     strMaterial = strMaterial.replace("'","' || chr(39) || '")
#
#     strSeq = """'TEC_' || LPAD(SEQ_EBM_COOPERATION_01.NEXTVAL,7,'0')"""
#
#     cursor.execute("""INSERT INTO TN_EBM_COOPERATION_T(MST_ID, CAS_NO, MATERIAL_NM, USE_CODE, PARTICIPANT_COUNT, STS, CATEGORY,CREATE_DATE,
#     CREATE_ID, UPDATE_DATE, UPDATE_ID) VALUES(%s,'%s','%s','%s','%s','%s','%s',to_date('%s','YYYY-MM-DD'),'%s',to_date('%s','YYYY-MM-DD'),'%s')"""
#                    % (strSeq, strCasno, strMaterial, '000', '0', '001', 'M', strTime, 'SYSTEM', strTime, 'SYSTEM'))
#     connection.commit()
# print("물질협의체가 존재하지 않는 것 생성 완료")


# ####################################################################################### 구성원수 일괄업데이트(로우수로 한다지만 꼭 있어야 할 듯)
# cursor.execute("""UPDATE TN_EBM_COOPERATION_T A SET A.PARTICIPANT_COUNT = (SELECT COUNT(*) FROM TN_EBM_COOPERATION_MAP_T WHERE MST_ID = A.MST_ID)""")
# connection.commit()
# print("구성원수가 업데이트 완료")


# ####################################################################################### 만약에 탈퇴한 인원수를 제외한 구성원수 업데이트 
# cursor.execute("""UPDATE TN_EBM_COOPERATION_T TA SET PARTICIPANT_COUNT =(SELECT COUNT(*) FROM TN_EBM_COOPERATION_MAP_T WHERE 
# 	((GROUP_ROLE, APPROVE_YN) NOT IN (('008', 'Y'))) AND MST_ID = TA.MST_ID)""")
# connection.commit()
# print("탈퇴한 업체를 제외한 구성원수가 업데이트 완료")


# ####################################################################################### 구성원수가 0인것 삭제
# cursor.execute("""DELETE FROM TN_EBM_COOPERATION_T WHERE PARTICIPANT_COUNT=0""")
# connection.commit()
# print("물질협의체 구성원수가 0인것 삭제")


# ####################################################################################### 공동협의체 테이블에 전체 임포트시 수행, 일반 SQL툴로 안 될때
# wb = op.load_workbook('c:\\work\\PY_MYD_G\\IDT\\20210205-173843.xlsx')
# ws = wb.worksheets[0]
# strTime = time.strftime('%Y-%m-%d', time.localtime(time.time()))

##### TN_EBM_COOPERATION_T 전체 임포트
# for ritem in range(2, ws.max_row + 1):
#     mstid = ws.cell(row=ritem, column=1).value
#     casno = ws.cell(row=ritem, column=2).value
#     mn = ws.cell(row=ritem, column=3).value
#     mn = mn.replace("'", "' || chr(39) || '")
#     us = ws.cell(row=ritem, column=4).value
#     pc = ws.cell(row=ritem, column=5).value
#     st = ws.cell(row=ritem, column=6).value
#     cg = ws.cell(row=ritem, column=7).value
#     ci = ws.cell(row=ritem, column=9).value
#     ui = ws.cell(row=ritem, column=11).value
#
#     cursor.execute("""INSERT INTO TN_EBM_COOPERATION_T VALUES('%s','%s','%s','%s','%s','%s','%s',to_date('%s','YYYY-MM-DD'),'%s',to_date('%s','YYYY-MM-DD'),'%s')"""
#                 % (mstid, casno, mn, us, pc, st, cg, strTime, ci, strTime, ui))
#     connection.commit()
#     print(ritem)
##### TN_EBM_COOPERATION_MAP_T 전체 임포트
# for ritem in range(2, ws.max_row + 1):
#     mstid = ws.cell(row=ritem, column=1).value
#     bz = ws.cell(row=ritem, column=2).value
#     gr = ws.cell(row=ritem, column=3).value
#     ay = ws.cell(row=ritem, column=4).value
#     uy = ws.cell(row=ritem, column=5).value
#     iy = ws.cell(row=ritem, column=6).value
#     ry = ws.cell(row=ritem, column=7).value
#     hy = ws.cell(row=ritem, column=8).value
#     cd = ws.cell(row=ritem, column=9).value
#     ci = ws.cell(row=ritem, column=10).value
#     ud = ws.cell(row=ritem, column=11).value
#     ui = ws.cell(row=ritem, column=12).value
#     ay2 = ws.cell(row=ritem, column=13).value
#     ti = ws.cell(row=ritem, column=14).value
#     ay3 = ws.cell(row=ritem, column=15).value
#
#     cursor.execute("""INSERT INTO TN_EBM_COOPERATION_MAP_T(MST_ID, BIZ_NO, GROUP_ROLE, AGREE_YN, USE_YN, INDIVISUAL_SUBMIT_YN,
#                REQUEST_PLAN_YN, HOPE_REPRESENT_YN, CREATE_DATE, CREATE_ID, UPDATE_DATE, UPDATE_ID, APPROVE_YN, TARGET_MST_ID, AGENT_YN)
#                VALUES('%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', to_date('%s','YYYY-MM-DD'), '%s', to_date('%s','YYYY-MM-DD'), '%s', '%s','%s', '%s')
#                """ % (mstid, bz, gr, ay, uy, iy, ry, hy, strTime, ci, strTime, ui, ay2,ti,ay3))
#     connection.commit()
#     print(ritem)

# 수입요건면제(권혜림 연구원)
# 권혜림 연구원 자유판매증명원 통계 뽑는 것

import datetime
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import random
import platform
import openpyxl as op
import cx_Oracle
import os

# 기술원 운영서버
os.putenv('NLS_LANG', '.UTF8')
connection = cx_Oracle.connect('chemp/chemp!1299@10.10.20.10:1521/chemp')
cursor  = connection.cursor()

# 현재 실행 위치 담기
G_ExFilePos = os.getcwd()

# 오늘날짜 세팅
now = datetime.datetime.now()
strTime = now.strftime('%Y-%m-%d')

# 플랫폼 담기
G_Platform = platform.system()

G_SplitDef = ''

# 작업항목마다 여기다 적는다.
if G_Platform == 'Windows':
    G_ExFilePos = G_ExFilePos + '\\All\\'
    G_SplitDef = '\\'
else:
    G_ExFilePos = G_ExFilePos + '/All/'
    G_SplitDef = '/'

wb = op.load_workbook(G_ExFilePos + "권혜림자유판매증명원.xlsx")
ws = wb.worksheets[0]

cursor.execute("""SELECT 
A.CERT_SERIAL_NO "증명서 일련번호(신고번호)",
A.CREATE_DATE "제출일",
A.REGIST_DATE "접수일",
A.RCT_NO "접수번호",
A.STS_DATE "발급일",
GET_CODE_NM('STE001', A.STS) "상태",
A.CN_NM "상호_명",--상호_명
REPLACE(A.BIZ_NO,'-','') "사업자번호", --사업자번호
A.RPS_NM "대표자명", --대표자명
A.APLCNT_NM "신고자성명", --신고자성명
A.APLCNT_TELNO_ARNO || '-' || A.APLCNT_TELNO_MIDNO || '-' || A.APLCNT_TELNO_ENDNO "신고자전화번호", --신고자_전화번호
A.APLCNT_EMAIL_ID || '@' || A.APLCNT_EMAIL_DOMN "담당자 이메일", --담당자 이메일
'(' || A.BPLC_ZIP || ') ' || A.BPLC_ADDR || ' ' || A.BPLC_D_ADDR "소재지 주소", --소재지 주소
A.BPLC_TELNO_ARNO || '-' || A.BPLC_TELNO_MIDNO || '-' || A.BPLC_TELNO_ENDNO "소재지전화번호", --소재지전화번호
A.BPLC_FAX_ARNO || '-' || A.BPLC_FAX_MIDNO || '-' || A.BPLC_FAX_ENDNO "소재지팩스번호", --소재지팩스번호
A.TARGET_NM "제품명",--제품명
A.TARGET_NM_ENG "제품명영문", --제품명영문
A.TARGET_TYPE "품목",--품목또는유형
A.TARGET_TYPE_ENG "품목(영문)", --품목또는유형_영문
A.TARGET_NO "신고번호/승인번호", --신고번호/승인번호
A.TARGET_ENDDT "유효기간",
A.MAKE_COMPANY_NM "제조회사명", --제조회사명
A.MAKE_COMPANY_NM_ENG "제조회사명(영문)", --제조회사명_영문
A.MAKE_COMPANY_ADDRESS1 || ' '|| A.MAKE_COMPANY_ADDRESS2 "제조회사주소", --제조회사주소
A.MAKE_COMPANY_ADDRESS1_ENG "제조회사주소(영문)",--제조회사주소영문
A.OUTCOME_NM "수출회사명", --수출회사명
A.OUTCOME_NM_ENG "수출회사명(영문)", --수출회사명영문
A.OUTCOME_ADDRESS1 || ' ' || A.OUTCOME_ADDRESS2 "수출회사주소", --수출회사주소
A.OUTCOME_ADDRESS1_ENG "수출회사주소(영문)", --수출회사주소영문
--A.INCOME_NM "수입업체명", --수입업체명
--A.INCOME_ADDRESS1 || ' ' || A.INCOME_ADDRESS2 "수입업체주소", --수입업체주소
A.INCOME_NM_ENG "수입업체명(영문)", --수입업체영문명
--A.INCOME_COUNTRY, --수입국
GET_CODE_NM('CPC003', A.INCOME_COUNTRY) "수입국",
A.INCOME_ADDRESS1_ENG "수입업체주소(영문)" --수입업체주소영문명
FROM tn_fsc_mst A
WHERE  A.STS != 'ST001'
AND A.target_kind IN ('001')
-- AND TO_CHAR(A.STS_DATE,'YYYY-MM-DD') LIKE '2021%'
ORDER BY 1
    """)

rowcount = 3

for row in cursor:
    print(row)
    strListValue = []
    for col in range(0,len(cursor.description)):
        if col == 0:
            strNo=str(row[col])

        va = str(row[col])
        if '' in va:
            va = va.replace('', ' ')
        if '' in va:
            va = va.replace('', ' ')
        if '' in va:
            va = va.replace('', ' ')
        if va == 'None':
            va=''
        ws.cell(row=rowcount, column=col+1, value=va)

    rowcount = rowcount + 1

wb.save(G_ExFilePos + "권혜림자유판매증명원.xlsx")
wb.close()
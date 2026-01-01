# 권혜림 연구원 위해 우려제품 넣기
# 지금은 거의 사용하지 않는다. 법에서 위해 우려제품 사용하지 않음

import platform
import openpyxl as op
import cx_Oracle
import datetime
import win32com.client as win32
import pandas as pd
from functools import reduce
import Hwp_Read_PL as phwp
import os, re
import shutil

# import re, csv
# import statsmodels.formula.api as smf
# import matplotlib.pyplot as plt
# import numpy as np
# import Excel_Basic_PL as EBP

# 기술원 개발서버
# os.putenv('NLS_LANG', '.UTF8')
# connection = cx_Oracle.connect('chemp/chemp@192.168.50.60:1521/orcl')
# cursor = connection.cursor()

# 기술원 운영서버
# os.putenv('NLS_LANG', '.UTF8')
# connection = cx_Oracle.connect('chemp/chemp!1299@10.10.20.10:1521/chemp')
# cursor = connection.cursor()


# 현재 실행 위치 담기
G_ExFilePos = os.getcwd()

# 오늘날짜 세팅
now = datetime.datetime.now()

# 플랫폼 담기
G_Platform = platform.system()

G_FilePdfPos = ''  # 시험성적서 파일 경로
G_AptSample = ''  # 전국아파트 샘플 파일
G_SplitDef = ''

# 작업항목마다 여기다 적는다.
if G_Platform == 'Windows':
    G_FilePdfPos = G_ExFilePos + '\\IDT\\PDF\\'
    G_AptSample = G_ExFilePos + '\\ODT\\Sample\\'
    G_SplitDef = '\\'
else:
    G_FilePdfPos = G_ExFilePos + '/IDT/PDF/'
    G_AptSample = G_ExFilePos + '/ODT/Sample/'
    G_SplitDef = '/'


wb = op.load_workbook(G_AptSample + "위해우려제품.xlsx")
ws = wb.worksheets[0]

for rowitem in range(1, ws.max_row + 1):
    strnewmode = str(ws.cell(row=rowitem, column=1).value)  # 신규인지 변경인지 확인
    strmstidold = str(ws.cell(row=rowitem, column=2).value)  # 변경전 자가검사번호
    strmstidnew = str(ws.cell(row=rowitem, column=11).value)  # 변경후 자가검사번호
    strmpdname = str(ws.cell(row=rowitem, column=12).value)  # 제품명
    strregdate = str(ws.cell(row=rowitem, column=13).value)  # 최초발급일
    stryugigan = str(ws.cell(row=rowitem, column=14).value)  # 유효기간
    strcompnm = str(ws.cell(row=rowitem, column=15).value)  # 회사명
    strbiznm = str(ws.cell(row=rowitem, column=16).value)  # 사업자번호
    strceo = str(ws.cell(row=rowitem, column=17).value)  # 대표이사
    strstaff = str(ws.cell(row=rowitem, column=18).value)  # 담당자
    strphone = str(ws.cell(row=rowitem, column=19).value)  # 담당자 연락처
    stremail = str(ws.cell(row=rowitem, column=20).value)  # 이메일

    strnewmode = strnewmode.strip()
    strmstidold = strmstidold.strip()
    strmstidnew = strmstidnew.strip()
    strmpdname = strmpdname.strip()
    strregdate = strregdate.strip()
    stryugigan = stryugigan.strip()
    strcompnm = strcompnm.strip()
    strbiznm = strbiznm.strip()
    strceo = strceo.strip()
    strstaff = strstaff.strip()
    strphone = strphone.strip()
    stremail = stremail.strip()

    strregdate = strregdate[0:10]  # 등록날짜 날짜 10자리로 나누기
    stryugigan = stryugigan[0:10]  # 유효기간 날짜 10자리로 나누기
    strphoneorg = strphone
    strphone = strphone.split('-')  # 전화번호 나누기
    stremailorg = stremail
    stremail = stremail.split('@')  # 이메일 나누기

    if strnewmode == '변경':
        cursor.execute("""UPDATE TN_IRN_MST_BAK SET MST_ID='%s', PDT_NAME='%s', REGDATE='%s', BIZ_NO='%s', 
            CMPNYNM='%s', CXFC='%s', MBERNM='%s',CBFRSTCERLTELNO='%s',CBMDDLCERLTELNO='%s',CBLASTCERLTELNO='%s',CBEMAILID='%s',CBEMAILDOMNNM='%s', 
            EFFTPRD_TRM_DY='%s', BIZRNO='%s' WHERE MST_ID = '%s'""" % (strmstidnew, strmpdname, strregdate,
                                                                        strbiznm, strcompnm, strceo, strstaff,
                                                                        strphone[0], strphone[1], strphone[2],
                                                                        stremail[0], stremail[1], stryugigan,
                                                                        strbiznm, strmstidold))
        connection.commit()

        cursor.execute("""UPDATE TN_HARM_MST SET SELF_CHECK_NO = '%s',PRODUCT_NAME = '%s', REGDATE = '%s', COMPANY_NAME='%s',
            INCOME_COMPANY_NAME = '%s', BIZ_NO = '%s', CEO_NAME = '%s',CHARGE_NAME = '%s', CHARGE_PHONE = '%s',EMAIL = '%s',NEW_CHANGE='변경' WHERE SELF_CHECK_NO='%s'"""
                        % (strmstidnew, strmpdname, strregdate, strcompnm, strcompnm, strbiznm, strceo, strstaff,
                            strphoneorg, stremailorg, strmstidold))
        connection.commit()

        cursor.execute("""UPDATE TN_IRN_MST_ETC SET MST_ID = '%s',EF_DATE = '%s' WHERE MST_ID='%s'"""
                        % (strmstidnew, stryugigan, strmstidold))

        connection.commit()

    else:
        # 위해우려제품 본 테이블 인서트
        cursor.execute(
            """INSERT INTO TN_HARM_MST(NO, SELF_CHECK_NO) VALUES((SELECT MAX(NO)+1 FROM TN_HARM_MST),'%s' )""" % (
                strmstidnew))
        connection.commit()

        # 위해우려제품 업데이트(변경과는 달리 기존 SELF_CHECK_NO는 하지 않음)
        cursor.execute("""UPDATE TN_HARM_MST SET PRODUCT_NAME = '%s', REGDATE = '%s', COMPANY_NAME='%s',
            INCOME_COMPANY_NAME = '%s', BIZ_NO = '%s', CEO_NAME = '%s',CHARGE_NAME = '%s', CHARGE_PHONE = '%s',EMAIL = '%s',NEW_CHANGE='신규' WHERE SELF_CHECK_NO='%s'"""
                        % (strmpdname, strregdate, strcompnm, strcompnm, strbiznm, strceo, strstaff, strphoneorg,
                            stremailorg, strmstidnew))
        connection.commit()

        # 수입요건확인번호 테이블에 인서트
        cursor.execute("""INSERT INTO TN_IRN_MST_BAK(NO, TYPE, MST_ID, SAVE_YN, SEND_YN, EST_NO) 
            VALUES(TN_IRN_MST_SQ04.NEXTVAL,'1', '%s','N','N','1')""" % (strmstidnew))
        connection.commit()

        # 인서트한 no를 추출
        cursor.execute("""SELECT NO FROM TN_IRN_MST_BAK WHERE MST_ID='%s'""" % (strmstidnew))
        strno = cursor.fetchone()
        strno = str(strno[0])

        cursor.execute("""UPDATE TN_IRN_MST_BAK SET INCOME_REQUEST_NO_SEQ=REPLACE(TO_CHAR(TN_IRN_MST_SQ05.NEXTVAL,'0000000'),' ','')
            WHERE TYPE='1' AND NO='%s'""" % (strno))
        connection.commit()

        cursor.execute("""UPDATE TN_IRN_MST_BAK
            SET INCOME_REQUEST_NO = '2' || '423' || '00' || SUBSTR(TO_CHAR(SYSDATE, 'YYYY'),3,2) || INCOME_REQUEST_NO_SEQ || MOD((11 - MOD((4*9) + (2*8) + (3*7) + (0*6) + (0*5) 
            + (SUBSTR(SUBSTR(TO_CHAR(SYSDATE, 'YYYY'),3,2),1,1) * 4)
            + (SUBSTR(SUBSTR(TO_CHAR(SYSDATE, 'YYYY'),3,2),2,1) * 3)
            + (SUBSTR(INCOME_REQUEST_NO_SEQ, 1,1) * 9)
            + (SUBSTR(INCOME_REQUEST_NO_SEQ, 2,1) * 8)
            + (SUBSTR(INCOME_REQUEST_NO_SEQ, 3,1) * 7)
            + (SUBSTR(INCOME_REQUEST_NO_SEQ, 4,1) * 6)
            + (SUBSTR(INCOME_REQUEST_NO_SEQ, 5,1) * 5)
            + (SUBSTR(INCOME_REQUEST_NO_SEQ, 6,1) * 4)
            + (SUBSTR(INCOME_REQUEST_NO_SEQ, 7,1) * 3),11)),10)
            WHERE TYPE='1' AND NO='%s'""" % (strno))
        connection.commit()

        cursor.execute("""UPDATE TN_IRN_MST_BAK SET PDT_NAME='%s', REGDATE='%s', BIZ_NO='%s', 
            CMPNYNM='%s', CXFC='%s', MBERNM='%s',CBFRSTCERLTELNO='%s',CBMDDLCERLTELNO='%s',CBLASTCERLTELNO='%s',CBEMAILID='%s',CBEMAILDOMNNM='%s', 
            EFFTPRD_TRM_DY='%s', BIZRNO='%s' WHERE MST_ID = '%s'""" % (strmpdname, strregdate,
                                                                        strbiznm, strcompnm, strceo, strstaff,
                                                                        strphone[0], strphone[1], strphone[2],
                                                                        stremail[0], stremail[1], stryugigan,
                                                                        strbiznm, strmstidnew))
        connection.commit()

        cursor.execute(
            """INSERT INTO TN_IRN_MST_ETC(MST_ID,EF_DATE) VALUES('%s','%s')""" % (strmstidnew, stryugigan))
        connection.commit()
    print('자거검사번호=%s' % (strmstidnew))

wb.close()

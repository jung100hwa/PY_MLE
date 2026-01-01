# 승인유예고시 및 자진취하건 업데이트
# 승인유예고시 초기데이터를 구축하는 것으로 한번만 수행하면 됨
# 이건은 아마도 한번만 수행될 것임
# 이후부터는 기존 살생물물질 신고에서 승인유예고시 하면 이 테이블로 인서트
# 이 테이블을 보는 화면에서 자진취하가 이루어짐

import platform
import openpyxl as op
import cx_Oracle
import datetime
import os


# 기술원 개발서버
os.putenv('NLS_LANG', '.UTF8')
connection = cx_Oracle.connect('chemp/chemp@192.168.50.60:1521/orcl')
cursor = connection.cursor()

# 기술원 운영서버
# os.putenv('NLS_LANG', '.UTF8')
# connection = cx_Oracle.connect('chemp/chemp!1299@10.10.20.10:1521/chemp')
# cursor = connection.cursor()


# 로컬설버
# os.putenv('NLS_LANG', '.UTF8')
# connection = cx_Oracle.connect('scott/tiger@127.0.0.1:1521/XE')
# cursor = connection.cursor()


# 현재 실행 위치 담기
G_ExFilePos = os.getcwd()

# 오늘날짜 세팅
now = datetime.datetime.now()
strTime = now.strftime('%Y-%m-%d')

# 플랫폼 담기
G_Platform = platform.system()

G_SplitDef = ''

# 작업항목마다 여기다 적는다.
if G_Platform == 'Windows':
    G_ExFilePos = G_ExFilePos + '\\GODT\\'
    G_SplitDef = '\\'
else:
    G_ExFilePos = G_ExFilePos + '/GODT/'
    G_SplitDef = '/'


# 엑셀파일 열기
wb = op.load_workbook(G_ExFilePos + "gc2.xlsx")
ws = wb.worksheets[0]


for rowitem in range(3, ws.max_row + 1):
    print("=======>" + str(rowitem))
    # 변수정의
    yun = ''                    # 연번
    m_nm = ''                   # 물질명
    casno = ''                  # casno
    itemcd = ''                 # 제품유형
    gd = ''                     # 승인유예기간
    rctno = ''                  # 기존살생물물질 신고 접수번호
    pig = ''                    # 제조수입 구분
    pdg = ''                    # 물질, 제품, 처리제품 구분
    pun = ''                    # 신고기업명
    org = ''                    # 선임구분(0-일반, 1-선임)
    mkn = ''                    # 제조자
    gosid = ''                  # 고시일자
    pland = ''                  # 물질승인신청 기한(신고하고 1년)
    gcg = ''                    # 고시인지 취하인지(0-고시, 1-취하)
    ong = ''                    # 기존인지 신규인지(0-이미고시된것, 1-신규고시한것)
    ci = ''                     # 처리자
    cd = ''                     # 처리일자
    ui = ''                     # 수정자
    ud = ''                     # 수정일자

    yun         = str(ws.cell(row=rowitem, column=1).value).strip()
    m_nm        = str(ws.cell(row=rowitem, column=2).value).strip()
    casno       = str(ws.cell(row=rowitem, column=3).value).strip()
    itemcd      = str(ws.cell(row=rowitem, column=4).value).strip()
    gd          = str(ws.cell(row=rowitem, column=5).value).strip()
    rctno       = str(ws.cell(row=rowitem, column=6).value).strip()
    pig         = str(ws.cell(row=rowitem, column=7).value).strip()
    pdg         = str(ws.cell(row=rowitem, column=8).value).strip()
    pun         = str(ws.cell(row=rowitem, column=9).value).strip()
    org         = str(ws.cell(row=rowitem, column=10).value).strip()
    mkn         = str(ws.cell(row=rowitem, column=11).value).strip()
    gosid       = str(ws.cell(row=rowitem, column=12).value).strip()
    pland       = str(ws.cell(row=rowitem, column=13).value).strip()
    gcg         = str(ws.cell(row=rowitem, column=14).value).strip()
    ong         = str(ws.cell(row=rowitem, column=15).value).strip()
    ci          = "mgr"
    cd          = strTime
    ui          = "mgr"
    ud          = strTime

    # 각 변수값 조정
    m_nm = m_nm.replace("'", "' || chr(39) || '")
    mkn = mkn.replace("'", "' || chr(39) || '")
    pun = pun.replace("'", "' || chr(39) || '")


    itemcd = itemcd.replace(' ', '')
    if itemcd == '4-선박·수중시설용오염방지제':
        itemcd = '4-가.선박·수중시설용오염방지제'

    # 유형코드를 조회
    cursor.execute("""SELECT CODE FROM COMTCCMMNDETAILCODE WHERE CODE_ID='BCC003' AND REPLACE(CODE_NM,' ','')='%s'""" % (itemcd))
    ptcode = cursor.fetchone()
    if ptcode:
        itemcd = str(ptcode[0])
    else:
        itemcd = '999'

    if gd != "None":
        if len(gd) == 11:
            gd = gd[0:len(gd) - 1]
            gd = gd.replace('.', '-')

    rctno = 'EAS-N-' + rctno

    if org != 'None':
        pun = org # 선임이면 선임자가 신고자
        org = '1'
    else:
        org = '0'

    if gosid != "None":
        gosid = gosid[0:10]
    else:
        gosid = strTime

    if pland !='별도공지':
        pland = pland[0:10]

    seq = "(SELECT NVL(MAX(SEQ)+1,1) FROM TN_EBM_GOSICANCEL_LIST)"

    try:
        cursor.execute("""INSERT INTO TN_EBM_GOSICANCEL_LIST(SEQ, YUN, MATERIAL_NAME ,CAS_NO, ITEM_CD, GRACE_DATE, 
        RCT_NO, PI_GUBUN, PD_GUBUN, PU_NAME, OR_GUBUN, MK_NAME, GOSI_DATE, PLAN_DATE, GOCANEL_GUBUN, OLDNEW_GUBUN, 
        CREATE_ID, CREATE_DATE, UPDATE_ID, UPDATE_DATE) VALUES(%s,'%s', '%s','%s','%s','%s', '%s','%s','%s',
        '%s', '%s','%s','%s','%s', '%s','%s','%s', '%s','%s','%s')""" % (seq,yun,m_nm,casno,itemcd,gd, rctno,pig,pdg,pun,org,mkn,gosid,pland,gcg,ong,ci,cd,ui,ud))
        connection.commit()
    except:
        print(rowitem)
        break

wb.save(G_ExFilePos + "gc2.xlsx")
wb.close()

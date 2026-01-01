# 현재 데이터베이스 국제평가대상물질 업데이트

import platform
import openpyxl as op
import cx_Oracle
import datetime
import os


# 기술원 개발서버
# os.putenv('NLS_LANG', '.UTF8')
# connection = cx_Oracle.connect('chemp/chemp@192.168.50.60:1521/orcl')
# cursor = connection.cursor()

# 기술원 운영서버
os.putenv('NLS_LANG', '.UTF8')
connection = cx_Oracle.connect('chemp/chemp!1299@10.10.20.10:1521/chemp')
cursor = connection.cursor()


# 로컬설버
# os.putenv('NLS_LANG', '.UTF8')
# connection = cx_Oracle.connect('scott/tiger@127.0.0.1:1521/XE')
# cursor = connection.cursor()


# 현재 실행 위치 담기
G_ExFilePos = os.getcwd()

# 오늘날짜 세팅
now = datetime.datetime.now()
strTime = now.strftime('%Y-%m-%d')

# 플랫폼 담기
G_Platform = platform.system()


G_SplitDef = ''

# 작업항목마다 여기다 적는다.
if G_Platform == 'Windows':
    G_ExFilePos = G_ExFilePos + '\\GODT\\'
    G_SplitDef = '\\'
else:
    G_ExFilePos = G_ExFilePos + '/GODT/'
    G_SplitDef = '/'


# 엑셀파일 열기
wb = op.load_workbook(G_ExFilePos + "국평물질.xlsx")
ws = wb.worksheets[0]

# 변수정의
strMaterial     = ''           # 화학물질명
strCasno        = ''           # CAS_NO


for rowitem in range(2, ws.max_row + 1):
    strMaterial = str(ws.cell(row=rowitem, column=1).value).strip()             # 물질명
    strCasno    = str(ws.cell(row=rowitem, column=2).value).strip()                # CAS No

    print('%s====>%s' % (rowitem-1, strCasno))

    if strCasno:
        cursor.execute("""UPDATE TN_EBM_COOPERATION_MAP_T SET P_GUBUN='Y' WHERE MST_ID IN (SELECT MST_ID FROM TN_EBM_COOPERATION_T WHERE CAS_NO='%s')""" %(strCasno))
        connection.commit()

wb.save(G_ExFilePos + "국평물질.xlsx")
wb.close()
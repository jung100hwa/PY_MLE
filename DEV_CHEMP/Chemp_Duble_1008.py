# 고유아이디 변경 건

import platform
import openpyxl as op
import cx_Oracle
import datetime
import os


# 기술원 개발서버
# os.putenv('NLS_LANG', '.UTF8')
# connection = cx_Oracle.connect('chemp/chemp@192.168.50.60:1521/orcl')
# cursor = connection.cursor()

# 기술원 운영서버
os.putenv('NLS_LANG', '.UTF8')
connection = cx_Oracle.connect('chemp/chemp!1299@10.10.20.10:1521/chemp')
cursor = connection.cursor()


# 로컬설버
# os.putenv('NLS_LANG', '.UTF8')
# connection = cx_Oracle.connect('scott/tiger@127.0.0.1:1521/XE')
# cursor = connection.cursor()


# 현재 실행 위치 담기
G_ExFilePos = os.getcwd()

# 오늘날짜 세팅
now = datetime.datetime.now()
strTime = now.strftime('%Y-%m-%d')

# 플랫폼 담기
G_Platform = platform.system()


G_SplitDef = ''

# 작업항목마다 여기다 적는다.
if G_Platform == 'Windows':
    G_ExFilePos = G_ExFilePos + '\\GODT\\'
    G_SplitDef = '\\'
else:
    G_ExFilePos = G_ExFilePos + '/GODT/'
    G_SplitDef = '/'


# 엑셀파일 열기
wb = op.load_workbook(G_ExFilePos + "고유번호변경_1021.xlsx")
ws = wb.worksheets[0]

for rowitem in range(1, ws.max_row + 1):
    # 변수정의
    strCnm   = ''         # 물질명
    strCasno = ''         # 변경될 고유번호

    strCnm     = str(ws.cell(row=rowitem, column=1).value).strip()                   # 물질명
    strCasno   = str(ws.cell(row=rowitem, column=3).value).strip()                   # 변경될 고유번호


    print('=============> %s' % (rowitem))
    
    # 물질승인신청테이블에 존재하는지 검토
    cursor.execute("""SELECT MTTR_NM FROM TN_NBMMST WHERE MTTR_NM='%s'""" % (strCnm))
    sss = cursor.fetchone()
    if sss:
        strCnm = str(sss[0])
        ws.cell(row=rowitem, column=4, value='존재함')

    # 승인유예고시테이블
    cursor.execute("""SELECT MATERIAL_NAME FROM TN_EBM_COMPANY_LIST WHERE MATERIAL_NAME='%s'""" % (strCnm))
    sss = cursor.fetchone()
    if sss:
        strCnm = str(sss[0])
        ws.cell(row=rowitem, column=5, value='존재함')

    # 물질승인신청계획서
    cursor.execute("""SELECT MATERIAL_NAME FROM TN_NBP_MST WHERE MATERIAL_NAME='%s'""" % (strCnm))
    sss = cursor.fetchone()
    if sss:
        strCnm = str(sss[0])
        ws.cell(row=rowitem, column=6, value='존재함')

    # 공동협의체
    cursor.execute("""SELECT MATERIAL_NM FROM TN_EBM_COOPERATION_T WHERE MATERIAL_NM='%s'""" % (strCnm))
    sss = cursor.fetchone()
    if sss:
        strCnm = str(sss[0])
        ws.cell(row=rowitem, column=7, value='존재함')

    # 선임테이블
    cursor.execute("""SELECT MATERIAL_NAME FROM TN_APD_MST WHERE MATERIAL_NAME='%s'""" % (strCnm))
    sss = cursor.fetchone()
    if sss:
        strCnm = str(sss[0])
        ws.cell(row=rowitem, column=8, value='존재함')

wb.save(G_ExFilePos + "고유번호변경_1021.xlsx")
wb.close()

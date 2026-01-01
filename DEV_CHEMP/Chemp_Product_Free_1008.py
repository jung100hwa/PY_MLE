# 수입요건면제(권혜림 연구원)
# 권혜림 연구원 수입요건 통계 뽑는 것

import datetime
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import random
import platform
import openpyxl as op
import cx_Oracle
import os

# 기술원 운영서버
os.putenv('NLS_LANG', '.UTF8')
connection = cx_Oracle.connect('chemp/chemp!1299@10.10.20.10:1521/chemp')
cursor  = connection.cursor()
cursor2 = connection.cursor()

# 현재 실행 위치 담기
G_ExFilePos = os.getcwd()

# 오늘날짜 세팅
now = datetime.datetime.now()
strTime = now.strftime('%Y-%m-%d')

# 플랫폼 담기
G_Platform = platform.system()

G_SplitDef = ''

# 작업항목마다 여기다 적는다.
if G_Platform == 'Windows':
    G_ExFilePos = G_ExFilePos + '\\All\\'
    G_SplitDef = '\\'
else:
    G_ExFilePos = G_ExFilePos + '/All/'
    G_SplitDef = '/'

wb = op.load_workbook(G_ExFilePos + "권혜림수입요건면제.xlsx")
ws = wb.worksheets[0]

cursor.execute("""SELECT
	TO_CHAR(REGT_DT ,'YYYY-MM-DD') AS "제출일",
	TO_CHAR(ACCEPT_DATE,'YYYY-MM-DD') AS "접수일",
	TA.ACCEPT_NO AS "접수번호",
	TA.STS_DATE AS "발급일",
	TA.EXEMP_NO AS "발급번호",
	(SELECT CODE_NM FROM COMTCCMMNDETAILCODE WHERE CODE_ID='IRN007' AND CODE=TA.PRO_STS) AS "상태",
	(SELECT CMPNY_NM FROM COMTNENTRPRSMBER WHERE ENTRPRS_MBER_ID=TA.ENTRPRS_MBER_ID) AS "회사명",
	(SELECT BIZRNO FROM COMTNENTRPRSMBER WHERE ENTRPRS_MBER_ID=TA.ENTRPRS_MBER_ID) AS "사업자등록번호",
	(SELECT CXFC FROM COMTNENTRPRSMBER WHERE ENTRPRS_MBER_ID=TA.ENTRPRS_MBER_ID) AS "대표자",
	TA.USE_PLACE AS "주소",
	(SELECT CB_FRST_TELNO || '-' || CB_MDDL_TELNO || '-' || CB_LAST_TELNO FROM COMTNENTRPRSMBER WHERE ENTRPRS_MBER_ID=TA.ENTRPRS_MBER_ID) AS "전화번호",
	(SELECT CB_FRST_FAX_TELNO || '-' || CB_MDDL_FAX_TELNO || '-' || CB_LAST_FAX_TELNO FROM COMTNENTRPRSMBER WHERE ENTRPRS_MBER_ID=TA.ENTRPRS_MBER_ID) AS "팩스번호",
	(SELECT CODE_NM FROM COMTCCMMNDETAILCODE WHERE CODE_ID='CPC003' AND CODE=TA.MAKE_COUNTRY) AS "제조국명",
	TA.MAKE_NAME AS "회사명",
	TA.PDT_NAME AS "제품명",
    GET_CODE_NM('NLC007', TA.PDT_CATEGORY) AS "품목분류",
    GET_CODE_NM(TA.PDT_CATEGORY, TA.PDT_ITEM) AS "품목",
    GET_CODE_NM(TA.PDT_ITEM,TA.USE_ITEM) AS "용도",
    GET_CODE_NM(TA.USE_ITEM,TA.USE_DETIAL_ITEM) AS "세부용도",
    (TA.PDT_EA || ' ' || TA.PDT_EA_TYPE) AS "수량",
    TA.PDT_AMOUNT AS "금액",
    TA.USE_PLACE AS "사용장소",
    (SELECT CODE_NM FROM COMTCCMMNDETAILCODE WHERE CODE_ID='IRN005' AND CODE=TA.EXEMP_CN) AS "면제내용",
    (SELECT CODE_NM FROM COMTCCMMNDETAILCODE WHERE CODE_ID='IRN006' AND CODE=TA.EXEMP_RESN) AS "면제확인 신청사용",
    CASE WHEN TA.PRD_LABEL_YN='Y' THEN '부착' ELSE '미부착' END AS "제품라벨",
    (SELECT CODE_NM FROM COMTCCMMNDETAILCODE WHERE CODE_ID='CPC003' AND CODE=TA.EXPORT_COUNTRY) AS "수출국명"
FROM TN_IRN_EXEMP TA WHERE TA.EXEMP_CN='01' AND STS_DATE LIKE '2021%'
    """)

rowcount = 3

for row in cursor:
    print(row)
    strListValue = []
    for col in range(0,len(cursor.description)):
        if col == 0:
            strNo=str(row[col])

        va = str(row[col])
        if '' in va:
            va = va.replace('', ' ')
        if '' in va:
            va = va.replace('', ' ')
        if '' in va:
            va = va.replace('', ' ')
        if va == 'None':
            va=''
        ws.cell(row=rowcount, column=col+1, value=va)

    rowcount = rowcount + 1

wb.save(G_ExFilePos + "권혜림수입요건면제.xlsx")
wb.close()
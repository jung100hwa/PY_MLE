# 선임자에 따른 승인유예고시 항목 조사
# 국외제조자 중심의 공동협의체 구성. 결국 앞으로는 이것만 사용
# 선임자만 별도로 엑셀을 만들어서 돌려야 함

import platform
import openpyxl as op
import cx_Oracle
import datetime
import os

# 기술원 개발서버
# os.putenv('NLS_LANG', '.UTF8')
# connection = cx_Oracle.connect('chemp/chemp@192.168.50.60:1521/orcl')
# cursor = connection.cursor()

# 기술원 운영서버
os.putenv('NLS_LANG', '.UTF8')
connection = cx_Oracle.connect('chemp/chemp!1299@10.10.20.10:1521/chemp')
cursor = connection.cursor()

# 로컬설버
# os.putenv('NLS_LANG', '.UTF8')
# connection = cx_Oracle.connect('scott/tiger@127.0.0.1:1521/XE')
# cursor = connection.cursor()

# 현재 실행 위치 담기
G_ExFilePos = os.getcwd()

# 오늘날짜 세팅
now = datetime.datetime.now()
strTime = now.strftime('%Y-%m-%d')

# 플랫폼 담기
G_Platform = platform.system()
G_SplitDef = ''

# 작업항목마다 여기다 적는다.
if G_Platform == 'Windows':
    G_ExFilePos = G_ExFilePos + '\\GODT\\'
    G_SplitDef = '\\'
else:
    G_ExFilePos = G_ExFilePos + '/GODT/'
    G_SplitDef = '/'

# 엑셀파일 열기
wb = op.load_workbook(G_ExFilePos + "선임고시_0930.xlsx")
ws = wb.worksheets[0]

strMstid = ''
strMainMstid = ''

for rowitem in range(2, ws.max_row+1):
    print("===========> " + str(rowitem))

    strNo = str(ws.cell(row=rowitem, column=1).value).strip()           # 연번
    strMaterial = str(ws.cell(row=rowitem, column=2).value).strip()     # 물질명
    strCasno = str(ws.cell(row=rowitem, column=3).value).strip()        # CAS No

    # 중간에 cas_no값 합쳐짐. 기존 살생물물질이 합쳐지고 붙고 함
    strCasno2 = ''
    if strCasno:
        strCasno = strCasno.strip()
        strCasno2 = strCasno
        if strCasno == '27083-27-8' or strCasno == '32289-58-0':
            strCasno = '27083-27-8, 32289-58-0'
        if strCasno == '105827-78-9' or strCasno == '138261-41-3':
            strCasno = '138261-41-3(105827-78-9)'
        # if strCasno == '고유번호 없음-6' or strCasno == '7681-52-9':
        #     strCasno = '7681-52-9'
        if strCasno == '고유번호 없음-7' or strCasno == '7778-54-3':
            strCasno = '7778-54-3'
    else:
        continue

    #  접수번호로 조회하면 선임일 경우 선임자 정보가 나옴, 선임자가 아닐 경우는 신고기업 정보
    strUsecode = str(ws.cell(row=rowitem, column=4).value).strip()          # 제품유형코드
    strGigan = str(ws.cell(row=rowitem, column=5).value).strip()            # 승인유예기간
    strRectno = str(ws.cell(row=rowitem, column=6).value).strip()           # 접수번호
    strComp1 = str(ws.cell(row=rowitem, column=7).value).strip()            # 신고기업
    strComp2 = str(ws.cell(row=rowitem, column=8).value).strip()            # 선임자
    strZezo = str(ws.cell(row=rowitem, column=9).value).strip()             # 국외제조자


    # 제품유형코드를 정제
    strUsecode = strUsecode.replace(' ', '')
    if strUsecode=='4-선박·수중시설용오염방지제':
        strUsecode = '4-가.선박·수중시설용오염방지제'

    # 날짜 끝에 . 없애기, 승인유예고시 엑셀에 꼭 날짜 마지막에 '.'를 찍어 줌. 이상하네.
    if strGigan:
        if len(strGigan) == 11:
            strGigan = strGigan[0:len(strGigan) - 1]
            strGigan = strGigan.replace('.', '-')

    # 접수번호를 형식에 맞게 수정
    strRectno = 'EAS-N-' + strRectno

    # 화학물질명에는 수만은 특수문자...아래는 ' 를 처리한다.
    strMaterial = strMaterial.replace("'", "' || chr(39) || '")

    # 유형코드를 조회
    cursor.execute("""SELECT CODE FROM COMTCCMMNDETAILCODE WHERE CODE_ID='BCC003' AND REPLACE(CODE_NM,' ','')='%s'""" % (strUsecode))
    ptcode = cursor.fetchone()
    if ptcode:
        strUsecode = str(ptcode[0])
    else:
        continue


    # 국외제조자도 특수문자 처리를 해야 함.
    strZezo = strZezo.replace("'", "' || chr(39) || '")

    # 여기서 접수번호에 해당하는 신청인 정보를 구한다. 만약에 선임자라고 하면 선임자 정보, 선임자가 아니면 일반 신청인 정보
    cursor.execute("""SELECT WR_ID FROM TN_EBMMST WHERE RCPTNO='%s'""" % (strRectno))
    strWrid = cursor.fetchone()
    strBizno = ''
    if strWrid:
        strWrid = str(strWrid[0])
        cursor.execute("""SELECT BIZRNO FROM COMTNENTRPRSMBER WHERE ENTRPRS_MBER_ID='%s'""" % (strWrid))
        strBizno = cursor.fetchone()
        if strBizno:
            strBizno = str(strBizno[0])
            strBizno = strBizno.replace('-', '')
        else:  # 사업자 번호를 기업회원에 없으면 일반회원에서 구한다.
            cursor.execute(
                """SELECT BIZRNO FROM COMTNENTRPRSMBER WHERE ENTRPRS_MBER_ID=(SELECT ENTRPRS_MBER_ID  FROM COMTNGNRLMBER WHERE MBER_ID='%s')"""
                % (strWrid))
            strBizno = cursor.fetchone()
            if strBizno:
                strBizno = str(strBizno[0])
                strBizno = strBizno.replace('-', '')
            else:
                continue  # 사업자번호가 없으면 중단
    else:  # 사용자 아이디가 없으면 중단
        continue


    # 선임자일경우 strComp2 = 'Y'
    strComp2 = strComp2.replace(" ", "")
    if strComp2 == "None" or strComp2 == '':
        strComp2 = 'N'
    else:
        cursor.execute("""SELECT MANAGE_NO FROM TN_APD_MST WHERE STS = '005'  AND (CAS_NO='%s' OR CAS_NO='%s') AND ITEM_CD='%s'
        AND BIZRNO='%s' AND APP_CN ='%s'""" % (strCasno, strCasno2, strUsecode, strBizno, strZezo))
        strMangeno = cursor.fetchone()
        if strMangeno:
            strComp2 = strMangeno[0]
        else:
            strComp2 = '선임자조회 안됨'

    # 제조자 정보
    # strZezo = strZezo.replace(" ", "")    # 절대하면 안된다.
    if strZezo == "None" or strZezo == '':
        strZezo = "-"

    # TN_EBM_COMPANY_LIST에 추가 한다. 먼저 있는지 검토한다.
    cursor.execute(
        """SELECT COUNT(*) FROM TN_EBM_COMPANY_LIST WHERE RCT_NO='%s' AND (CAS_NO='%s' OR CAS_NO='%s') AND ITEM_CD='%s' AND DEL_YN = 'N'"""
        % (strRectno, strCasno, strCasno2, strUsecode))
    strExist = cursor.fetchone()


    if strExist:
        strExist = int(strExist[0])

        # TN_EBM_COMPANY에 없으면 추가
        if strExist == 0:
            # 아래 부분은 통합DB를 만들려고 TN_EBM_COMPANY_LIST이 테이블에도 국외제조자를 넣어 줬던것인데 일단 현재는 사용하지 않음
            # cursor.execute("""INSERT INTO TN_EBM_COMPANY_LIST(RCT_NO, CAS_NO, ITEM_CD, MATERIAL_NAME, NO, DEL_YN, GRACE_DATE, AGENT_NAME)
            # VALUES('%s','%s','%s','%s', %s,'%s','%s', '%s')""" % (strRectno, strCasno, strUsecode, strMaterial, strNo, 'N', strGigan, strZezo))

            cursor.execute("""INSERT INTO TN_EBM_COMPANY_LIST(RCT_NO, CAS_NO, ITEM_CD, MATERIAL_NAME, NO, DEL_YN, GRACE_DATE)
            VALUES('%s','%s','%s','%s', %s,'%s','%s')""" % (strRectno, strCasno, strUsecode, strMaterial, strNo, 'N', strGigan))
            connection.commit()
        # 있으면 국외제조자만 업데이트
        # else:
        #     # 여기도 마찬가지 통합DB를 위해 컬럼을 추가 했는데. 현재는 사용하지 않음
        #     cursor.execute("""UPDATE TN_EBM_COMPANY_LIST SET AGENT_NAME ='%s' WHERE RCT_NO='%s' AND (CAS_NO='%s' OR CAS_NO='%s') AND ITEM_CD='%s' AND DEL_YN = 'N'"""
        #                    % (strZezo, strRectno, strCasno, strCasno2, strUsecode))
        #     connection.commit()

    # 제품유형이 있는지 확인한다.
    cursor.execute("""SELECT MST_ID FROM TN_EBM_COOPERATION_T WHERE CAS_NO='%s' AND USE_CODE='%s'""" % (strCasno, strUsecode))
    strMstid = cursor.fetchone()

    if strMstid:  # 제품유형이 존재하면
        strMstid = str(strMstid[0])

        # 업체가 제품유형협의체에 있는지 조사
        cursor.execute("""SELECT MST_ID FROM TN_EBM_COOPERATION_MAP_T WHERE MST_ID='%s' AND BIZ_NO='%s' AND (
        AGENT_NAME ='-' or AGENT_NAME='%s')""" % (strMstid, strBizno, strZezo))
        sss = cursor.fetchone()

        # 유형협의체에 기업이 존재하면 선임, '-'를 업데이트한다.
        if sss:
            cursor.execute("""UPDATE TN_EBM_COOPERATION_MAP_T SET AGENT_YN='%s', AGENT_NAME='%s' WHERE MST_ID='%s' AND BIZ_NO='%s' AND (
            AGENT_NAME ='-' or AGENT_NAME='%s')""" % (strComp2, strZezo, strMstid, strBizno, strZezo))
            connection.commit()
        else:
            #  업체가 제품유형협의체에 존재하지 않으면 제품유형에 업체를 추가한다.
            #  선임일 경우 agent_yn에 y로 표시한다.
            cursor.execute("""INSERT INTO TN_EBM_COOPERATION_MAP_T(MST_ID, BIZ_NO, GROUP_ROLE, AGREE_YN, USE_YN, INDIVISUAL_SUBMIT_YN,
            REQUEST_PLAN_YN, HOPE_REPRESENT_YN, CREATE_DATE, CREATE_ID, UPDATE_DATE, UPDATE_ID, APPROVE_YN, AGENT_YN, AGENT_NAME)
            VALUES('%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', to_date('%s','YYYY-MM-DD'), '%s', to_date('%s','YYYY-MM-DD'), '%s', '%s','%s', '%s')
            """ % (strMstid, strBizno, '009', 'Y', 'Y', 'N', 'N', '', strTime, strWrid, strTime, strWrid, 'Y', strComp2, strZezo))
            connection.commit()

        # 유형협의체의 참여자수를 증가시킨다. !! 지금은 소스가 변경되서 큰 의미는 없다
        cursor.execute("""UPDATE TN_EBM_COOPERATION_T SET PARTICIPANT_COUNT = PARTICIPANT_COUNT + 1 WHERE MST_ID='%s'""" % (strMstid))
        connection.commit()

    else:  # 제품 유형이 없을 때 신규로 만들어 넣는다.
        strSeq = """'TEC_' || LPAD(SEQ_EBM_COOPERATION_01.NEXTVAL,7,'0')"""
        cursor.execute("""INSERT INTO TN_EBM_COOPERATION_T(MST_ID, CAS_NO, MATERIAL_NM, USE_CODE, PARTICIPANT_COUNT, STS, CATEGORY,
        CREATE_DATE, CREATE_ID, UPDATE_DATE, UPDATE_ID) VALUES(%s,'%s','%s','%s','%s','%s','%s',to_date('%s','YYYY-MM-DD'),'%s',to_date('%s','YYYY-MM-DD'), '%s')"""
                       % (strSeq, strCasno, strMaterial, strUsecode, '0', '001', 'C', strTime, 'SYSTEM', strTime,'SYSTEM'))
        connection.commit()

        # 유형협의체 마스터 코드를 다시 불러온다. 새로 생성했기 때문에
        cursor.execute("""SELECT MST_ID FROM TN_EBM_COOPERATION_T WHERE CAS_NO='%s' AND USE_CODE='%s'""" % (strCasno, strUsecode))
        strMstid = cursor.fetchone()
        strMstid = str(strMstid[0])

        # 해당 제품유형코드에 업체정보를 추가한다.
        cursor.execute("""INSERT INTO TN_EBM_COOPERATION_MAP_T(MST_ID, BIZ_NO, GROUP_ROLE, AGREE_YN, USE_YN, INDIVISUAL_SUBMIT_YN,
        REQUEST_PLAN_YN, HOPE_REPRESENT_YN, CREATE_DATE, CREATE_ID, UPDATE_DATE, UPDATE_ID, APPROVE_YN, AGENT_YN, AGENT_NAME)
        VALUES('%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', to_date('%s','YYYY-MM-DD'), '%s', to_date('%s','YYYY-MM-DD'), '%s', '%s','%s', '%s')
        """ % (strMstid, strBizno, '009', 'Y', 'Y', 'N', 'N', '', strTime, strWrid, strTime, strWrid, 'Y', strComp2, strZezo))
        connection.commit()

        # 유형협의체의 참여자수를 증가시킨다. !! 지금은 소스가 변경되서 큰 의미는 없다
        cursor.execute("""UPDATE TN_EBM_COOPERATION_T SET PARTICIPANT_COUNT = PARTICIPANT_COUNT + 1 WHERE MST_ID='%s'""" % (strMstid))
        connection.commit()

    print('%s=== %s === %s===%s' % (rowitem, strMstid, strWrid, strBizno))

wb.save(G_ExFilePos + "선임고시_0930.xlsx")
wb.close()

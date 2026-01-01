# 조승연 책임. 데이터 추출
# 뽑아논 엑셀을 수정함. 살생물물질, 그외물질
import platform
from re import T
import openpyxl as op
import datetime
import os
from openpyxl.descriptors import base
import pandas as pd

# 현재 실행 위치 담기
G_ExFilePos = os.getcwd()

# 오늘날짜 세팅
now = datetime.datetime.now()
strTime = now.strftime("%Y-%m-%d")

# 플랫폼 담기
G_Platform = platform.system()

G_AptSample = ""  # 파일위치
G_SplitDef = ""

# 작업항목마다 여기다 적는다.
if G_Platform == "Windows":
    G_AptSample = G_ExFilePos + "\\ALL\\"
    G_SplitDef = "\\"
else:
    G_AptSample = G_ExFilePos + "/ALL/"
    G_SplitDef = "/"


m_list = [1, 2, 3, 4, 5, 6, 7, 8, 14, 15, 16, 17, 18, 19, 20, 21]

wb = op.load_workbook(G_AptSample + "살생물제품현황_20220117.xlsx")
ws = wb.worksheets[0]
basev = ''
currv = ''
chtv = True
for rowitem in range(2, ws.max_row + 1):

    print('=====> ' + str(rowitem))

    if chtv:
        for item in range(0, len(m_list)):      # 그룹중 첫번째 행
            imsi = str(ws.cell(row=rowitem, column=m_list[item]).value).strip()
            basev = basev + imsi
        chtv = False
    else:
        for item in range(0, len(m_list)):      # 두번째행부터
            imsi = str(ws.cell(row=rowitem, column=m_list[item]).value).strip()
            currv = currv + imsi

    if basev == currv:
        for colitem in range(1, ws.max_column + 1):
            if colitem in m_list:
                ws.cell(row=rowitem, column=colitem, value='')

    if basev != currv and len(currv) > 0:
        basev = currv

    currv = ''

wb.save(G_AptSample + "살생물제품현황_20220117_merge.xlsx")
wb.close()
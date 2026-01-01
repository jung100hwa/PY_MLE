# 오화용 연구원 초록누리 데이터 뽑는 

import datetime
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import random
import platform
import openpyxl as op
import cx_Oracle
import os

# 기술원 운영서버
os.putenv('NLS_LANG', '.UTF8')
connection = cx_Oracle.connect('chemp/chemp!1299@10.10.20.10:1521/chemp')
cursor  = connection.cursor()
cursor2 = connection.cursor()

# 현재 실행 위치 담기
G_ExFilePos = os.getcwd()

# 오늘날짜 세팅
now = datetime.datetime.now()
strTime = now.strftime('%Y-%m-%d')

# 플랫폼 담기
G_Platform = platform.system()

G_SplitDef = ''

# 작업항목마다 여기다 적는다.
if G_Platform == 'Windows':
    G_ExFilePos = G_ExFilePos + '\\All\\'
    G_SplitDef = '\\'
else:
    G_ExFilePos = G_ExFilePos + '/All/'
    G_SplitDef = '/'

wb = op.load_workbook(G_ExFilePos + "초록누리.xlsx")
ws = wb.worksheets[0]

cursor.execute("""
	SELECT ROW_NUMBER() OVER(ORDER BY ISSUDY, RCT_NO ),
	RCT_NO "접수번호",
	GET_CODE_NM('TES001',
	EXPR_INSPCT_OG_NM) AS "시험검사기관",
	CNS_NO AS "신고번호",
	EFFTPRD_BEGIN_DY AS "확인결과서 유효기간 시작일"
	--	  , decode(ST, 'represent', '대표', 'derive', '파생') AS "대표/파생 구분" 
,
	CASE
		WHEN ST = 'represent'
		AND cl = 'register' THEN '대표'
		WHEN st = 'represent'
		AND cl = 'change' THEN '변경(대표)'
		WHEN st = 'derive'
		AND cl = 'register' THEN '파생'
		WHEN st = 'derive'
		AND cl = 'change' THEN '변경(파생)'
		ELSE ''
	END "제품구분"
	--, decode(ST, 'represent', '대표', 'derive', '파생') AS "제품구분" 
,
	decode(ST, 'represent', '대표', 'derive', '파생') AS "대표/파생 구분" ,
	decode(CL, 'register', '신규', 'change', '변경') AS "신규/변경 분류" ,
	GET_CODE_NM('',
	ITEM_C) AS "품목군" ,
	GET_CODE_NM('',
	DETAIL_ITEM_C) AS "품목" ,
	PDT_NM AS "제품명" ,
	CASE
		ST WHEN 'derive' THEN PDTNM
		WHEN 'represent' THEN
		CASE
			CL WHEN 'change' THEN PDTNM
		END
	END AS "파생/변경제품명" ,
	decode(CLDPRTCPKGMTR_YN, 'yeschild', '대상', 'nochild', '비대상') AS "어린이보호포장대상" ,
	items AS "용도" ,
	CASE
		WHEN DETAIL_DF IS NULL THEN DF
		ELSE DF || '-' || DETAIL_DF
	END AS "제형" ,
	decode(MF_ICM, 'jejo', '제조', 'suib', '수입') AS "제조/수입" ,
	NVL2(MF_NATION, MF_NATION, '-') AS "제조국" ,
	CN_NM AS "기업명" ,
	APLCNT_ID ,
	(
	SELECT
		b.BIZRNO
	FROM
		(
		SELECT
			A.USER_ID,
			B.BIZRNO
		FROM
			COMVNUSERMASTER A,
			COMTNENTRPRSMBER B
		WHERE
			A.ASSGN_ID = B.ENTRPRS_MBER_ID ) b
	WHERE
		aplcnt_id = b.user_id ) "사업자번호" ,
	RCTDY AS "접수일자" ,
	ISSUDY AS "발급일자"
	--	  , 
	--	  ( 
	--	  	SELECT B.SBSCRB_DE 
	--	  	FROM 
	--	  	( 
	--	  	SELECT A.USER_ID, B.SBSCRB_DE 
	--	  	FROM COMVNUSERMASTER A, COMTNENTRPRSMBER B 
	--	  	WHERE A.ASSGN_ID  = B.ENTRPRS_MBER_ID 
	--	  	) b 
	--	  	WHERE aplcnt_id = b.user_id 
	--	  ) "기업관리자 등록일" 
	--	  , 
	--	  ( 
	--	  	SELECT GET_CODE_NM('COM026', B.ENTRPRS_SE_CODE) 
	--	  	FROM 
	--	  	( 
	--	  	SELECT A.USER_ID, B.ENTRPRS_SE_CODE 
	--	  	FROM COMVNUSERMASTER A, COMTNENTRPRSMBER B 
	--	  	WHERE A.ASSGN_ID  = B.ENTRPRS_MBER_ID 
	--	  	) b 
	--	  	WHERE aplcnt_id = b.user_id 
	--	  ) "기업구분코드" 
,
	bplc_addr "소재지주소" ,
	bplc_telno "소재지연락처" ,
	aplcnt_telno "담당자 연락처"
FROM
	(
	SELECT
		*
	FROM
		(
		SELECT
			a.MST_ID ,
			a.EST_NO ,
			a.RCT_NO ,
			a.CNS_NO ,
			a.ISSUDY,
			--TO_CHAR(a.MODDY, 'YYYY-MM-DD') AS MODDY , 
 a.CN_NM ,
			(
			SELECT
				PDTNM
			FROM
				TN_NLC_MST
			WHERE
				MST_ID = a.MST_ID
				AND EST_NO = '1' ) AS PDT_NM ,
			a.PDTNM ,
			a.MF_ICM ,
			a.RCTDY ,
			a.CLDPRTCPKGMTR_YN ,
			a.MF_NATION ,
			b.ITEM_C ,
			b.DETAIL_ITEM_C ,
			a.DF ,
			a.DETAIL_DF ,
			a.ST ,
			a.CL ,
			c.EXPR_INSPCT_OG_NM ,
			c.EFFTPRD_BEGIN_DY ,
			d.items ,
			a.APLCNT_ID ,
			a.BPLC_ADDR || BPLC_D_ADDR BPLC_ADDR ,
			a.BPLC_TELNO_ARNO || '-' || a.BPLC_TELNO_MIDNO || '-' || a.BPLC_TELNO_ENDNO bplc_telno ,
			a.aplcnt_telno_arno || '-' || a.aplcnt_telno_midno || '-' || a.aplcnt_telno_endno aplcnt_telno
		FROM
			TN_NLC_MST a,
			TN_NLC_ITEM b,
			TN_NLC_INSPCTSCRE c,
			(
			SELECT
				MST_ID,
				EST_NO,
				LISTAGG(ITEMS, ', ') WITHIN GROUP (
			ORDER BY
				ITEMS) AS ITEMS
			FROM
				(
				SELECT
					MST_ID,
					EST_NO,
					ITEM_C,
					DETAIL_ITEM_C,
					PRPOS_C,
					DETAIL_PRPOS_C ,
					CASE
						WHEN ITEM_C IS NOT NULL
							AND DETAIL_ITEM_C IS NOT NULL
							AND PRPOS_C IS NOT NULL
							AND DETAIL_PRPOS_C IS NOT NULL THEN GET_CODE_NM('',
							PRPOS_C) || '-' || GET_CODE_NM('',
							DETAIL_PRPOS_C) || DECODE(DETAIL_PROPOS_ETC, '', '', '(' || DETAIL_PROPOS_ETC || ')')
							WHEN ITEM_C IS NOT NULL
								AND DETAIL_ITEM_C IS NOT NULL
								AND PRPOS_C IS NOT NULL
								AND DETAIL_PRPOS_C IS NULL THEN GET_CODE_NM('',
								PRPOS_C) || DECODE(DETAIL_PROPOS_ETC, '', '', '(' || DETAIL_PROPOS_ETC || ')')
								WHEN ITEM_C IS NOT NULL
									AND DETAIL_ITEM_C IS NOT NULL
									AND PRPOS_C IS NULL
									AND DETAIL_PRPOS_C IS NULL THEN ''
									WHEN ITEM_C IS NOT NULL
									AND DETAIL_ITEM_C IS NULL
									AND PRPOS_C IS NULL
									AND DETAIL_PRPOS_C IS NULL THEN ''
									ELSE 'EXCEPTION : NO MORE OPTIONS!'
								END AS ITEMS
								--104207 

								FROM (
								--SELECT COUNT(*) 
								--FROM 
								--( 
								SELECT
									DISTINCT MST_ID,
									EST_NO,
									ITEM_C,
									DETAIL_ITEM_C,
									PRPOS_C,
									DETAIL_PRPOS_C,
									DETAIL_PROPOS_ETC
								FROM
									TN_NLC_ITEM
								GROUP BY
									MST_ID,
									EST_NO,
									ITEM_C,
									DETAIL_ITEM_C,
									PRPOS_C,
									DETAIL_PRPOS_C,
									DETAIL_PROPOS_ETC
									--) 
) )
			GROUP BY
				MST_ID,
				EST_NO ) d
		WHERE
			1 = 1
			AND a.DELETE_AT = 'N'
			AND a.MST_ID = b.MST_ID
			AND a.EST_NO = b.EST_NO
			AND a.MST_ID = c.MST_ID
			AND a.EST_NO = c.EST_NO
			AND b.ITEM_ID = c.ITEM_ID
			AND a.MST_ID = d.MST_ID
			AND a.EST_NO = d.EST_NO
			AND a.STS = 'ST008'
		ORDER BY
			a.MST_ID,
			a.EST_NO,
			b.ITEM_C,
			b.DETAIL_ITEM_C,
			b.PRPOS_C,
			b.DETAIL_PRPOS_C,
			c.INSPCTSCRE_ID,
			c.ITEM_ID )
	GROUP BY
		MST_ID ,
		EST_NO ,
		RCT_NO ,
		CNS_NO ,
		ISSUDY ,
		CN_NM ,
		PDT_NM ,
		PDTNM ,
		MF_ICM ,
		RCTDY ,
		CLDPRTCPKGMTR_YN ,
		MF_NATION ,
		ITEM_C ,
		DETAIL_ITEM_C ,
		DF ,
		DETAIL_DF ,
		ST ,
		CL ,
		EXPR_INSPCT_OG_NM ,
		EFFTPRD_BEGIN_DY ,
		items ,
		aplcnt_id ,
		BPLC_ADDR ,
		bplc_telno ,
		aplcnt_telno
		--HAVING MODDY <= '2021-01-31'
		HAVING ISSUDY BETWEEN '2019-01-01' AND '2021-12-31')
    """)

rowcount = 2

for row in cursor:
    print(row)
    strListValue = []
    for col in range(0,len(cursor.description)):
        if col == 0:
            strNo=str(row[col])

        va = str(row[col])
        if '' in va:
            va = va.replace('', ' ')
        if '' in va:
            va = va.replace('', ' ')
        if '' in va:
            va = va.replace('', ' ')
        if va == 'None':
            va=''
        ws.cell(row=rowcount, column=col+1, value=va)

    rowcount = rowcount + 1

wb.save(G_ExFilePos + "초록누리.xlsx")
wb.close()
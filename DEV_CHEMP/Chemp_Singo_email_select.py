# 조승현 책임 신고대상 신청자 이메일 정보 추출
# 엑셀 추출 이후에 중복데이터 삭제 필요


import datetime
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import random
import platform
import openpyxl as op
import cx_Oracle
import os

# 기술원 운영서버
os.putenv('NLS_LANG', '.UTF8')
connection = cx_Oracle.connect('chemp/chemp!1299@10.10.20.10:1521/chemp')
cursor  = connection.cursor()
cursor2 = connection.cursor()

# 현재 실행 위치 담기
G_ExFilePos = os.getcwd()

# 오늘날짜 세팅
now = datetime.datetime.now()
strTime = now.strftime('%Y-%m-%d')

# 플랫폼 담기
G_Platform = platform.system()

G_SplitDef = ''

# 작업항목마다 여기다 적는다.
if G_Platform == 'Windows':
    G_ExFilePos = G_ExFilePos + '\\All\\'
    G_SplitDef = '\\'
else:
    G_ExFilePos = G_ExFilePos + '/All/'
    G_SplitDef = '/'

wb = op.load_workbook(G_ExFilePos + "신고대상신청자이메일정보.xlsx")
ws = wb.worksheets[0]

rowcount = 2

cursor.execute("""SELECT CN_NM, APLCNT_NAME, APLCNT_EMAIL_ID || '@' || APLCNT_EMAIL_DOMN FROM TN_NLC_MST""")
for row in cursor:
	print(row)
	for col in range(0,len(cursor.description)):
		va = str(row[col])
		ws.cell(row=rowcount, column=col+1, value=va)
	rowcount = rowcount + 1

wb.save(G_ExFilePos + "신고대상신청자이메일정보.xlsx")
wb.close()